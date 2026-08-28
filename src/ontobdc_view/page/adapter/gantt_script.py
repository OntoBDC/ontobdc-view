from __future__ import annotations

from .container import (
    IFC_WORK_SCHEDULE_RUNTIME,
    chrome_controls_source,
    connection_state_source,
    container_connection_source,
)

from typing import Callable, Dict

from ontobdc_view.shared.adapter.vendor import (
    VENDOR_SHEET_JS_NAME,
    vendor_asset_source,
)
from ontobdc_view.shared.domain.port.gantt_script import GanttScriptPort


class GanttScriptAdapter(GanttScriptPort):
    """Returns the IfcWorkSchedule Gantt Page's runtime JS, split by responsibility.

    Each of the 4 scripts below is one state of
    `GanttScriptGenerationProcessState` (ontobdc-wip). The build-time
    Capability for that state calls `script_source(name)` to get the
    content, then writes it to
    `.__ontobdc__/asset/ifc_work_schedule_view/<name>.js` inside the container —
    this class only builds text, it never touches a filesystem itself.

    All 4 attach their exports onto a single shared
    `window.OntoBDCGanttViewRuntime` namespace object (functions
    directly, cross-cutting mutable state under `.state`), since none of
    these load as ES modules — matching the plain-`<script>`-tag loading
    every other Tile/Page asset in this project already uses. Each script
    logs a `console.log` line when it starts and finishes attaching its
    exports, so the browser console shows exactly which generation state a
    loaded page is running, at what point.
    """

    _BUILDERS: Dict[str, Callable[["GanttScriptAdapter"], str]] = {
        # The Page loads the vendored SheetJS build by name like any other
        # runtime file, so it resolves here like any other: a name the Page
        # lists but this adapter cannot produce is a runtime that goes
        # missing at load time with nothing failing at build time.
        VENDOR_SHEET_JS_NAME: lambda self: vendor_asset_source(
            VENDOR_SHEET_JS_NAME
        ),
        "i18n_apply": lambda self: self._i18n_apply_source(),
        "graph_reader": lambda self: self._graph_reader_source(),
        "container_connection": lambda self: container_connection_source(
            IFC_WORK_SCHEDULE_RUNTIME
        ),
        "connection_state": lambda self: connection_state_source(
            IFC_WORK_SCHEDULE_RUNTIME
        ),
        "chrome_controls": lambda self: chrome_controls_source(
            IFC_WORK_SCHEDULE_RUNTIME
        ),
        "pyodide_runtime": lambda self: self._pyodide_runtime_source(),
        "task_table_timeline": lambda self: self._task_table_timeline_source(),
        "dependency_arrows": lambda self: self._dependency_arrows_source(),
    }

    def script_source(self, name: str) -> str:
        builder = self._BUILDERS.get(name)
        if builder is None:
            raise ValueError(f"Unknown gantt script name: {name!r}")
        return builder(self)

    def _i18n_apply_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var rt = window.OntoBDCGanttViewRuntime = window.OntoBDCGanttViewRuntime || { state: {} };
  console.log("[gantt-view] state start: i18n_script_generated");

  rt.state = Object.assign(
    {
      graph: [],
      enrichedTasks: [],
      rawSequences: [],
      taskBarData: [],
      taskIndexByGlobalId: {},
    },
    rt.state || {},
  );

  rt.IBIM_NS = "https://infobim.org/ontology/ns#";
  rt.OBDC_NS = "http://ontobdc.org/ontology/domain/ontobdc/ns.ttl#";
  rt.DCTERMS_TITLE = "http://purl.org/dc/terms/title";
  rt.DCTERMS_IDENTIFIER = "http://purl.org/dc/terms/identifier";
  rt.DCTERMS_DESCRIPTION = "http://purl.org/dc/terms/description";

  var I18N = JSON.parse(document.getElementById("ontobdc-i18n")?.textContent || "{}");

  function t(key, vars) {
    var locale = document.documentElement.lang || document.documentElement.dataset.language || "en";
    var table = I18N[locale] || I18N.en || {};
    var text = table[key] ?? key;
    if (vars) {
      for (var pair of Object.entries(vars)) text = text.replaceAll("{" + pair[0] + "}", pair[1]);
    }
    return text;
  }

  rt.t = t;

  function applyI18n() {
    var nodes = document.querySelectorAll("[data-i18n]");
    for (var i = 0; i < nodes.length; i++) {
      var n = nodes[i];
      var k = n.getAttribute("data-i18n");
      if (!k) continue;
      if (n.classList && n.classList.contains("connect-btn")) {
        if (rt.ensureConnectButtonInnerStatus) rt.ensureConnectButtonInnerStatus();
        if (rt.setConnectButtonLabel) {
          rt.setConnectButtonLabel(t(k));
        } else {
          n.textContent = t(k);
        }
      } else {
        n.textContent = t(k);
      }
    }
    var titleNodes = document.querySelectorAll("[data-i18n-title]");
    for (var j = 0; j < titleNodes.length; j++) {
      var tn = titleNodes[j];
      var kt = tn.getAttribute("data-i18n-title");
      if (kt) tn.setAttribute("title", t(kt));
    }
    var labelNodes = document.querySelectorAll("[data-i18n-aria-label]");
    for (var m = 0; m < labelNodes.length; m++) {
      var ln = labelNodes[m];
      var kl = ln.getAttribute("data-i18n-aria-label");
      if (kl) ln.setAttribute("aria-label", t(kl));
    }
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", applyI18n);
  } else {
    applyI18n();
  }

  console.log("[gantt-view] state done: i18n_script_generated");
})(window);
"""

    def _graph_reader_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var rt = window.OntoBDCGanttViewRuntime = window.OntoBDCGanttViewRuntime || { state: {} };
  console.log("[gantt-view] state start: graph_reader_script_generated");

  // Must match `.gantt-table tbody tr { height: 34px; }` in
  // ifc_work_schedule_view.css — this drives the SVG bar Y offsets on the
  // right, the CSS drives the actual <tr> height on the left; any gap
  // between the two compounds per row and drifts the bars away from their
  // WBS rows the further down the list you scroll.
  rt.ROW_HEIGHT = 34;
  rt.HEADER_HEIGHT = 72;
  rt.HEADER_ROW_HEIGHT_YEAR = 24;
  rt.HEADER_ROW_HEIGHT_MONTH = 24;
  rt.HEADER_ROW_HEIGHT_DAY = 24;
  rt.TICK_WIDTH_DAY = 40;
  rt.MILLIS_PER_DAY = 86400000;

  rt.IFC_WORK_SCHEDULE_TYPE = "https://infobim.org/ontology/ns#IfcWorkSchedule";
  rt.IFC_TASK_TYPES = [
    "https://standards.buildingsmart.org/IFC/DEV/IFC4_3/OWL#IfcTask",
    "https://infobim.org/ontology/ns#IfcTask",
    "http://standards.buildingsmart.org/IFC/DEV/IFC4_3/OWL#IfcTask",
    "IfcTask"
  ];
  rt.IFC_TASK_TIME_TYPES = [
    "https://standards.buildingsmart.org/IFC/DEV/IFC4_3/OWL#IfcTaskTime",
    "https://infobim.org/ontology/ns#IfcTaskTime",
    "http://standards.buildingsmart.org/IFC/DEV/IFC4_3/OWL#IfcTaskTime",
    "IfcTaskTime"
  ];
  rt.IFC_REL_SEQUENCE_TYPES = [
    "https://standards.buildingsmart.org/IFC/DEV/IFC4_3/OWL#IfcRelSequence",
    "https://infobim.org/ontology/ns#IfcRelSequence",
    "http://standards.buildingsmart.org/IFC/DEV/IFC4_3/OWL#IfcRelSequence",
    "IfcRelSequence"
  ];

  rt.DC_ALIASES = {
    "Name": ["http://purl.org/dc/terms/title", "schema:name", "rdfs:label"],
    "Description": ["http://purl.org/dc/terms/description", "schema:description"],
    "Identification": ["http://purl.org/dc/terms/identifier", "schema:identifier"],
    "Title": ["http://purl.org/dc/terms/title", "schema:name"],
    "Identifier": ["http://purl.org/dc/terms/identifier", "schema:identifier"]
  };

  function parseJsonLd() {
    var el = document.getElementById("ontobdc-page-jsonld");
    if (!el) return [];
    try {
      var parsed = JSON.parse(el.textContent);
      if (Array.isArray(parsed)) return parsed;
      if (parsed && parsed["@graph"] && Array.isArray(parsed["@graph"])) {
        return parsed["@graph"];
      }
      return [parsed];
    } catch (e) {
      return [];
    }
  }
  rt.parseJsonLd = parseJsonLd;

  function typeMatches(node, typeUris) {
    if (!node) return false;
    var nodeTypes = node["@type"];
    if (!nodeTypes) return false;
    var arr = Array.isArray(nodeTypes) ? nodeTypes : [nodeTypes];
    for (var i = 0; i < arr.length; i++) {
      var t = arr[i];
      if (typeof t === "string") {
        if (typeUris.includes(t)) return true;
        for (var j = 0; j < typeUris.length; j++) {
          var tu = typeUris[j];
          if (t.endsWith("#" + tu) || t.endsWith("/" + tu)) return true;
        }
      }
    }
    return false;
  }
  rt.typeMatches = typeMatches;

  function findByType(graph, typeUris) {
    var list = Array.isArray(typeUris) ? typeUris : [typeUris];
    var out = [];
    for (var i = 0; i < graph.length; i++) {
      if (typeMatches(graph[i], list)) out.push(graph[i]);
    }
    return out;
  }
  rt.findByType = findByType;

  function extractLiteral(val) {
    if (val === null || val === undefined) return "";
    if (typeof val === "string") return val;
    if (typeof val === "number" || typeof val === "boolean") return String(val);
    if (Array.isArray(val)) {
      for (var i = 0; i < val.length; i++) {
        var r = extractLiteral(val[i]);
        if (r !== "") return r;
      }
      return "";
    }
    if (typeof val === "object") {
      if ("@value" in val) return extractLiteral(val["@value"]);
      if ("value" in val) return extractLiteral(val["value"]);
    }
    return "";
  }
  rt.extractLiteral = extractLiteral;

  function literalValue(node, identifier) {
    if (!node || !identifier) return "";
    var idLower = identifier.toLowerCase();
    var idCamel = identifier.charAt(0).toUpperCase() + identifier.slice(1);
    var idCamelLower = idCamel.toLowerCase();
    var exactPatterns = [
      identifier,
      "#" + identifier,
      "/" + identifier,
      identifier.toLowerCase(),
      "#" + identifier.toLowerCase(),
      "/" + identifier.toLowerCase(),
      idCamel,
      "#" + idCamel,
      "/" + idCamel,
      identifier.charAt(0).toLowerCase() + identifier.slice(1),
      identifier + "Field",
      "#" + identifier + "Field",
      "/" + identifier + "Field",
      idCamel + "Field",
      "#" + idCamel + "Field",
      "/" + idCamel + "Field"
    ];
    var keys = Object.keys(node);
    for (var i = 0; i < keys.length; i++) {
      var key = keys[i];
      for (var j = 0; j < exactPatterns.length; j++) {
        var p = exactPatterns[j];
        if (key === p || key.endsWith(p)) {
          var val = node[key];
          return extractLiteral(val);
        }
      }
    }
    for (var k = 0; k < keys.length; k++) {
      var lowerKey = keys[k].toLowerCase();
      var stripped = lowerKey.replace(/^.*#/, "").replace(/^.*\//, "").replace(/field$/, "");
      if (stripped === idLower || stripped === idCamelLower) {
        var val2 = node[keys[k]];
        var r = extractLiteral(val2);
        if (r !== "") return r;
      }
    }
    for (var m = 0; m < keys.length; m++) {
      var lowerKey2 = keys[m].toLowerCase();
      var stripped2 = lowerKey2.replace(/^.*#/, "").replace(/^.*\//, "");
      if (stripped2.includes(idLower) || stripped2.includes(idCamelLower)) {
        var val3 = node[keys[m]];
        var result = extractLiteral(val3);
        if (result !== "") return result;
      }
    }
    return "";
  }
  rt.literalValue = literalValue;

  function anyLiteral(node, identifiers) {
    if (!node) return "";
    var list = Array.isArray(identifiers) ? identifiers : [identifiers];
    for (var i = 0; i < list.length; i++) {
      var v = literalValue(node, list[i]);
      if (v !== "") return v;
    }
    return "";
  }
  rt.anyLiteral = anyLiteral;

  function aliasAwareLiteral(node, identifier) {
    var direct = literalValue(node, identifier);
    if (direct !== "") return direct;
    var aliases = rt.DC_ALIASES[identifier] || rt.DC_ALIASES[identifier.charAt(0).toUpperCase() + identifier.slice(1)] || [];
    for (var i = 0; i < aliases.length; i++) {
      var alias = aliases[i];
      var keys = Object.keys(node || {});
      for (var j = 0; j < keys.length; j++) {
        var key = keys[j];
        var aliasShort = alias.split(":").pop();
        if (key === alias || key.endsWith("#" + aliasShort) || key.endsWith("/" + aliasShort) || key.includes(alias)) {
          var r = extractLiteral(node[key]);
          if (r !== "") return r;
        }
      }
    }
    return "";
  }
  rt.aliasAwareLiteral = aliasAwareLiteral;

  function referenceId(node) {
    if (!node) return "";
    if (typeof node === "string") return node;
    if (Array.isArray(node)) {
      for (var i = 0; i < node.length; i++) {
        var r = referenceId(node[i]);
        if (r !== "") return r;
      }
      return "";
    }
    if (typeof node === "object") {
      if ("@id" in node) return node["@id"];
      if ("id" in node) return String(node["id"]);
    }
    return "";
  }
  rt.referenceId = referenceId;

  function parseDate(str) {
    if (!str) return null;
    var s = String(str).trim();
    if (!s) return null;
    var d = new Date(s);
    if (!isNaN(d.getTime())) return d;
    var match = s.match(/^(\d{4})-(\d{2})-(\d{2})/);
    if (match) {
      d = new Date(Number(match[1]), Number(match[2]) - 1, Number(match[3]));
      if (!isNaN(d.getTime())) return d;
    }
    return null;
  }
  rt.parseDate = parseDate;

  function formatDate(d) {
    if (!d) return "";
    var y = d.getFullYear();
    var m = String(d.getMonth() + 1).padStart(2, "0");
    var day = String(d.getDate()).padStart(2, "0");
    return y + "-" + m + "-" + day;
  }
  rt.formatDate = formatDate;

  function diffDays(a, b) {
    if (!a || !b) return 0;
    var aDay = new Date(a.getFullYear(), a.getMonth(), a.getDate()).getTime();
    var bDay = new Date(b.getFullYear(), b.getMonth(), b.getDate()).getTime();
    return Math.round((bDay - aDay) / rt.MILLIS_PER_DAY);
  }
  rt.diffDays = diffDays;

  function getGlobalId(node) {
    return literalValue(node, "GlobalId") || literalValue(node, "globalId") || "";
  }
  rt.getGlobalId = getGlobalId;

  function runParse() {
    var graph = parseJsonLd();
    rt.state.graph = graph;

    var schedules = findByType(graph, rt.IFC_WORK_SCHEDULE_TYPE);
    var schedule = schedules[0] || null;
    // The embedded graph and the connected workbook each contribute an
    // IfcWorkSchedule node under a different @id, so they never dedupe. The
    // embedded one carries dcterms:title/identifier; the workbook one carries
    // the live GlobalId and the editable Name/Description. Fold just those
    // onto the primary node so getGlobalId() resolves and an inline edit (or a
    // reconnect after one) shows the saved value instead of snapping back to
    // the baked-in title.
    if (schedule && schedules.length > 1) {
      var LIVE_SCHEDULE_KEYS = [
        rt.IBIM_NS + "GlobalId",
        rt.IBIM_NS + "Name",
        rt.IBIM_NS + "Description",
      ];
      for (var si = 1; si < schedules.length; si++) {
        for (var ki = 0; ki < LIVE_SCHEDULE_KEYS.length; ki++) {
          var liveKey = LIVE_SCHEDULE_KEYS[ki];
          if (schedules[si] && schedules[si][liveKey] !== undefined) {
            schedule[liveKey] = schedules[si][liveKey];
          }
        }
      }
    }
    var rawTasks = findByType(graph, rt.IFC_TASK_TYPES);
    var rawTaskTimes = findByType(graph, rt.IFC_TASK_TIME_TYPES);
    var rawSequences = findByType(graph, rt.IFC_REL_SEQUENCE_TYPES);
    rt.state.schedule = schedule;
    rt.state.rawTasks = rawTasks;
    rt.state.rawTaskTimes = rawTaskTimes;
    rt.state.rawSequences = rawSequences;

    var tasksByGlobalId = {};
    for (var i = 0; i < rawTasks.length; i++) {
      var t = rawTasks[i];
      var gid = getGlobalId(t);
      if (gid) tasksByGlobalId[gid] = t;
    }

    var taskTimeByGlobalId = {};
    for (var j = 0; j < rawTaskTimes.length; j++) {
      var tt = rawTaskTimes[j];
      var gid2 = getGlobalId(tt);
      if (gid2) taskTimeByGlobalId[gid2] = tt;
    }

    var taskByTaskTimeGid = {};
    for (var k = 0; k < rawTasks.length; k++) {
      var t2 = rawTasks[k];
      var taskTimeRef = anyLiteral(t2, ["TaskTimeGlobalId", "TaskTime"]) || referenceId(t2["TaskTime"] || t2["taskTime"] || t2["IfcTaskTime"] || null);
      var timeGidMatch = taskTimeRef.match(/([A-Za-z0-9_]{5,})/);
      var timeGid = timeGidMatch ? timeGidMatch[1] : "";
      if (!timeGid) {
        var ttGidKeys = Object.keys(taskTimeByGlobalId);
        for (var m = 0; m < ttGidKeys.length; m++) {
          if (taskTimeRef.includes(ttGidKeys[m])) {
            timeGid = ttGidKeys[m];
            break;
          }
        }
      }
      if (timeGid) taskByTaskTimeGid[timeGid] = t2;
    }

    var tasksByTimeRef = {};
    for (var n = 0; n < rawTasks.length; n++) {
      var t3 = rawTasks[n];
      var candidateKeys = ["TaskTime", "taskTime", "IfcTaskTime"];
      for (var p = 0; p < candidateKeys.length; p++) {
        var refs = t3[candidateKeys[p]];
        if (!refs) continue;
        var refsList = Array.isArray(refs) ? refs : [refs];
        for (var q = 0; q < refsList.length; q++) {
          var refId = referenceId(refsList[q]);
          if (refId) {
            tasksByTimeRef[refId] = t3;
            var ttGidKeys2 = Object.keys(taskTimeByGlobalId);
            for (var r = 0; r < ttGidKeys2.length; r++) {
              if (refId.includes(ttGidKeys2[r]) || refId.endsWith(ttGidKeys2[r])) {
                taskByTaskTimeGid[ttGidKeys2[r]] = t3;
              }
            }
          }
        }
      }
    }

    for (var s = 0; s < rawTaskTimes.length; s++) {
      var tt2 = rawTaskTimes[s];
      var ttId = tt2["@id"] || "";
      if (ttId && tasksByTimeRef[ttId]) {
        var t4 = tasksByTimeRef[ttId];
        var gid3 = getGlobalId(tt2);
        if (gid3 && !taskByTaskTimeGid[gid3]) {
          taskByTaskTimeGid[gid3] = t4;
        }
      }
    }

    var enrichedTasks = [];
    for (var w = 0; w < rawTasks.length; w++) {
      var task = rawTasks[w];
      var globalId = getGlobalId(task);
      var identification = aliasAwareLiteral(task, "Identification") || aliasAwareLiteral(task, "Identifier") || "";
      var name = aliasAwareLiteral(task, "Name") || aliasAwareLiteral(task, "Title") || "";
      var description = aliasAwareLiteral(task, "Description") || "";

      var taskTime = null;
      var ck = ["TaskTime", "taskTime", "IfcTaskTime", "TaskTimeGlobalId"];
      outerLoop:
      for (var x = 0; x < ck.length; x++) {
        var refs2 = task[ck[x]];
        if (!refs2) continue;
        var refsList2 = Array.isArray(refs2) ? refs2 : [refs2];
        for (var y = 0; y < refsList2.length; y++) {
          var refId2 = referenceId(refsList2[y]) || extractLiteral(refsList2[y]);
          for (var z = 0; z < rawTaskTimes.length; z++) {
            var tt3 = rawTaskTimes[z];
            var ttId2 = tt3["@id"] || "";
            var ttGid = getGlobalId(tt3);
            if ((refId2 && (ttId2 === refId2 || (ttGid && refId2.includes(ttGid)))) ||
                (ttGid && taskByTaskTimeGid[ttGid] === task)) {
              taskTime = tt3;
              break outerLoop;
            }
          }
        }
        if (taskTime) break;
      }
      if (!taskTime) {
        var ttGidMatch2 = Object.keys(taskByTaskTimeGid).find(function (g) { return taskByTaskTimeGid[g] === task; });
        if (ttGidMatch2) taskTime = taskTimeByGlobalId[ttGidMatch2] || null;
      }
      if (!taskTime && rawTasks.length > 0 && rawTaskTimes.length === rawTasks.length) {
        var linkedCount = 0;
        for (var yk = 0; yk < enrichedTasks.length; yk++) {
          if (enrichedTasks[yk] && enrichedTasks[yk].rawTime) linkedCount++;
        }
        if (linkedCount / rawTasks.length < 0.3) {
          taskTime = rawTaskTimes[w] || null;
        }
      }

      var scheduleStart = taskTime ? parseDate(literalValue(taskTime, "ScheduleStart")) : null;
      var scheduleFinish = taskTime ? parseDate(literalValue(taskTime, "ScheduleFinish")) : null;
      var actualStart = taskTime ? parseDate(literalValue(taskTime, "ActualStart")) : null;
      var actualFinish = taskTime ? parseDate(literalValue(taskTime, "ActualFinish")) : null;
      var completionStr = taskTime ? literalValue(taskTime, "Completion") : "";
      if (completionStr === "") completionStr = taskTime ? literalValue(taskTime, "PercentComplete") : "";
      var completion = 0;
      if (completionStr !== "") {
        var num = parseFloat(String(completionStr).replace("%", ""));
        if (!isNaN(num)) {
          if (num > 0 && num < 1) num = num * 100;
          completion = Math.max(0, Math.min(100, num));
        }
      }
      var isMilestoneStrTask = anyLiteral(task, ["IsMilestone", "Milestone"]);
      var isMilestoneStrTT = taskTime ? (anyLiteral(taskTime, ["IsMilestone", "Milestone"]) || "") : "";
      var isMilestoneStr = isMilestoneStrTask !== "" ? isMilestoneStrTask : isMilestoneStrTT;
      var isMilestoneLower = String(isMilestoneStr || "").trim().toLowerCase();
      var isMilestone = (
        isMilestoneLower === "true"
        || isMilestoneLower === "1"
        || isMilestoneLower === "yes"
        || isMilestoneLower === "sim"
        || isMilestoneLower === "s"
        || isMilestoneLower === "t"
      );
      var isCriticalStrTask = anyLiteral(task, ["IsCritical", "Critical"]);
      var isCriticalStrTT = taskTime ? (anyLiteral(taskTime, ["IsCritical", "Critical"]) || "") : "";
      var isCriticalStr = isCriticalStrTask !== "" ? isCriticalStrTask : isCriticalStrTT;
      var isCriticalLower = String(isCriticalStr || "").trim().toLowerCase();
      var isCritical = (
        isCriticalLower === "true"
        || isCriticalLower === "1"
        || isCriticalLower === "yes"
        || isCriticalLower === "sim"
        || isCriticalLower === "s"
        || isCriticalLower === "t"
      );

      var duration_days = 0;
      if (scheduleStart && scheduleFinish) {
        duration_days = diffDays(scheduleStart, scheduleFinish);
      }
      if (duration_days < 0) duration_days = 0;
      if (duration_days < 1 && isMilestone) {
        isMilestone = true;
      }
      if (duration_days === 0 && scheduleStart) {
        duration_days = 1;
      }

      enrichedTasks.push({
        globalId: globalId,
        identification: identification,
        name: name,
        description: description,
        scheduleStart: scheduleStart,
        scheduleFinish: scheduleFinish,
        actualStart: actualStart,
        actualFinish: actualFinish,
        completion: completion,
        isMilestone: isMilestone,
        isCritical: isCritical,
        duration_days: duration_days,
        raw: task,
        rawTime: taskTime
      });
    }

    enrichedTasks.sort(function (a, b) {
      var ak = (a.identification || a.name || "").toString();
      var bk = (b.identification || b.name || "").toString();
      return ak.localeCompare(bk, undefined, { numeric: true, sensitivity: "base" });
    });

    rt.state.enrichedTasks = enrichedTasks;
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", runParse);
  } else {
    runParse();
  }

  // Exported so the Page can re-derive everything after its graph changes —
  // connecting a container folder or refreshing from the workbook replaces
  // the embedded graph, and without this the new nodes would sit in the DOM
  // unread.
  rt.runParse = runParse;

  console.log("[gantt-view] state done: graph_reader_script_generated");
})(window);
"""

    def _pyodide_runtime_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCGanttViewRuntime = window.OntoBDCGanttViewRuntime || { state: {} };
  console.log("[ifc-work-schedule-view] state start: pyodide_runtime_script_generated");

  var __ontobdcIdentityT = function __ontobdcIdentityT(key, vars) {
    var text = String(key || "");
    if (vars) {
      try {
        Object.keys(vars).forEach(function (k) {
          text = text.split("{" + k + "}").join(String(vars[k]));
        });
      } catch (e) { /* noop */ }
    }
    return text;
  };
  if (typeof runtime.t !== "function") runtime.t = __ontobdcIdentityT;
  // Resolve runtime.t at call time, not now: this IIFE can run before
  // i18n_apply.js has attached the real translator (the loader inserts every
  // script at once), so a captured reference would freeze to the identity
  // fallback and every error message would surface as its raw i18n key.
  var t = function t(key, vars) {
    var fn = (typeof runtime.t === "function") ? runtime.t : __ontobdcIdentityT;
    return fn(key, vars);
  };

  // Normalize payload names: VIEW_PAYLOAD / WORKSTREAM_PAYLOAD / window.infoBimIfcWorkScheduleView
  // into one runtime.VIEW_PAYLOAD. Connection state may set runtime.WORKSTREAM_PAYLOAD after
  // this IIFE runs, and the HTML inline may expose window.infoBimIfcWorkScheduleView.
  if (!runtime.VIEW_PAYLOAD) {
    runtime.VIEW_PAYLOAD = runtime.WORKSTREAM_PAYLOAD || window.infoBimIfcWorkScheduleView || null;
  }
  var VIEW_PAYLOAD = runtime.VIEW_PAYLOAD;

  // Lazy wrapper for resolveContainerHandle so we don't crash if
  // container_connection.js hasn't defined runtime.resolveContainerHandle yet.
  // container_connection's resolveContainerHandle(selectedHandle) throws
  // "not this schedule's dataset" when called with no handle (its BFS starts
  // from `undefined`). Every standalone refresh here — the header Refresh
  // button, and now createTask()/deleteTask() via refreshFromWorkbook() —
  // calls it with no argument, so feed it the handle the connection already
  // resolved instead of letting it re-discover from nothing.
  var resolveContainerHandle = function __ontobdcLazyResolveContainerHandle(selectedHandle) {
    var fn = (typeof runtime.resolveContainerHandle === "function") ? runtime.resolveContainerHandle : null;
    if (!fn) {
      var err = new Error("No connected container available");
      return Promise.reject(err);
    }
    var handleArg = selectedHandle
      || (runtime.state && (runtime.state.datasetHandle || runtime.state.rawContainerHandle))
      || undefined;
    try {
      var result = fn.call(runtime, handleArg);
      return Promise.resolve(result);
    } catch (caughtErr) {
      return Promise.reject(caughtErr);
    }
  };

  function resolvePyodideLocalUrl() {
    var fallbackCdn = "https://cdn.jsdelivr.net/pyodide/v0.27.2/full/pyodide.js";
    var pyodideRel = "ifc_work_schedule_view/pyodide/v0.27.2/full/pyodide.js";
    try {
      var loaderBase = window.__ONTOBDC_ASSET_BASE_URL__ || (window.__ontobdcGanttLoader && window.__ontobdcGanttLoader.assetBase);
      var base = loaderBase;
      if (!base) {
        var tmpBase = document.createElement("a");
        tmpBase.href = "../asset/";
        base = tmpBase.href;
      }
      if (base && base.charAt(base.length - 1) !== "/") base += "/";
      var tmp = document.createElement("a");
      tmp.href = base + pyodideRel;
      var jsUrl = tmp.href;
      if (jsUrl && /\/pyodide\.js$/.test(jsUrl)) {
        var idx = jsUrl.lastIndexOf("/pyodide.js");
        return {
          scriptUrl: jsUrl,
          indexUrl: jsUrl.substring(0, idx + 1)
        };
      }
    } catch (e) { /* noop */ }
    return { scriptUrl: fallbackCdn, indexUrl: "https://cdn.jsdelivr.net/pyodide/v0.27.2/full/" };
  }
  var PYODIDE_URLS = resolvePyodideLocalUrl();
  const PYODIDE_CDN_URL = PYODIDE_URLS.scriptUrl;
  const PYODIDE_INDEX_URL = PYODIDE_URLS.indexUrl;

  let pyodideInstance = null;
  let pyodideLoadPromise = null;
  let pyodideHasOpenpyxl = false;

  // pyodide.globals is ONE shared namespace across every caller, so a
  // globals.set() from a second call can land between the first call's
  // globals.set() and its runPythonAsync() actually executing — the first
  // script then silently reads the second call's globals. Every
  // globals.set()+runPythonAsync() pair runs under this lock, exactly as
  // the WorkStream Page's runtime does.
  let pyodideOperationQueue = Promise.resolve();

  function withPyodideLock(taskFn) {
    const settled = pyodideOperationQueue.then(taskFn, taskFn);
    pyodideOperationQueue = settled.then(() => {}, () => {});
    return settled;
  }

  function loadScriptTag(src) {
    return new Promise((resolve, reject) => {
      const script = document.createElement("script");
      script.src = src;
      script.onload = resolve;
      script.onerror = () => reject(new Error(`Failed to load ${src}`));
      document.head.appendChild(script);
    });
  }

  async function ensurePyodide(options) {
    const withOpenpyxl = Boolean(options && options.withOpenpyxl);
    if (pyodideInstance && (!withOpenpyxl || pyodideHasOpenpyxl)) {
      return pyodideInstance;
    }
    if (!pyodideLoadPromise) {
      pyodideLoadPromise = (async () => {
        const isFileProtocol =
          typeof location !== "undefined" &&
          location &&
          typeof location.protocol === "string" &&
          location.protocol === "file:";
        if (typeof loadPyodide !== "function") {
          if (isFileProtocol) {
            throw new Error(
              "Pyodide unavailable on file:// protocol. If SheetJS " +
              "workbook parsing also fails, open this page via HTTP " +
              "(python -m http.server 8765) instead."
            );
          }
          await loadScriptTag(PYODIDE_CDN_URL);
        }
        const instance = await loadPyodide({ indexURL: PYODIDE_INDEX_URL });
        await instance.loadPackage("micropip");
        await instance.runPythonAsync("import micropip\nawait micropip.install('rdflib')");
        return instance;
      })();
    }
    pyodideInstance = await pyodideLoadPromise;
    if (withOpenpyxl && !pyodideHasOpenpyxl) {
      await pyodideInstance.runPythonAsync("import micropip\nawait micropip.install('openpyxl')");
      pyodideHasOpenpyxl = true;
    }
    return pyodideInstance;
  }

  // ---------------------------------------------------------------------------
  // SheetJS-native workbook parser. Equivalent to WORKBOOK_PARSE_SCRIPT but
  // runs entirely in the browser and therefore needs NO fetch(), no dynamic
  // imports, no WASM boot. Works on file:// because <script src=> loads it
  // as a classic script tag. Output shape is identical:
  //   { nodes: [{ @id, @type, [IBIM+column]: [{@value}] }],
  //     workbookPath: "relative/path/to.xlsx",
  //     counts: { IfcWorkSchedule, IfcTask, IfcTaskTime, IfcRelSequence } }
  // Pyodide is retained as a fallback path for deployments where SheetJS is
  // not bundled or for workbooks that openpyxl can read but SheetJS cannot.
  // ---------------------------------------------------------------------------
  const IBIM = "https://infobim.org/ontology/ns#";
  const GANTT_ENTITIES = ["IfcWorkSchedule", "IfcTask", "IfcTaskTime", "IfcRelSequence"];
  const WORKBOOK_EXT_RE = /\.xlsx$/i;

  function hasSheetJS() {
    return !!(window.XLSX && typeof window.XLSX.read === "function");
  }

  async function collectFilesRecursively(dirHandle) {
    var collected = [];
    async function walk(current, rel) {
      if (!current || typeof current.values !== "function") return;
      try {
        for await (var entry of current.values()) {
          if (!entry) continue;
          var nextRel = rel ? rel + "/" + entry.name : entry.name;
          if (entry.kind === "file") {
            collected.push({ handle: entry, relPath: nextRel });
          } else if (entry.kind === "directory") {
            try { await walk(entry, nextRel); } catch (subErr) {}
          }
        }
      } catch (err) {}
    }
    await walk(dirHandle, "");
    return collected;
  }

  function readFileAsArrayBuffer(fileHandle) {
    return fileHandle.getFile().then(function (file) {
      return file.arrayBuffer();
    });
  }

  function scoreWorkbookSheets(sheetnames) {
    var score = 0;
    for (var i = 0; i < GANTT_ENTITIES.length; i++) {
      if (sheetnames.indexOf(GANTT_ENTITIES[i]) !== -1) score += 1;
    }
    return score;
  }

  function readDatapackageTextOrNull(text) {
    if (!text) return null;
    try { return JSON.parse(text); }
    catch (e) { return null; }
  }

  async function readDatapackageFromHandle(handle) {
    var candidateNames = [".__ontobdc__", ".__infobim__"];
    for (var i = 0; i < candidateNames.length; i++) {
      try {
        var metaDir = await handle.getDirectoryHandle(candidateNames[i]);
        var pkgFile = await metaDir.getFileHandle("datapackage.json");
        var blob = await pkgFile.getFile();
        var txt = await blob.text();
        var parsed = readDatapackageTextOrNull(txt);
        if (parsed && typeof parsed === "object") return parsed;
      } catch (e) {
        if (e && (e.name === "NotFoundError" || e.name === "TypeMismatchError")) continue;
        if (e && e.name === "NotAllowedError") throw e;
      }
    }
    return { resources: [] };
  }

  function buildResourceIndex(resources) {
    var byId = {};
    if (!Array.isArray(resources)) return byId;
    for (var i = 0; i < resources.length; i++) {
      var item = resources[i];
      if (!item || typeof item !== "object") continue;
      var identifier = (item.entityIdentifier || item.name || "").toString();
      if (identifier) byId[identifier] = item;
    }
    return byId;
  }

  function parseWorkbookFromArrayBuffer(buffer) {
    var XLSX = window.XLSX;
    var wb = XLSX.read(buffer, { type: "array", cellDates: true });
    var sheetnames = wb.SheetNames;
    return {
      workbook: wb,
      sheetnames: sheetnames,
      sheetRows: function (sheetName) {
        if (!sheetName || sheetnames.indexOf(sheetName) === -1) return [];
        var ws = wb.Sheets[sheetName];
        if (!ws) return [];
        var jsonArr = XLSX.utils.sheet_to_json(ws, { header: 1, raw: true, defval: null });
        if (!jsonArr || jsonArr.length === 0) return [];
        var rawHeader = jsonArr.shift() || [];
        var header = [];
        for (var h = 0; h < rawHeader.length; h++) {
          var val = rawHeader[h];
          header.push(val === null || val === undefined ? "" : String(val).trim());
        }
        if (header.every(function (col) { return !col; })) return [];
        var records = [];
        for (var r = 0; r < jsonArr.length; r++) {
          var row = jsonArr[r];
          if (!row || row.every(function (cell) { return cell === null || cell === undefined || cell === ""; })) continue;
          var record = {};
          for (var c = 0; c < header.length; c++) {
            var col = header[c];
            if (!col) continue;
            var value = c < row.length ? row[c] : null;
            if (value === null || value === undefined || value === "") continue;
            var text = "";
            if (typeof value === "string") {
              text = value.trim();
            } else if (value instanceof Date) {
              var y = value.getFullYear();
              var m = String(value.getMonth() + 1).padStart(2, "0");
              var d = String(value.getDate()).padStart(2, "0");
              var hh = String(value.getHours()).padStart(2, "0");
              var mm = String(value.getMinutes()).padStart(2, "0");
              var ss = String(value.getSeconds()).padStart(2, "0");
              if (hh === "00" && mm === "00" && ss === "00") {
                text = y + "-" + m + "-" + d;
              } else {
                text = y + "-" + m + "-" + d + "T" + hh + ":" + mm + ":" + ss;
              }
            } else if (typeof value === "number" && Number.isFinite(value)) {
              text = value.toString();
            } else if (typeof value === "boolean") {
              text = value ? "TRUE" : "FALSE";
            } else {
              text = String(value).trim();
            }
            if (text) record[col] = text;
          }
          if (Object.keys(record).length > 0) records.push(record);
        }
        return records;
      }
    };
  }

  async function sheetJsFindWorkbook(containerHandle, datapackage, resourceIndex) {
    var rootResource = resourceIndex.IfcWorkSchedule;
    if (rootResource && rootResource.path) {
      try {
        var pathSegments = String(rootResource.path).split("/").filter(Boolean);
        var cursor = containerHandle;
        for (var i = 0; i < pathSegments.length - 1; i++) {
          cursor = await cursor.getDirectoryHandle(pathSegments[i]);
        }
        var leaf = pathSegments[pathSegments.length - 1];
        var fileHandle = await cursor.getFileHandle(leaf);
        if (fileHandle && fileHandle.kind === "file") {
          return { fileHandle: fileHandle, relPath: String(rootResource.path) };
        }
      } catch (e) { /* fall through to rglob */ }
    }
    var allFiles = await collectFilesRecursively(containerHandle);
    var candidates = allFiles.filter(function (f) {
      return WORKBOOK_EXT_RE.test(f.handle.name) && !f.handle.name.startsWith("~$");
    });
    candidates.sort(function (a, b) {
      return a.relPath.localeCompare(b.relPath);
    });
    var best = null;
    var bestScore = -1;
    for (var j = 0; j < candidates.length; j++) {
      try {
        var ab = await readFileAsArrayBuffer(candidates[j].handle);
        var parsed = parseWorkbookFromArrayBuffer(ab);
        var score = scoreWorkbookSheets(parsed.sheetnames);
        if (score > bestScore) {
          bestScore = score;
          best = { fileHandle: candidates[j].handle, relPath: candidates[j].relPath, parsed: parsed };
        }
        if (score === GANTT_ENTITIES.length) break;
      } catch (wbErr) { /* skip unreadable */ }
    }
    return best;
  }

  function buildGanttNodesFromWorkbook(viewPayload, parsedWorkbook, resourceIndex) {
    var byIdentifier = Object.assign({}, resourceIndex);
    for (var i = 0; i < GANTT_ENTITIES.length; i++) {
      var entityName = GANTT_ENTITIES[i];
      if (!byIdentifier[entityName]) {
        byIdentifier[entityName] = { dialect: { excel: { sheet: entityName } } };
      }
    }
    var scheduleUri = (viewPayload && viewPayload.scheduleUri) ? viewPayload.scheduleUri : "urn:ontobdc:schedule";
    var nodes = [];
    for (var k = 0; k < GANTT_ENTITIES.length; k++) {
      var entityName = GANTT_ENTITIES[k];
      var resource = byIdentifier[entityName];
      if (!resource) continue;
      var sheetName = (((resource.dialect || {}).excel || {}).sheet) || entityName;
      var rows = parsedWorkbook.sheetRows(sheetName);
      for (var r = 0; r < rows.length; r++) {
        var record = rows[r];
        var globalId = (record && record.GlobalId) ? record.GlobalId : (entityName + "-" + r);
        var node = {
          "@id": scheduleUri + "/" + entityName + "/" + globalId,
          "@type": [IBIM + entityName]
        };
        var columns = Object.keys(record || {});
        for (var c = 0; c < columns.length; c++) {
          var col = columns[c];
          if (!col) continue;
          node[IBIM + col] = [{ "@value": record[col] }];
        }
        nodes.push(node);
      }
    }
    var counts = {};
    for (var j = 0; j < GANTT_ENTITIES.length; j++) {
      var countName = GANTT_ENTITIES[j];
      var typeUri = IBIM + countName;
      var n = 0;
      for (var z = 0; z < nodes.length; z++) {
        var types = nodes[z]["@type"];
        if (Array.isArray(types) && types[0] === typeUri) n += 1;
      }
      counts[countName] = n;
    }
    return { nodes: nodes, counts: counts };
  }

  async function sheetJsRefreshFromWorkbook() {
    // Re-normalize payload at call time — connection_state.js or the HTML
    // inline payload may have set runtime.WORKSTREAM_PAYLOAD / window.infoBimIfcWorkScheduleView
    // after this IIFE captured its top-level VIEW_PAYLOAD copy.
    if (!runtime.VIEW_PAYLOAD) {
      runtime.VIEW_PAYLOAD = runtime.WORKSTREAM_PAYLOAD || window.infoBimIfcWorkScheduleView || null;
    }
    var currentViewPayload = runtime.VIEW_PAYLOAD;
    if (!currentViewPayload) throw new Error("No IfcWorkSchedule context payload present on this page.");
    var handle = await resolveContainerHandle();
    if (!handle) return null;
    var datapackage = await readDatapackageFromHandle(handle);
    var resources = (datapackage && Array.isArray(datapackage.resources)) ? datapackage.resources : [];
    var resourceIndex = buildResourceIndex(resources);
    var best = await sheetJsFindWorkbook(handle, datapackage, resourceIndex);
    if (!best || !best.fileHandle) {
      throw new Error(
        "No IfcWorkSchedule workbook found in the connected folder. " +
        "Expected either a datapackage.json with an IfcWorkSchedule resource " +
        "or an *.xlsx file with sheets IfcWorkSchedule / IfcTask / IfcTaskTime / IfcRelSequence."
      );
    }
    var parsed = best.parsed;
    if (!parsed) {
      var ab = await readFileAsArrayBuffer(best.fileHandle);
      parsed = parseWorkbookFromArrayBuffer(ab);
    }
    var built = buildGanttNodesFromWorkbook(currentViewPayload, parsed, resourceIndex);
    // Stashed so saveTaskDates() below can write back into this exact same
    // file/workbook object later without re-scanning the whole connected
    // folder again on every single date edit.
    runtime.state.workbookFileHandle = best.fileHandle;
    runtime.state.workbookSheetJs = parsed.workbook;
    runtime.state.workbookRelPath = best.relPath || "";
    return {
      nodes: built.nodes,
      workbookPath: best.relPath || "",
      counts: built.counts
    };
  }

  // ---------------------------------------------------------------------------
  // Task date editing (the Gantt's double-click modal, task_table_timeline.js)
  // writes ScheduleStart/ScheduleFinish back into the *same* connected xlsx
  // file it read them from, via SheetJS + the File System Access API write
  // stream already granted by container_connection.js's directory picker —
  // no pyodide/openpyxl round-trip needed since SheetJS can both read and
  // write .xlsx natively in the browser.
  // ---------------------------------------------------------------------------
  // completion is a 0-100 number, or null to leave the cell untouched. When it
  // is given but the IfcTaskTime sheet has neither a Completion nor a
  // PercentComplete column, a Completion column is appended to the header so the
  // value has somewhere to live.
  async function writeScheduleCellsToWorkbook(taskTimeGlobalId, startIso, finishIso, completion) {
    var fileHandle = runtime.state.workbookFileHandle;
    var wb = runtime.state.workbookSheetJs;
    if (!fileHandle || !wb) {
      throw new Error(t("ganttErrNoFolder"));
    }
    if (!window.XLSX || !window.XLSX.utils) {
      throw new Error(t("ganttErrSheetJsUnavailable"));
    }
    var XLSX = window.XLSX;
    if (wb.SheetNames.indexOf("IfcTaskTime") === -1) {
      throw new Error(t("ganttErrSheetMissing", { sheet: "IfcTaskTime" }));
    }
    var ws = wb.Sheets["IfcTaskTime"];
    var range = XLSX.utils.decode_range(ws["!ref"]);
    var headerRow = range.s.r;
    var colByName = {};
    for (var c = range.s.c; c <= range.e.c; c++) {
      var headerCell = ws[XLSX.utils.encode_cell({ r: headerRow, c: c })];
      var colName = (headerCell && headerCell.v !== undefined && headerCell.v !== null)
        ? String(headerCell.v).trim() : "";
      if (colName) colByName[colName] = c;
    }
    var gidCol = colByName["GlobalId"];
    var startCol = colByName["ScheduleStart"];
    var finishCol = colByName["ScheduleFinish"];
    if (gidCol === undefined || startCol === undefined || finishCol === undefined) {
      throw new Error(t("ganttErrColumnsMissing", { sheet: "IfcTaskTime", columns: "GlobalId/ScheduleStart/ScheduleFinish" }));
    }
    var targetRow = -1;
    for (var r = headerRow + 1; r <= range.e.r; r++) {
      var gidCell = ws[XLSX.utils.encode_cell({ r: r, c: gidCol })];
      if (gidCell && String(gidCell.v).trim() === taskTimeGlobalId) {
        targetRow = r;
        break;
      }
    }
    if (targetRow === -1) {
      throw new Error(t("ganttErrRowNotFound", { sheet: "IfcTaskTime", globalId: taskTimeGlobalId }));
    }
    ws[XLSX.utils.encode_cell({ r: targetRow, c: startCol })] = { t: "s", v: startIso };
    ws[XLSX.utils.encode_cell({ r: targetRow, c: finishCol })] = { t: "s", v: finishIso };

    if (completion !== null && completion !== undefined) {
      var completionCol = colByName["Completion"];
      if (completionCol === undefined) completionCol = colByName["PercentComplete"];
      if (completionCol === undefined) {
        completionCol = range.e.c + 1;
        ws[XLSX.utils.encode_cell({ r: headerRow, c: completionCol })] = { t: "s", v: "Completion" };
        range.e.c = completionCol;
        ws["!ref"] = XLSX.utils.encode_range(range);
      }
      ws[XLSX.utils.encode_cell({ r: targetRow, c: completionCol })] = { t: "n", v: completion };
    }

    var out = XLSX.write(wb, { type: "array", bookType: "xlsx" });
    var writable = await fileHandle.createWritable();
    await writable.write(out);
    await writable.close();
    if (runtime.scheduleSurfaceRegeneration) {
      runtime.scheduleSurfaceRegeneration("ifc_work_schedule_task_time");
    }
  }

  function _isoDateOnly(d) {
    var y = d.getFullYear();
    var m = String(d.getMonth() + 1).padStart(2, "0");
    var day = String(d.getDate()).padStart(2, "0");
    return y + "-" + m + "-" + day + "T00:00:00";
  }

  // newCompletion is a 0-100 number, or null/undefined to leave completion as it
  // is (the date-only edit path).
  async function saveTaskDates(task, newStart, newFinish, newCompletion) {
    if (!task) throw new Error(t("ganttErrNoTaskSelected"));
    var taskTimeGlobalId = task.rawTime ? runtime.getGlobalId(task.rawTime) : "";
    if (!taskTimeGlobalId) {
      throw new Error(t("ganttErrNoTaskTime"));
    }
    var startIso = _isoDateOnly(newStart);
    var finishIso = _isoDateOnly(newFinish);
    var completion = (newCompletion === null || newCompletion === undefined)
      ? null : Math.max(0, Math.min(100, Number(newCompletion)));

    await writeScheduleCellsToWorkbook(taskTimeGlobalId, startIso, finishIso, completion);

    // Keep the in-memory JSON-LD node consistent too (task.rawTime is the
    // very node object living inside runtime.state.graph, not a copy — see
    // graph_reader.js), so a later runParse() would re-derive the same
    // values instead of reverting this edit before the next real refresh.
    task.rawTime[runtime.IBIM_NS + "ScheduleStart"] = [{ "@value": startIso }];
    task.rawTime[runtime.IBIM_NS + "ScheduleFinish"] = [{ "@value": finishIso }];
    if (completion !== null) {
      task.rawTime[runtime.IBIM_NS + "Completion"] = [{ "@value": completion }];
      task.completion = completion;
    }
    task.scheduleStart = newStart;
    task.scheduleFinish = newFinish;
    var days = runtime.diffDays(newStart, newFinish);
    if (days < 0) days = 0;
    if (days === 0) days = 1;
    task.duration_days = days;
  }
  runtime.saveTaskDates = saveTaskDates;

  // ---------------------------------------------------------------------------
  // Task rename / WBS edit (the modal title pencil and the WBS field,
  // task_table_timeline.js) — same write path as saveTaskDates above, but
  // against a column on the IfcTask sheet itself (Name and Identification both
  // live on the task, not on its IfcTaskTime row). columnCandidates is a
  // priority list of acceptable header names; createIfMissing appends the first
  // candidate as a new header column when the sheet carries none of them.
  // ---------------------------------------------------------------------------
  async function writeTaskCellToWorkbook(taskGlobalId, columnCandidates, value, createIfMissing) {
    var fileHandle = runtime.state.workbookFileHandle;
    var wb = runtime.state.workbookSheetJs;
    if (!fileHandle || !wb) {
      throw new Error(t("ganttErrNoFolder"));
    }
    if (!window.XLSX || !window.XLSX.utils) {
      throw new Error(t("ganttErrSheetJsUnavailable"));
    }
    var XLSX = window.XLSX;
    if (wb.SheetNames.indexOf("IfcTask") === -1) {
      throw new Error(t("ganttErrSheetMissing", { sheet: "IfcTask" }));
    }
    var ws = wb.Sheets["IfcTask"];
    var range = XLSX.utils.decode_range(ws["!ref"]);
    var headerRow = range.s.r;
    var colByName = {};
    for (var c = range.s.c; c <= range.e.c; c++) {
      var headerCell = ws[XLSX.utils.encode_cell({ r: headerRow, c: c })];
      var colName = (headerCell && headerCell.v !== undefined && headerCell.v !== null)
        ? String(headerCell.v).trim() : "";
      if (colName) colByName[colName] = c;
    }
    var gidCol = colByName["GlobalId"];
    var valueCol;
    for (var i = 0; i < columnCandidates.length; i++) {
      if (colByName[columnCandidates[i]] !== undefined) {
        valueCol = colByName[columnCandidates[i]];
        break;
      }
    }
    if (valueCol === undefined && createIfMissing) {
      valueCol = range.e.c + 1;
      ws[XLSX.utils.encode_cell({ r: headerRow, c: valueCol })] = { t: "s", v: columnCandidates[0] };
      range.e.c = valueCol;
      ws["!ref"] = XLSX.utils.encode_range(range);
    }
    if (gidCol === undefined || valueCol === undefined) {
      throw new Error(t("ganttErrColumnsMissing", { sheet: "IfcTask", columns: "GlobalId/" + columnCandidates[0] }));
    }
    var targetRow = -1;
    for (var r = headerRow + 1; r <= range.e.r; r++) {
      var gidCell = ws[XLSX.utils.encode_cell({ r: r, c: gidCol })];
      if (gidCell && String(gidCell.v).trim() === taskGlobalId) {
        targetRow = r;
        break;
      }
    }
    if (targetRow === -1) {
      throw new Error(t("ganttErrRowNotFound", { sheet: "IfcTask", globalId: taskGlobalId }));
    }
    ws[XLSX.utils.encode_cell({ r: targetRow, c: valueCol })] = { t: "s", v: value };

    var out = XLSX.write(wb, { type: "array", bookType: "xlsx" });
    var writable = await fileHandle.createWritable();
    await writable.write(out);
    await writable.close();
    if (runtime.scheduleSurfaceRegeneration) {
      runtime.scheduleSurfaceRegeneration("ifc_work_schedule_task");
    }
  }

  async function saveTaskName(task, newName) {
    if (!task) throw new Error(t("ganttErrNoTaskSelected"));
    if (!task.globalId) {
      throw new Error(t("ganttErrNoGlobalId"));
    }
    await writeTaskCellToWorkbook(task.globalId, ["Name"], newName, false);

    // Same reasoning as saveTaskDates: task.raw is the live IfcTask node
    // inside runtime.state.graph, so patching it here keeps a later
    // runParse() from reverting the rename before the next real refresh.
    if (task.raw) {
      task.raw[runtime.IBIM_NS + "Name"] = [{ "@value": newName }];
    }
    task.name = newName;
  }
  runtime.saveTaskName = saveTaskName;

  async function writeScheduleMetadataToWorkbook(scheduleGlobalId, columnCandidates, value) {
    var fileHandle = runtime.state.workbookFileHandle;
    var wb = runtime.state.workbookSheetJs;
    if (!fileHandle || !wb) throw new Error(t("ganttErrNoFolder"));
    if (!window.XLSX || !window.XLSX.utils) throw new Error(t("ganttErrSheetJsUnavailable"));
    var XLSX = window.XLSX;
    var sheetName = "IfcWorkSchedule";
    if (wb.SheetNames.indexOf(sheetName) === -1) {
      throw new Error(t("ganttErrSheetMissing", { sheet: sheetName }));
    }
    var ws = wb.Sheets[sheetName];
    var range = XLSX.utils.decode_range(ws["!ref"]);
    var headerRow = range.s.r;
    var columns = {};
    for (var c = range.s.c; c <= range.e.c; c++) {
      var headerCell = ws[XLSX.utils.encode_cell({ r: headerRow, c: c })];
      var header = headerCell && headerCell.v != null ? String(headerCell.v).trim() : "";
      if (header) columns[header] = c;
    }
    var gidColumn = columns.GlobalId;
    var valueColumn;
    for (var i = 0; i < columnCandidates.length; i++) {
      if (columns[columnCandidates[i]] !== undefined) {
        valueColumn = columns[columnCandidates[i]];
        break;
      }
    }
    if (valueColumn === undefined) {
      valueColumn = range.e.c + 1;
      ws[XLSX.utils.encode_cell({ r: headerRow, c: valueColumn })] = { t: "s", v: columnCandidates[0] };
      range.e.c = valueColumn;
      ws["!ref"] = XLSX.utils.encode_range(range);
    }
    if (gidColumn === undefined) {
      throw new Error(t("ganttErrColumnsMissing", { sheet: sheetName, columns: "GlobalId" }));
    }
    var targetRow = -1;
    for (var r = headerRow + 1; r <= range.e.r; r++) {
      var gidCell = ws[XLSX.utils.encode_cell({ r: r, c: gidColumn })];
      if (gidCell && String(gidCell.v).trim() === scheduleGlobalId) { targetRow = r; break; }
    }
    if (targetRow === -1) {
      throw new Error(t("ganttErrRowNotFound", { sheet: sheetName, globalId: scheduleGlobalId }));
    }
    ws[XLSX.utils.encode_cell({ r: targetRow, c: valueColumn })] = { t: "s", v: value };
    var output = XLSX.write(wb, { type: "array", bookType: "xlsx" });
    var writable = await fileHandle.createWritable();
    await writable.write(output);
    await writable.close();
    if (runtime.scheduleSurfaceRegeneration) {
      runtime.scheduleSurfaceRegeneration("ifc_work_schedule_metadata");
    }
  }

  async function saveScheduleField(column, value) {
    if (column !== "Name" && column !== "Description") {
      throw new Error("Unsupported IfcWorkSchedule column: " + column);
    }
    var schedule = runtime.state.schedule;
    // The embedded graph node identifies the schedule with dcterms:identifier,
    // not a GlobalId property (only the connected workbook row carries that),
    // so fall back to the alias-aware identifier — it is the same compressed
    // IFC GUID the IfcWorkSchedule sheet keys its GlobalId column on.
    var globalId = schedule
      ? (runtime.getGlobalId(schedule)
        || runtime.aliasAwareLiteral(schedule, "Identification")
        || runtime.literalValue(schedule, "identifier")
        || "")
      : "";
    if (!schedule || !globalId) throw new Error(t("ganttErrNoScheduleGlobalId"));
    await writeScheduleMetadataToWorkbook(globalId, [column], value);
    schedule[runtime.IBIM_NS + column] = [{ "@value": value }];
  }
  runtime.saveScheduleField = saveScheduleField;

  async function saveTaskIdentification(task, newIdentification) {
    if (!task) throw new Error(t("ganttErrNoTaskSelected"));
    if (!task.globalId) {
      throw new Error(t("ganttErrNoGlobalId"));
    }
    await writeTaskCellToWorkbook(task.globalId, ["Identification", "Identifier"], newIdentification, true);

    if (task.raw) {
      task.raw[runtime.IBIM_NS + "Identification"] = [{ "@value": newIdentification }];
    }
    task.identification = newIdentification;
  }
  runtime.saveTaskIdentification = saveTaskIdentification;

  // ---------------------------------------------------------------------------
  // Task creation (the "+" button in the Page header, task_table_timeline.js)
  // — appends one new row to IfcTask and one to IfcTaskTime rather than
  // patching the in-memory graph itself: the caller re-reads the workbook via
  // refreshFromWorkbook() afterward, which is simpler and safer than hand
  // -building a JSON-LD node shaped exactly like graph_reader.js expects.
  // ---------------------------------------------------------------------------
  function _randomGlobalId() {
    // Not a byte-perfect compressed-IFC GUID (see ifcopenshell.guid.new()
    // for that) -- nothing downstream validates the format, only that it
    // is a unique string, so a plain random token in the same alphabet is
    // enough and avoids pulling a GUID library into the browser bundle.
    var alphabet = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz_$";
    var bytes = new Uint8Array(22);
    if (window.crypto && typeof window.crypto.getRandomValues === "function") {
      window.crypto.getRandomValues(bytes);
    } else {
      for (var i = 0; i < bytes.length; i++) bytes[i] = Math.floor(Math.random() * 256);
    }
    var out = "";
    for (var j = 0; j < bytes.length; j++) out += alphabet[bytes[j] % alphabet.length];
    return out;
  }

  function _appendRowToSheet(wb, sheetName, valuesByColumn) {
    var XLSX = window.XLSX;
    if (wb.SheetNames.indexOf(sheetName) === -1) {
      throw new Error(t("ganttErrSheetMissing", { sheet: sheetName }));
    }
    var ws = wb.Sheets[sheetName];
    var range = XLSX.utils.decode_range(ws["!ref"]);
    var headerRow = range.s.r;
    var colByName = {};
    for (var c = range.s.c; c <= range.e.c; c++) {
      var headerCell = ws[XLSX.utils.encode_cell({ r: headerRow, c: c })];
      var colName = (headerCell && headerCell.v !== undefined && headerCell.v !== null)
        ? String(headerCell.v).trim() : "";
      if (colName) colByName[colName] = c;
    }
    var newRow = range.e.r + 1;
    var maxCol = range.e.c;
    var keys = Object.keys(valuesByColumn);
    for (var k = 0; k < keys.length; k++) {
      var colName2 = keys[k];
      var value = valuesByColumn[colName2];
      if (value === undefined || value === null || value === "") continue;
      var colIdx = colByName[colName2];
      if (colIdx === undefined) {
        // Sheet has no such column yet -- append it to the header row so the
        // value is not silently dropped (e.g. Identification / WBS on a
        // workbook whose IfcTask sheet never carried that column).
        colIdx = maxCol + 1;
        colByName[colName2] = colIdx;
        ws[XLSX.utils.encode_cell({ r: headerRow, c: colIdx })] = { t: "s", v: colName2 };
        maxCol = colIdx;
      }
      var cellType = typeof value === "boolean" ? "b" : (typeof value === "number" ? "n" : "s");
      ws[XLSX.utils.encode_cell({ r: newRow, c: colIdx })] = { t: cellType, v: value };
      if (colIdx > maxCol) maxCol = colIdx;
    }
    range.e.r = newRow;
    if (maxCol > range.e.c) range.e.c = maxCol;
    ws["!ref"] = XLSX.utils.encode_range(range);
  }

  async function createTask(fields) {
    var fileHandle = runtime.state.workbookFileHandle;
    var wb = runtime.state.workbookSheetJs;
    if (!fileHandle || !wb) {
      throw new Error(t("ganttErrNoFolder"));
    }
    if (!window.XLSX || !window.XLSX.utils) {
      throw new Error(t("ganttErrSheetJsUnavailable"));
    }
    var name = String((fields && fields.name) || "").trim();
    if (!name) {
      throw new Error(t("ganttErrNameRequired"));
    }
    var taskGlobalId = _randomGlobalId();
    var taskTimeGlobalId = _randomGlobalId();

    _appendRowToSheet(wb, "IfcTask", {
      GlobalId: taskGlobalId,
      Identification: (fields && fields.identification) || undefined,
      Name: name,
      IsMilestone: false,
      TaskTimeGlobalId: taskTimeGlobalId,
    });
    _appendRowToSheet(wb, "IfcTaskTime", {
      GlobalId: taskTimeGlobalId,
      Name: name,
      ScheduleStart: (fields && fields.scheduleStart) || undefined,
      ScheduleFinish: (fields && fields.scheduleFinish) || undefined,
    });

    var out = window.XLSX.write(wb, { type: "array", bookType: "xlsx" });
    var writable = await fileHandle.createWritable();
    await writable.write(out);
    await writable.close();
    if (runtime.scheduleSurfaceRegeneration) {
      runtime.scheduleSurfaceRegeneration("ifc_work_schedule_task_created");
    }
  }
  runtime.createTask = createTask;

  // ---------------------------------------------------------------------------
  // Task deletion (the trash icon at the bottom of the edit modal,
  // task_table_timeline.js) — drops the task's IfcTask row, its linked
  // IfcTaskTime row and any IfcRelSequence row that still points at it, then
  // (like createTask) leaves the caller to refreshFromWorkbook() so the screen
  // re-renders from the rewritten file.
  // ---------------------------------------------------------------------------
  function _deleteRowsFromSheet(wb, sheetName, shouldDeleteRow) {
    var XLSX = window.XLSX;
    if (wb.SheetNames.indexOf(sheetName) === -1) return 0;
    var ws = wb.Sheets[sheetName];
    var aoa = XLSX.utils.sheet_to_json(ws, { header: 1, raw: true, defval: null });
    if (!aoa.length) return 0;
    var header = aoa[0] || [];
    var kept = [header];
    var removed = 0;
    for (var r = 1; r < aoa.length; r++) {
      var cells = aoa[r] || [];
      var rowObj = {};
      for (var c = 0; c < header.length; c++) {
        var key = (header[c] === undefined || header[c] === null) ? "" : String(header[c]).trim();
        if (key) rowObj[key] = cells[c];
      }
      if (shouldDeleteRow(rowObj)) { removed++; continue; }
      kept.push(cells);
    }
    if (removed > 0) wb.Sheets[sheetName] = XLSX.utils.aoa_to_sheet(kept);
    return removed;
  }

  async function deleteTask(task) {
    var fileHandle = runtime.state.workbookFileHandle;
    var wb = runtime.state.workbookSheetJs;
    if (!fileHandle || !wb) {
      throw new Error(t("ganttErrNoFolder"));
    }
    if (!window.XLSX || !window.XLSX.utils) {
      throw new Error(t("ganttErrSheetJsUnavailable"));
    }
    if (!task) throw new Error(t("ganttErrNoTaskSelected"));
    var taskGlobalId = task.globalId || "";
    if (!taskGlobalId) throw new Error(t("ganttErrNoGlobalId"));
    var taskTimeGlobalId = task.rawTime ? runtime.getGlobalId(task.rawTime) : "";

    _deleteRowsFromSheet(wb, "IfcTask", function (row) {
      return String(row.GlobalId || "").trim() === taskGlobalId;
    });
    if (taskTimeGlobalId) {
      _deleteRowsFromSheet(wb, "IfcTaskTime", function (row) {
        return String(row.GlobalId || "").trim() === taskTimeGlobalId;
      });
    }
    _deleteRowsFromSheet(wb, "IfcRelSequence", function (row) {
      var refs = [
        row.RelatingProcess, row.RelatingProcessGlobalId, row.Predecessor,
        row.RelatedProcess, row.RelatedProcessGlobalId, row.Successor
      ];
      for (var i = 0; i < refs.length; i++) {
        if (refs[i] !== undefined && refs[i] !== null
          && String(refs[i]).indexOf(taskGlobalId) !== -1) {
          return true;
        }
      }
      return false;
    });

    var out = window.XLSX.write(wb, { type: "array", bookType: "xlsx" });
    var writable = await fileHandle.createWritable();
    await writable.write(out);
    await writable.close();
    if (runtime.scheduleSurfaceRegeneration) {
      runtime.scheduleSurfaceRegeneration("ifc_work_schedule_task_deleted");
    }
  }
  runtime.deleteTask = deleteTask;

  // Reads the connected container's own workbook and returns the schedule,
  // its tasks, their task times and their sequence relations as JSON-LD
  // nodes — the same shape the Page's embedded graph already carries, so
  // graph_reader consumes them unchanged. The column-to-property mapping
  // is taken from datapackage.json's own resource schema rather than from
  // a hardcoded column list, so adding a field to the facade contract is
  // enough for it to reach the Page.
  const WORKBOOK_PARSE_SCRIPT = `
import json
from pathlib import Path

from openpyxl import load_workbook

IBIM = "https://infobim.org/ontology/ns#"
GANTT_ENTITIES = ("IfcWorkSchedule", "IfcTask", "IfcTaskTime", "IfcRelSequence")

payload = json.loads(view_payload_json)
container = Path(container_mount_path)

datapackage_path = container / payload.get("datapackagePath", ".__ontobdc__/datapackage.json")
datapackage = json.loads(datapackage_path.read_text(encoding="utf-8")) if datapackage_path.is_file() else {"resources": []}

resources = datapackage.get("resources", []) or []
by_identifier = {}
for item in resources:
    identifier = item.get("entityIdentifier") or item.get("name") or ""
    if identifier:
        by_identifier[identifier] = item


def _find_workbook_fallback(root: Path):
    best = None
    best_score = -1
    for path in sorted(root.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        except Exception:
            continue
        try:
            score = sum(1 for s in GANTT_ENTITIES if s in wb.sheetnames)
        finally:
            wb.close()
        if score > best_score:
            best_score = score
            best = path
        if score == len(GANTT_ENTITIES):
            break
    return best


root_resource = by_identifier.get("IfcWorkSchedule")
if root_resource is not None:
    workbook_path = (datapackage_path.parent / root_resource.get("path", "")).resolve()
else:
    found = _find_workbook_fallback(container)
    if found is None:
        raise FileNotFoundError(
            "No IfcWorkSchedule workbook found in the connected folder. "
            "Expected either a datapackage.json with an IfcWorkSchedule resource "
            "or an *.xlsx file with sheets IfcWorkSchedule / IfcTask / IfcTaskTime / IfcRelSequence."
        )
    workbook_path = found.resolve()
    for name in GANTT_ENTITIES:
        if name not in by_identifier:
            by_identifier[name] = {"dialect": {"excel": {"sheet": name}}}

if not workbook_path.is_file():
    raise FileNotFoundError("workbook not found: " + str(workbook_path))

workbook = load_workbook(filename=str(workbook_path), data_only=True, read_only=True)


def sheet_rows(resource):
    sheet_name = (resource.get("dialect", {}).get("excel", {}) or {}).get("sheet")
    if not sheet_name or sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return []
    records = []
    for raw in rows:
        record = {}
        for column, value in zip(header, raw):
            if not column or value is None:
                continue
            text = value.isoformat() if hasattr(value, "isoformat") else str(value).strip()
            if text:
                record[column] = text
        if record:
            records.append(record)
    return records


def node_for(entity_name, record, index):
    global_id = record.get("GlobalId") or (entity_name + "-" + str(index))
    node = {
        "@id": payload.get("scheduleUri", "urn:ontobdc:schedule") + "/" + entity_name + "/" + global_id,
        "@type": [IBIM + entity_name],
    }
    for column, value in record.items():
        node[IBIM + column] = [{"@value": value}]
    return node


nodes = []
for entity_name in GANTT_ENTITIES:
    resource = by_identifier.get(entity_name)
    if resource is None:
        continue
    for index, record in enumerate(sheet_rows(resource)):
        nodes.append(node_for(entity_name, record, index))

workbook.close()

json.dumps({
    "nodes": nodes,
    "workbookPath": str(workbook_path),
    "counts": {
        name: len([n for n in nodes if n["@type"][0] == IBIM + name])
        for name in GANTT_ENTITIES
    },
})
`;

  async function mountContainer(pyodide, handle) {
    if (runtime.state.activeMountPath) {
      try {
        pyodide.FS.unmount(runtime.state.activeMountPath);
      } catch (unmountError) {
        console.warn("[ifc-work-schedule-view] unmount failed", unmountError);
      }
    }
    const mountPath = `/container_${Date.now()}`;
    pyodide.FS.mkdirTree(mountPath);
    await pyodide.mountNativeFS(mountPath, handle);
    runtime.state.activeMountPath = mountPath;
    return mountPath;
  }

  // Replaces the Page's embedded graph with the workbook's own nodes and
  // re-runs the parse. The embedded graph stays the single source every
  // downstream script reads, so nothing below this has to know a workbook
  // exists.
  function applyNodes(nodes) {
    const element = document.getElementById("ontobdc-page-jsonld");
    if (!element) return false;
    let existing = [];
    try {
      const parsed = JSON.parse(element.textContent);
      existing = Array.isArray(parsed) ? parsed : [parsed];
    } catch (parseError) {
      existing = [];
    }
    const fromWorkbook = new Set(nodes.map((node) => node["@id"]));
    const merged = existing.filter((node) => !fromWorkbook.has(node && node["@id"])).concat(nodes);
    element.textContent = JSON.stringify(merged);
    runtime.state.workbookNodeCount = nodes.length;
    // No optional call here: graph_reader always loads before this script,
    // and a missing runParse means the workbook was read and then silently
    // thrown away — the failure has to be loud.
    for (const step of ["runParse", "runRender", "runArrows"]) {
      if (typeof runtime[step] !== "function") {
        throw new Error(
          "the gantt runtime did not export " + step
          + "; the workbook nodes cannot be rendered."
        );
      }
    }
    // Parse, then draw, then arrows — arrows depend on the bar geometries
    // the render has just produced, which is why they load last too.
    runtime.runParse();
    runtime.runRender();
    runtime.runArrows();
    return true;
  }

  async function refreshFromWorkbook() {
    if (!VIEW_PAYLOAD) {
      throw new Error("No IfcWorkSchedule context payload present on this page.");
    }
    const handle = await resolveContainerHandle();
    if (!handle) return null;

    if (hasSheetJS()) {
      const result = await sheetJsRefreshFromWorkbook();
      applyNodes((result && result.nodes) || []);
      runtime.state.workbook = result || null;
      if (result && result.counts) {
        console.log("[ifc-work-schedule-view] workbook parsed (SheetJS)", result.counts);
      }
      return result;
    }

    const pyodide = await ensurePyodide({ withOpenpyxl: true });
    const mountPath = await mountContainer(pyodide, handle);

    const raw = await withPyodideLock(() => {
      pyodide.globals.set("view_payload_json", JSON.stringify(VIEW_PAYLOAD));
      pyodide.globals.set("container_mount_path", mountPath);
      return pyodide.runPythonAsync(WORKBOOK_PARSE_SCRIPT);
    });

    const result = JSON.parse(raw);
    applyNodes(result.nodes || []);
    runtime.state.workbook = result;
    console.log("[ifc-work-schedule-view] workbook parsed (Pyodide fallback)", result.counts);
    return result;
  }

  Object.assign(runtime, {
    ensurePyodide: ensurePyodide,
    withPyodideLock: withPyodideLock,
    refreshFromWorkbook: refreshFromWorkbook,
    applyNodes: applyNodes,
    WORKBOOK_PARSE_SCRIPT: WORKBOOK_PARSE_SCRIPT,
    hasSheetJS: hasSheetJS,
    sheetJsRefreshFromWorkbook: sheetJsRefreshFromWorkbook,
    applyLiveWorkbookResult: function applyLiveWorkbookResult(result) {
      applyNodes((result && result.nodes) || []);
      runtime.state.workbook = result || null;
      if (result && result.counts) {
        console.log("[ifc-work-schedule-view] workbook parsed", result.counts);
      }
    }
  });
  console.log("[ifc-work-schedule-view] state end: pyodide_runtime_script_generated");
}(window));
"""

    def _task_table_timeline_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var rt = window.OntoBDCGanttViewRuntime = window.OntoBDCGanttViewRuntime || { state: {} };
  console.log("[gantt-view] state start: task_table_timeline_script_generated");

  // i18n_apply.js runs before this script and attaches the shared translator as
  // rt.t; fall back to the key itself if it somehow is not there yet.
  function t(key, vars) {
    if (typeof rt.t === "function") return rt.t(key, vars);
    var text = String(key);
    if (vars) { for (var k in vars) { text = text.split("{" + k + "}").join(String(vars[k])); } }
    return text;
  }

  function numOr(v, fallback) {
    var n = Number(v);
    if (isFinite(n) && !isNaN(n)) return n;
    var fb = Number(fallback);
    if (isFinite(fb) && !isNaN(fb)) return fb;
    return 0;
  }

  function createSvgEl(tag, attrs) {
    var el = document.createElementNS("http://www.w3.org/2000/svg", tag);
    if (attrs) {
      var keys = Object.keys(attrs);
      for (var i = 0; i < keys.length; i++) {
        var key = keys[i];
        var raw = attrs[key];
        var val;
        if (typeof raw === "number") {
          var n1 = numOr(raw, 0);
          // Clamp size/length attributes to non-negative so the SVG renderer
          // never logs "A negative value is not valid" for width/height/x/y
          // when a task has a zero or reversed date range on the workbook.
          if (key === "width" || key === "height" || key === "rx" || key === "ry" || key === "stroke-width" || key === "x1" || key === "x2" || key === "y1" || key === "y2") {
            n1 = Math.max(0, n1);
          }
          val = String(n1);
        } else if (typeof raw === "string" && raw !== "" && !isNaN(Number(raw))) {
          var n2 = numOr(raw, 0);
          if (key === "width" || key === "height" || key === "rx" || key === "ry" || key === "stroke-width" || key === "x1" || key === "x2" || key === "y1" || key === "y2") {
            n2 = Math.max(0, n2);
          }
          val = String(n2);
        } else {
          val = raw === null || raw === undefined ? "" : String(raw);
        }
        el.setAttribute(key, val);
      }
    }
    return el;
  }
  rt.createSvgEl = createSvgEl;

  // Double-click on a task (its left-table row or its right-panel bar /
  // milestone diamond) opens this small modal to change ScheduleStart /
  // ScheduleFinish. Built once and reused across every runRender() pass,
  // since the tbody/svg elements it hangs its listeners off of are
  // recreated on every pass but the dialog itself only needs to exist once.
  var editDialogEl = null;
  var editDialogWbsInput = null;
  var editDialogStartInput = null;
  var editDialogFinishInput = null;
  var editDialogCompletionInput = null;
  var editDialogCompletionField = null;
  var editDialogTitleTextEl = null;
  var editDialogRenameBtn = null;
  var editDialogTitleInput = null;
  var editDialogErrorEl = null;
  var editDialogSaveBtn = null;
  var editDialogDeleteBtn = null;
  var editDialogCurrentTask = null;
  var editDialogRenameSaving = false;

  function ensureEditDialog() {
    if (editDialogEl) return editDialogEl;

    var dialog = document.createElement("dialog");
    dialog.className = "onto-gantt-edit-dialog";

    var header = document.createElement("div");
    header.className = "onto-gantt-edit-dialog-header";

    var titleWrap = document.createElement("h3");
    titleWrap.className = "onto-gantt-edit-dialog-title";
    var titleText = document.createElement("span");
    titleText.className = "onto-gantt-edit-title-text";
    var renameBtn = document.createElement("button");
    renameBtn.type = "button";
    renameBtn.className = "onto-gantt-edit-rename-btn";
    renameBtn.setAttribute("aria-label", t("ganttActionRename"));
    renameBtn.textContent = "✎";
    var titleInput = document.createElement("input");
    titleInput.type = "text";
    titleInput.className = "onto-gantt-edit-title-input";
    titleInput.style.display = "none";
    titleWrap.appendChild(titleText);
    titleWrap.appendChild(renameBtn);
    titleWrap.appendChild(titleInput);

    var closeBtn = document.createElement("button");
    closeBtn.type = "button";
    closeBtn.setAttribute("aria-label", t("ganttActionClose"));
    closeBtn.textContent = "×";
    header.appendChild(titleWrap);
    header.appendChild(closeBtn);

    function enterRenameMode() {
      if (!editDialogCurrentTask) return;
      titleText.style.display = "none";
      renameBtn.style.display = "none";
      titleInput.style.display = "";
      titleInput.value = editDialogCurrentTask.name || "";
      titleInput.disabled = false;
      titleInput.focus();
      titleInput.select();
    }

    function exitRenameMode() {
      titleInput.style.display = "none";
      titleText.style.display = "";
      renameBtn.style.display = "";
    }

    renameBtn.addEventListener("click", enterRenameMode);

    titleInput.addEventListener("keydown", function (evt) {
      if (evt.key === "Enter") {
        evt.preventDefault();
        handleRenameSave();
      } else if (evt.key === "Escape") {
        evt.preventDefault();
        exitRenameMode();
      }
    });
    titleInput.addEventListener("blur", function () {
      if (editDialogRenameSaving) return;
      exitRenameMode();
    });

    function handleRenameSave() {
      var task = editDialogCurrentTask;
      if (!task) return;
      var newName = titleInput.value.trim();
      if (!newName) {
        showEditDialogError(t("ganttErrNameEmpty"));
        return;
      }
      if (newName === task.name) {
        exitRenameMode();
        return;
      }
      if (typeof rt.saveTaskName !== "function") {
        showEditDialogError(t("ganttErrSaveUnavailable"));
        return;
      }
      editDialogErrorEl.style.display = "none";
      editDialogRenameSaving = true;
      titleInput.disabled = true;
      Promise.resolve(rt.saveTaskName(task, newName)).then(function () {
        editDialogRenameSaving = false;
        titleInput.disabled = false;
        titleText.textContent = task.name;
        exitRenameMode();
        runRender();
        if (typeof rt.runArrows === "function") rt.runArrows();
      }).catch(function (err) {
        editDialogRenameSaving = false;
        titleInput.disabled = false;
        titleInput.focus();
        showEditDialogError((err && err.message) ? err.message : String(err));
      });
    }

    var body = document.createElement("div");
    body.className = "onto-gantt-edit-dialog-body";

    var wbsLabel = document.createElement("label");
    wbsLabel.className = "onto-gantt-edit-field";
    wbsLabel.textContent = t("ganttColumnWbs");
    var wbsInput = document.createElement("input");
    wbsInput.type = "text";
    wbsLabel.appendChild(wbsInput);

    var startLabel = document.createElement("label");
    startLabel.className = "onto-gantt-edit-field";
    startLabel.textContent = t("ganttFieldStart");
    var startInput = document.createElement("input");
    startInput.type = "date";
    startLabel.appendChild(startInput);

    var finishLabel = document.createElement("label");
    finishLabel.className = "onto-gantt-edit-field";
    finishLabel.textContent = t("ganttFieldFinish");
    var finishInput = document.createElement("input");
    finishInput.type = "date";
    finishLabel.appendChild(finishInput);

    var errorEl = document.createElement("div");
    errorEl.className = "onto-gantt-edit-error";

    var footer = document.createElement("div");
    footer.className = "onto-gantt-edit-dialog-footer";
    var cancelBtn = document.createElement("button");
    cancelBtn.type = "button";
    cancelBtn.className = "onto-gantt-edit-btn onto-gantt-edit-btn-secondary";
    cancelBtn.textContent = t("ganttActionCancel");
    var saveBtn = document.createElement("button");
    saveBtn.type = "button";
    saveBtn.className = "onto-gantt-edit-btn onto-gantt-edit-btn-primary";
    saveBtn.textContent = t("ganttActionSave");
    footer.appendChild(cancelBtn);
    footer.appendChild(saveBtn);

    var completionLabel = document.createElement("label");
    completionLabel.className = "onto-gantt-edit-field";
    completionLabel.textContent = t("ganttFieldCompletion");
    var completionInput = document.createElement("input");
    completionInput.type = "number";
    completionInput.min = "0";
    completionInput.max = "100";
    completionInput.step = "1";
    completionInput.inputMode = "numeric";
    completionLabel.appendChild(completionInput);

    // Destructive action, kept apart from the Cancel/Save row: a divider, then
    // an icon-only, right-aligned delete button below it.
    var divider = document.createElement("hr");
    divider.className = "onto-gantt-edit-divider";
    var dangerRow = document.createElement("div");
    dangerRow.className = "onto-gantt-edit-danger-row";
    var deleteBtn = document.createElement("button");
    deleteBtn.type = "button";
    deleteBtn.className = "onto-gantt-edit-delete-btn";
    deleteBtn.setAttribute("aria-label", t("ganttActionDelete"));
    deleteBtn.setAttribute("title", t("ganttActionDelete"));
    deleteBtn.innerHTML = '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="3 6 5 6 21 6"/><path d="M19 6l-1 14a2 2 0 0 1-2 2H8a2 2 0 0 1-2-2L5 6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"/><line x1="10" y1="11" x2="10" y2="17"/><line x1="14" y1="11" x2="14" y2="17"/></svg>';
    dangerRow.appendChild(deleteBtn);

    body.appendChild(wbsLabel);
    body.appendChild(startLabel);
    body.appendChild(finishLabel);
    body.appendChild(completionLabel);
    body.appendChild(errorEl);
    body.appendChild(footer);
    body.appendChild(divider);
    body.appendChild(dangerRow);

    dialog.appendChild(header);
    dialog.appendChild(body);
    document.body.appendChild(dialog);

    function closeDialog() {
      if (typeof dialog.close === "function") dialog.close();
      editDialogCurrentTask = null;
    }
    closeBtn.addEventListener("click", closeDialog);
    cancelBtn.addEventListener("click", closeDialog);
    dialog.addEventListener("cancel", function () { editDialogCurrentTask = null; });
    saveBtn.addEventListener("click", function () { handleSaveEdit(); });
    deleteBtn.addEventListener("click", function () { handleDeleteTask(); });

    editDialogEl = dialog;
    editDialogWbsInput = wbsInput;
    editDialogDeleteBtn = deleteBtn;
    editDialogStartInput = startInput;
    editDialogFinishInput = finishInput;
    editDialogCompletionInput = completionInput;
    editDialogCompletionField = completionLabel;
    editDialogTitleTextEl = titleText;
    editDialogRenameBtn = renameBtn;
    editDialogTitleInput = titleInput;
    editDialogErrorEl = errorEl;
    editDialogSaveBtn = saveBtn;
    return dialog;
  }

  function showEditDialogError(message) {
    editDialogErrorEl.textContent = message;
    editDialogErrorEl.style.display = "block";
  }

  function openEditDialog(task) {
    if (!task) return;
    var dialog = ensureEditDialog();
    editDialogCurrentTask = task;
    // Reset to display mode in case the dialog was left mid-rename from a
    // previously opened task.
    editDialogTitleInput.style.display = "none";
    editDialogTitleTextEl.style.display = "";
    editDialogRenameBtn.style.display = "";
    editDialogTitleTextEl.textContent = task.name || task.identification || t("ganttTaskFallbackName");
    editDialogWbsInput.value = task.identification || "";
    editDialogStartInput.value = rt.formatDate(task.scheduleStart) || "";
    editDialogFinishInput.value = rt.formatDate(task.scheduleFinish) || "";
    // A milestone is a point in time, so a completion percentage is meaningless
    // for it — hide the field rather than offer an edit that has no effect.
    editDialogCompletionField.style.display = task.isMilestone ? "none" : "";
    editDialogCompletionInput.value = (task.completion === null || task.completion === undefined)
      ? "" : String(Math.round(task.completion));
    editDialogErrorEl.textContent = "";
    editDialogErrorEl.style.display = "none";
    editDialogSaveBtn.disabled = false;
    editDialogSaveBtn.textContent = t("ganttActionSave");
    editDialogDeleteBtn.disabled = false;
    if (typeof dialog.showModal === "function") {
      dialog.showModal();
    } else {
      dialog.setAttribute("open", "open");
    }
  }
  // Exported for tests / manual triggering outside a real dblclick.
  rt.openGanttTaskEditDialog = openEditDialog;

  function handleSaveEdit() {
    var task = editDialogCurrentTask;
    if (!task) return;
    var startVal = editDialogStartInput.value;
    var finishVal = editDialogFinishInput.value;
    if (!startVal || !finishVal) {
      showEditDialogError(t("ganttErrBothDates"));
      return;
    }
    var newStart = rt.parseDate(startVal);
    var newFinish = rt.parseDate(finishVal);
    if (!newStart || !newFinish) {
      showEditDialogError(t("ganttErrInvalidDate"));
      return;
    }
    if (newFinish.getTime() < newStart.getTime()) {
      showEditDialogError(t("ganttErrFinishBeforeStart"));
      return;
    }
    var newCompletion = null;
    if (!task.isMilestone) {
      var completionVal = editDialogCompletionInput.value;
      if (completionVal !== "" && completionVal !== null && completionVal !== undefined) {
        var pct = parseFloat(String(completionVal).replace("%", "").replace(",", "."));
        if (isNaN(pct) || pct < 0 || pct > 100) {
          showEditDialogError(t("ganttErrCompletionRange"));
          return;
        }
        newCompletion = Math.max(0, Math.min(100, pct));
      }
    }
    if (typeof rt.saveTaskDates !== "function") {
      showEditDialogError(t("ganttErrSaveUnavailable"));
      return;
    }
    editDialogErrorEl.style.display = "none";
    editDialogSaveBtn.disabled = true;
    editDialogSaveBtn.textContent = t("ganttActionSaving");
    // WBS lives on the IfcTask sheet, dates/completion on IfcTaskTime — two
    // separate workbook writes, so only touch the WBS one when it actually
    // changed.
    var newWbs = editDialogWbsInput.value.trim();
    var pipeline = Promise.resolve();
    if (newWbs !== (task.identification || "") && typeof rt.saveTaskIdentification === "function") {
      pipeline = pipeline.then(function () { return rt.saveTaskIdentification(task, newWbs); });
    }
    pipeline.then(function () {
      return rt.saveTaskDates(task, newStart, newFinish, newCompletion);
    }).then(function () {
      editDialogSaveBtn.disabled = false;
      editDialogSaveBtn.textContent = t("ganttActionSave");
      if (editDialogEl && typeof editDialogEl.close === "function") editDialogEl.close();
      editDialogCurrentTask = null;
      runRender();
      if (typeof rt.runArrows === "function") rt.runArrows();
    }).catch(function (err) {
      editDialogSaveBtn.disabled = false;
      editDialogSaveBtn.textContent = t("ganttActionSave");
      showEditDialogError((err && err.message) ? err.message : String(err));
    });
  }

  function handleDeleteTask() {
    var task = editDialogCurrentTask;
    if (!task) return;
    if (typeof rt.deleteTask !== "function" || typeof rt.refreshFromWorkbook !== "function") {
      showEditDialogError(t("ganttErrDeleteUnavailable"));
      return;
    }
    if (typeof window.confirm === "function" && !window.confirm(t("ganttDeleteConfirm"))) {
      return;
    }
    editDialogErrorEl.style.display = "none";
    editDialogDeleteBtn.disabled = true;
    editDialogSaveBtn.disabled = true;
    Promise.resolve(rt.deleteTask(task)).then(function () {
      return rt.refreshFromWorkbook();
    }).then(function () {
      editDialogDeleteBtn.disabled = false;
      editDialogSaveBtn.disabled = false;
      if (editDialogEl && typeof editDialogEl.close === "function") editDialogEl.close();
      editDialogCurrentTask = null;
    }).catch(function (err) {
      editDialogDeleteBtn.disabled = false;
      editDialogSaveBtn.disabled = false;
      showEditDialogError((err && err.message) ? err.message : String(err));
    });
  }

  // The "+" button in the Page header (between the filter and fullscreen
  // buttons) opens this dialog to register a brand-new task. Unlike the
  // edit/rename dialogs above, creation doesn't patch enrichedTasks by hand
  // afterward -- it just asks rt.refreshFromWorkbook() to re-read the
  // connected folder, the same call the header's own refresh button makes,
  // so the new row reaches the screen through the exact same parse/render
  // path every other task already went through.
  var addTaskDialogEl = null;
  var addTaskWbsInput = null;
  var addTaskNameInput = null;
  var addTaskStartInput = null;
  var addTaskFinishInput = null;
  var addTaskErrorEl = null;
  var addTaskSubmitBtn = null;

  function ensureAddTaskDialog() {
    if (addTaskDialogEl) return addTaskDialogEl;

    var dialog = document.createElement("dialog");
    dialog.className = "onto-gantt-edit-dialog";

    var header = document.createElement("div");
    header.className = "onto-gantt-edit-dialog-header";
    var title = document.createElement("h3");
    title.className = "onto-gantt-edit-dialog-title";
    var titleText = document.createElement("span");
    titleText.className = "onto-gantt-edit-title-text";
    titleText.textContent = t("ganttNewTaskTitle");
    title.appendChild(titleText);
    var closeBtn = document.createElement("button");
    closeBtn.type = "button";
    closeBtn.setAttribute("aria-label", t("ganttActionClose"));
    closeBtn.textContent = "×";
    header.appendChild(title);
    header.appendChild(closeBtn);

    var body = document.createElement("div");
    body.className = "onto-gantt-edit-dialog-body";

    var wbsLabel = document.createElement("label");
    wbsLabel.className = "onto-gantt-edit-field";
    wbsLabel.textContent = t("ganttColumnWbs");
    var wbsInput = document.createElement("input");
    wbsInput.type = "text";
    wbsLabel.appendChild(wbsInput);

    var nameLabel = document.createElement("label");
    nameLabel.className = "onto-gantt-edit-field";
    nameLabel.textContent = t("ganttFieldName");
    var nameInput = document.createElement("input");
    nameInput.type = "text";
    nameInput.placeholder = t("ganttColumnTaskName");
    nameLabel.appendChild(nameInput);

    var startLabel = document.createElement("label");
    startLabel.className = "onto-gantt-edit-field";
    startLabel.textContent = t("ganttFieldStart");
    var startInput = document.createElement("input");
    startInput.type = "date";
    startLabel.appendChild(startInput);

    var finishLabel = document.createElement("label");
    finishLabel.className = "onto-gantt-edit-field";
    finishLabel.textContent = t("ganttFieldFinish");
    var finishInput = document.createElement("input");
    finishInput.type = "date";
    finishLabel.appendChild(finishInput);

    var errorEl = document.createElement("div");
    errorEl.className = "onto-gantt-edit-error";

    var footer = document.createElement("div");
    footer.className = "onto-gantt-edit-dialog-footer";
    var cancelBtn = document.createElement("button");
    cancelBtn.type = "button";
    cancelBtn.className = "onto-gantt-edit-btn onto-gantt-edit-btn-secondary";
    cancelBtn.textContent = t("ganttActionCancel");
    var submitBtn = document.createElement("button");
    submitBtn.type = "button";
    submitBtn.className = "onto-gantt-edit-btn onto-gantt-edit-btn-primary";
    submitBtn.textContent = t("ganttActionAdd");
    footer.appendChild(cancelBtn);
    footer.appendChild(submitBtn);

    body.appendChild(wbsLabel);
    body.appendChild(nameLabel);
    body.appendChild(startLabel);
    body.appendChild(finishLabel);
    body.appendChild(errorEl);
    body.appendChild(footer);

    dialog.appendChild(header);
    dialog.appendChild(body);
    document.body.appendChild(dialog);

    function closeDialog() {
      if (typeof dialog.close === "function") dialog.close();
    }
    closeBtn.addEventListener("click", closeDialog);
    cancelBtn.addEventListener("click", closeDialog);
    nameInput.addEventListener("keydown", function (evt) {
      if (evt.key === "Enter") {
        evt.preventDefault();
        handleCreateTask();
      }
    });
    submitBtn.addEventListener("click", function () { handleCreateTask(); });

    addTaskDialogEl = dialog;
    addTaskWbsInput = wbsInput;
    addTaskNameInput = nameInput;
    addTaskStartInput = startInput;
    addTaskFinishInput = finishInput;
    addTaskErrorEl = errorEl;
    addTaskSubmitBtn = submitBtn;
    return dialog;
  }

  function showAddTaskError(message) {
    addTaskErrorEl.textContent = message;
    addTaskErrorEl.style.display = "block";
  }

  function openAddTaskDialog() {
    var dialog = ensureAddTaskDialog();
    addTaskWbsInput.value = "";
    addTaskNameInput.value = "";
    addTaskStartInput.value = "";
    addTaskFinishInput.value = "";
    addTaskErrorEl.textContent = "";
    addTaskErrorEl.style.display = "none";
    addTaskSubmitBtn.disabled = false;
    addTaskSubmitBtn.textContent = t("ganttActionAdd");
    if (typeof dialog.showModal === "function") {
      dialog.showModal();
    } else {
      dialog.setAttribute("open", "open");
    }
    addTaskNameInput.focus();
  }
  // Exported for tests / manual triggering outside a real click.
  rt.openGanttAddTaskDialog = openAddTaskDialog;

  function handleCreateTask() {
    var name = addTaskNameInput.value.trim();
    if (!name) {
      showAddTaskError(t("ganttErrNameRequired"));
      return;
    }
    var startVal = addTaskStartInput.value;
    var finishVal = addTaskFinishInput.value;
    if (startVal && finishVal) {
      var start = rt.parseDate(startVal);
      var finish = rt.parseDate(finishVal);
      if (start && finish && finish.getTime() < start.getTime()) {
        showAddTaskError(t("ganttErrFinishBeforeStart"));
        return;
      }
    }
    if (typeof rt.createTask !== "function" || typeof rt.refreshFromWorkbook !== "function") {
      showAddTaskError(t("ganttErrCreateUnavailable"));
      return;
    }
    addTaskErrorEl.style.display = "none";
    addTaskSubmitBtn.disabled = true;
    addTaskSubmitBtn.textContent = t("ganttActionAdding");
    Promise.resolve(rt.createTask({
      name: name,
      identification: addTaskWbsInput.value.trim() || undefined,
      scheduleStart: startVal ? startVal + "T00:00:00" : undefined,
      scheduleFinish: finishVal ? finishVal + "T00:00:00" : undefined,
    })).then(function () {
      return rt.refreshFromWorkbook();
    }).then(function () {
      addTaskSubmitBtn.disabled = false;
      addTaskSubmitBtn.textContent = t("ganttActionAdd");
      if (addTaskDialogEl && typeof addTaskDialogEl.close === "function") addTaskDialogEl.close();
    }).catch(function (err) {
      addTaskSubmitBtn.disabled = false;
      addTaskSubmitBtn.textContent = t("ganttActionAdd");
      showAddTaskError((err && err.message) ? err.message : String(err));
    });
  }

  function wireAddTaskButton() {
    var btn = document.querySelector(".gantt-add-task-btn");
    if (!btn || btn.dataset.ontobdcWired === "true") return;
    btn.dataset.ontobdcWired = "true";
    btn.addEventListener("click", openAddTaskDialog);
  }
  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", wireAddTaskButton);
  } else {
    wireAddTaskButton();
  }

  function renderScheduleHeader(schedule) {
    if (!schedule) return;
    var identifier = rt.aliasAwareLiteral(schedule, "Identification")
      || rt.aliasAwareLiteral(schedule, "Identifier")
      || rt.getGlobalId(schedule);
    var name = rt.aliasAwareLiteral(schedule, "Name")
      || rt.aliasAwareLiteral(schedule, "Title")
      || identifier
      || t("breadcrumbSchedule");
    var description = rt.aliasAwareLiteral(schedule, "Description");
    document.title = name || document.title;

    var nameHost = document.querySelector(".name");
    rt.mountInlineEditor(nameHost, {
      value: name,
      required: true,
      editLabel: t("editField", { field: t("ganttFieldName") }),
      saveLabel: t("saveField", { field: t("ganttFieldName") }),
      cancelLabel: t("cancelEdit"),
      requiredMessage: t("fieldRequired", { field: t("ganttFieldName") }),
      emptyLabel: t("emptyValue"),
      onSave: function (value) { return rt.saveScheduleField("Name", value); },
    });
    var identifierHost = document.querySelector(".identifier");
    if (identifierHost) identifierHost.textContent = identifier;

    var fields = document.querySelector(".fields");
    if (!fields) return;
    fields.innerHTML = "";
    var row = document.createElement("div");
    row.className = "field";
    var label = document.createElement("div");
    label.className = "field-label";
    label.textContent = t("description");
    var valueHost = document.createElement("div");
    valueHost.className = "field-value";
    rt.mountInlineEditor(valueHost, {
      value: description,
      multiline: true,
      rows: 3,
      editLabel: t("editField", { field: t("description") }),
      saveLabel: t("saveField", { field: t("description") }),
      cancelLabel: t("cancelEdit"),
      emptyLabel: t("emptyValue"),
      onSave: function (value) { return rt.saveScheduleField("Description", value); },
    });
    row.append(label, valueHost);
    fields.appendChild(row);
  }

  function runRender() {
    var FALLBACK_HEADER_HEIGHT = 72;
    var FALLBACK_ROW_HEIGHT = 34;
    var FALLBACK_TICK_WIDTH_DAY = 28;
    var FALLBACK_MILLIS_PER_DAY = 86400000;
    var _hdr = Number(rt.HEADER_HEIGHT);
    rt.HEADER_HEIGHT = (isFinite(_hdr) && !isNaN(_hdr) && _hdr > 0) ? _hdr : FALLBACK_HEADER_HEIGHT;
    var _row = Number(rt.ROW_HEIGHT);
    rt.ROW_HEIGHT = (isFinite(_row) && !isNaN(_row) && _row > 0) ? _row : FALLBACK_ROW_HEIGHT;
    var _tick = Number(rt.TICK_WIDTH_DAY);
    rt.TICK_WIDTH_DAY = (isFinite(_tick) && !isNaN(_tick) && _tick > 0) ? _tick : FALLBACK_TICK_WIDTH_DAY;
    var _ms = Number(rt.MILLIS_PER_DAY);
    rt.MILLIS_PER_DAY = (isFinite(_ms) && !isNaN(_ms) && _ms > 0) ? _ms : FALLBACK_MILLIS_PER_DAY;

    // Idempotente: esta funcao roda no load e de novo a cada ingestao de
    // planilha. Sem zerar os alvos, a segunda passada desenha por cima da
    // primeira — rotulos de mes sobrepostos e dias repetidos.
    var previousTbody = document.getElementById("gantt-tbody");
    if (previousTbody) previousTbody.textContent = "";
    var previousHeader = document.getElementById("gantt-timeline-header");
    if (previousHeader) previousHeader.textContent = "";
    var previousSvg = document.getElementById("gantt-svg");
    if (previousSvg) previousSvg.textContent = "";

    var enrichedTasks = rt.state.enrichedTasks || [];
    var schedule = rt.state.schedule || null;
    renderScheduleHeader(schedule);

    var allDates = [];
    for (var i = 0; i < enrichedTasks.length; i++) {
      var task = enrichedTasks[i];
      if (task.scheduleStart) allDates.push(task.scheduleStart);
      if (task.scheduleFinish) allDates.push(task.scheduleFinish);
    }
    if (schedule) {
      var sStart = rt.parseDate(rt.literalValue(schedule, "StartTime"));
      var sFinish = rt.parseDate(rt.literalValue(schedule, "FinishTime"));
      if (sStart) allDates.push(sStart);
      if (sFinish) allDates.push(sFinish);
    }

    var minDate = null;
    var maxDate = null;
    if (allDates.length > 0) {
      var times = [];
      for (var j = 0; j < allDates.length; j++) times.push(allDates[j].getTime());
      minDate = new Date(Math.min.apply(null, times));
      maxDate = new Date(Math.max.apply(null, times));
    }
    if (!minDate || !maxDate) {
      var today = new Date();
      minDate = new Date(today.getFullYear(), today.getMonth(), 1);
      maxDate = new Date(today.getFullYear(), today.getMonth() + 1, 0);
    }
    minDate = new Date(minDate.getFullYear(), minDate.getMonth(), minDate.getDate());
    maxDate = new Date(maxDate.getFullYear(), maxDate.getMonth(), maxDate.getDate());

    var totalDays = Math.ceil((maxDate.getTime() - minDate.getTime()) / Math.max(1, numOr(rt.MILLIS_PER_DAY, 86400000))) + 1;
    if (!isFinite(totalDays) || totalDays < 1) totalDays = 31;
    var svgWidth = Math.max(800, numOr(totalDays, 31) * numOr(rt.TICK_WIDTH_DAY, 28));
    // No +HEADER_HEIGHT here: #gantt-timeline-header is a separate sibling
    // <div> (see the .html.j2 template) that already reserves its own
    // 72px in normal block flow above #gantt-svg — adding HEADER_HEIGHT a
    // second time *inside* the svg's own coordinate system just pushed
    // every row down by a redundant extra header's worth of blank space.
    var svgHeight = enrichedTasks.length * numOr(rt.ROW_HEIGHT, 34);
    if (!isFinite(svgWidth) || isNaN(svgWidth) || svgWidth < 0) svgWidth = 800;
    if (!isFinite(svgHeight) || isNaN(svgHeight) || svgHeight < 0) svgHeight = numOr(rt.ROW_HEIGHT, 34);
    rt.state.minDate = minDate;
    rt.state.maxDate = maxDate;
    rt.state.totalDays = totalDays;
    rt.state.svgWidth = svgWidth;
    rt.state.svgHeight = svgHeight;

    function setSvgAttr(el, name, value, fallback) {
      if (!el) return;
      el.setAttribute(name, String(numOr(value, fallback)));
    }

    var tbody = document.getElementById("gantt-tbody");
    if (tbody) {
      for (var k = 0; k < enrichedTasks.length; k++) {
        var task2 = enrichedTasks[k];
        var tr = document.createElement("tr");
        var tdWbs = document.createElement("td");
        tdWbs.className = "col-wbs";
        tdWbs.textContent = task2.identification || "";
        var tdName = document.createElement("td");
        tdName.className = "col-name";
        tdName.title = task2.description || task2.name || "";
        tdName.textContent = task2.name || "";
        var tdStart = document.createElement("td");
        tdStart.className = "col-start";
        tdStart.textContent = rt.formatDate(task2.scheduleStart);
        var tdFinish = document.createElement("td");
        tdFinish.className = "col-finish";
        tdFinish.textContent = rt.formatDate(task2.scheduleFinish);
        var tdDuration = document.createElement("td");
        tdDuration.className = "col-duration";
        var durVal = Math.max(0, task2.duration_days);
        tdDuration.textContent = durVal.toFixed(durVal % 1 === 0 ? 0 : 2) + "d";
        var tdCompletion = document.createElement("td");
        tdCompletion.className = "col-completion";
        tdCompletion.textContent = Math.round(task2.completion) + "%";
        tr.appendChild(tdWbs);
        tr.appendChild(tdName);
        tr.appendChild(tdStart);
        tr.appendChild(tdFinish);
        tr.appendChild(tdDuration);
        tr.appendChild(tdCompletion);
        tr.style.cursor = "pointer";
        (function (taskRef) {
          tr.addEventListener("dblclick", function () { openEditDialog(taskRef); });
        })(task2);
        tbody.appendChild(tr);
      }
    }

    var svg = document.getElementById("gantt-svg");
    var timelineHeader = document.getElementById("gantt-timeline-header");
    if (svg) {
      svg.setAttribute("viewBox", "0 0 " + numOr(svgWidth, 800) + " " + numOr(svgHeight, rt.ROW_HEIGHT));
      setSvgAttr(svg, "width", svgWidth, 800);
      setSvgAttr(svg, "height", svgHeight, rt.ROW_HEIGHT);
    }

    var taskIndexByGlobalId = {};
    for (var m = 0; m < enrichedTasks.length; m++) {
      if (enrichedTasks[m].globalId) {
        taskIndexByGlobalId[enrichedTasks[m].globalId] = m;
      }
    }
    rt.state.taskIndexByGlobalId = taskIndexByGlobalId;

    var taskBarData = [];
    for (var n = 0; n < enrichedTasks.length; n++) {
      var task3 = enrichedTasks[n];
      var rowY = numOr(n * rt.ROW_HEIGHT, 0);
      var startOffsetDays = task3.scheduleStart ? (task3.scheduleStart.getTime() - minDate.getTime()) / rt.MILLIS_PER_DAY : 0;
      var barX = numOr(startOffsetDays * rt.TICK_WIDTH_DAY + 6, 6);
      var durationDays = Math.max(0.25, isFinite(task3.duration_days) ? task3.duration_days : 1);
      var barWidth = Math.max(0, numOr(durationDays * rt.TICK_WIDTH_DAY - 12, rt.TICK_WIDTH_DAY / 2));
      taskBarData.push({
        index: n,
        rowY: rowY,
        barX: barX,
        barWidth: barWidth,
        durationDays: durationDays,
        task: task3,
        isMilestone: task3.isMilestone
      });
    }
    rt.state.taskBarData = taskBarData;

    if (svg) {
      for (var d = 0; d <= totalDays; d++) {
        var x = d * rt.TICK_WIDTH_DAY;
        var isMajor = d % 7 === 0;
        var line = createSvgEl("line", {
          x1: numOr(x, 0),
          y1: 0,
          x2: numOr(x, 0),
          y2: numOr(svgHeight, rt.ROW_HEIGHT),
          class: isMajor ? "gantt-grid-col-major" : "gantt-grid-col"
        });
        svg.appendChild(line);
      }
    }

    if (svg) {
      var todayNow = new Date();
      todayNow.setHours(0, 0, 0, 0);
      var msPerDay2 = 86400000;
      var diffMs2 = todayNow.getTime() - minDate.getTime();
      var todayIndexExact2 = diffMs2 / msPerDay2;
      if (todayIndexExact2 >= -1 && todayIndexExact2 <= (totalDays + 1)) {
        var todayX2 = todayIndexExact2 * rt.TICK_WIDTH_DAY + rt.TICK_WIDTH_DAY / 2;
        var todayTotalH = Math.max(rt.ROW_HEIGHT * Math.max(4, enrichedTasks.length + 2), svgHeight);
        var todayRect2 = createSvgEl("rect", {
          x: numOr(todayX2 - 1, 0),
          y: 0,
          width: 2,
          height: numOr(todayTotalH, rt.ROW_HEIGHT * 4),
          class: "gantt-today-line"
        });
        svg.appendChild(todayRect2);
        if (!window.__ontobdcGanttTodayX) {
          window.__ontobdcGanttTodayX = todayX2;
        }
      }
      for (var p = 0; p <= enrichedTasks.length; p++) {
        var y = numOr(p * rt.ROW_HEIGHT, 0);
        var rowLine = createSvgEl("line", {
          x1: 0,
          y1: numOr(y, 0),
          x2: numOr(svgWidth, 800),
          y2: numOr(y, 0),
          class: "gantt-row-line"
        });
        svg.appendChild(rowLine);
      }
    }

    if (timelineHeader) {
      timelineHeader.innerHTML = "";
      timelineHeader.style.width = numOr(svgWidth, 800) + "px";
      timelineHeader.style.height = numOr(rt.HEADER_HEIGHT, 72) + "px";
      timelineHeader.style.position = "relative";
      timelineHeader.style.overflow = "hidden";

      var pageLang = (window.ontobdcUrlState && window.ontobdcUrlState.lang) ||
                     document.documentElement.lang ||
                     "en";

      function titleCase(s) {
        if (!s) return "";
        return String(s).charAt(0).toUpperCase() + String(s).slice(1);
      }
      var yearFmt = new Intl.DateTimeFormat(pageLang, { year: "numeric" });
      var monthFmtLong = new Intl.DateTimeFormat(pageLang, { month: "long" });

      function cell(yearOrMonthOrDay, colSpan, label) {
        var th = document.createElement("th");
        th.className = "gantt-header-cell gantt-header-" + yearOrMonthOrDay;
        th.setAttribute("colspan", String(colSpan));
        var span = document.createElement("span");
        span.className = "gantt-header-label gantt-header-" + yearOrMonthOrDay + "-label";
        span.textContent = label;
        th.appendChild(span);
        return th;
      }

      var trYear = document.createElement("tr");
      var trMonth = document.createElement("tr");
      var trDay = document.createElement("tr");

      var cur = new Date(minDate.getFullYear(), minDate.getMonth(), minDate.getDate());
      var curYear = cur.getFullYear();
      var curMonth = cur.getMonth();
      var curYearStartIdx = 0;
      var curMonthStartIdx = 0;
      var qFinal = totalDays;

      for (var q = 0; q <= qFinal; q++) {
        var isEnd = q === qFinal;
        var dayNum, monthNum, yearNum;
        if (!isEnd) {
          dayNum = cur.getDate();
          monthNum = cur.getMonth();
          yearNum = cur.getFullYear();
        }

        var flushMonth = isEnd || monthNum !== curMonth;
        var flushYear = isEnd || yearNum !== curYear;

        if (flushMonth && q > curMonthStartIdx) {
          var spanDaysMonth = q - curMonthStartIdx;
          trMonth.appendChild(cell("month", spanDaysMonth, titleCase(monthFmtLong.format(new Date(curYear, curMonth, 1)))));
          curMonthStartIdx = q;
          if (!isEnd) curMonth = monthNum;
        }

        if (flushYear && q > curYearStartIdx) {
          var spanDaysYear = q - curYearStartIdx;
          trYear.appendChild(cell("year", spanDaysYear, titleCase(yearFmt.format(new Date(curYear, 0, 1)))));
          curYearStartIdx = q;
          if (!isEnd) curYear = yearNum;
        }

        if (!isEnd) {
          trDay.appendChild(cell("day", 1, String(dayNum)));
          cur.setDate(cur.getDate() + 1);
        }
      }

      var thead = document.createElement("thead");
      thead.appendChild(trYear);
      thead.appendChild(trMonth);
      thead.appendChild(trDay);
      var tbl = document.createElement("table");
      tbl.className = "gantt-timeline-table";
      tbl.style.width = numOr(svgWidth, 800) + "px";
      tbl.style.height = numOr(rt.HEADER_HEIGHT, 72) + "px";
      tbl.style.tableLayout = "fixed";
      tbl.style.borderCollapse = "collapse";
      tbl.appendChild(thead);
      timelineHeader.appendChild(tbl);
    }

    if (svg) {
      for (var w = 0; w < taskBarData.length; w++) {
        var data = taskBarData[w];
        var rowY2 = data.rowY;
        var barX2 = data.barX;
        var barWidth2 = data.barWidth;
        var isMilestone2 = data.isMilestone;
        var task4 = data.task;
        var centerY = numOr(rowY2, 0) + numOr(rt.ROW_HEIGHT, 34) / 2;
        var centerX = numOr(barX2, 6) + numOr(barWidth2, numOr(rt.TICK_WIDTH_DAY, 28) / 2) / 2;

        if (isMilestone2) {
          var size = 10;
          var pts = [
            [numOr(centerX, 0), numOr(centerY, 0) - size],
            [numOr(centerX, 0) + size, numOr(centerY, 0)],
            [numOr(centerX, 0), numOr(centerY, 0) + size],
            [numOr(centerX, 0) - size, numOr(centerY, 0)]
          ].map(function (pair) { return pair.join(","); }).join(" ");
          var diamond = createSvgEl("polygon", {
            points: pts,
            class: "gantt-marker-diamond"
          });
          svg.appendChild(diamond);
          (function (taskRef) {
            diamond.addEventListener("dblclick", function (evt) {
              if (evt && typeof evt.stopPropagation === "function") evt.stopPropagation();
              openEditDialog(taskRef);
            });
          })(task4);

          if (task4.name && barWidth2 > 20) {
            var textMil = createSvgEl("text", {
              x: numOr(centerX, 0) + size + 4,
              y: numOr(centerY, 0) + 4,
              class: "gantt-bar-text"
            });
            textMil.textContent = task4.name;
            svg.appendChild(textMil);
          }
        } else {
          var barClasses = ["gantt-bar"];
          if (task4.isCritical) barClasses.push("gantt-bar-critical");
          var bar = createSvgEl("rect", {
            x: numOr(barX2, 6),
            y: numOr(rowY2, 0) + 6,
            width: numOr(barWidth2, numOr(rt.TICK_WIDTH_DAY, 28) / 2),
            height: Math.max(0, numOr(rt.ROW_HEIGHT, 34) - 12),
            class: barClasses.join(" ")
          });
          svg.appendChild(bar);
          (function (taskRef) {
            bar.addEventListener("dblclick", function (evt) {
              if (evt && typeof evt.stopPropagation === "function") evt.stopPropagation();
              openEditDialog(taskRef);
            });
          })(task4);

          if (task4.completion > 0) {
            var progressWidth = Math.max(0, (numOr(barWidth2, numOr(rt.TICK_WIDTH_DAY, 28) / 2) * task4.completion) / 100);
            var progress = createSvgEl("rect", {
              x: numOr(barX2, 6),
              y: numOr(rowY2, 0) + 6,
              width: progressWidth,
              height: Math.max(0, numOr(rt.ROW_HEIGHT, 34) - 12),
              class: "gantt-bar-progress"
            });
            svg.appendChild(progress);
          }

          if (task4.name && barWidth2 > 60) {
            var textBar = createSvgEl("text", {
              x: numOr(barX2, 6) + 4,
              y: numOr(rowY2, 0) + 17,
              class: "gantt-bar-text"
            });
            var maxChars = Math.floor(barWidth2 / 6);
            textBar.textContent = task4.name.length > maxChars ? task4.name.substring(0, maxChars - 1) + "\u2026" : task4.name;
            svg.appendChild(textBar);
          }
        }
      }
    }

    var leftPanel = document.querySelector(".gantt-left");
    var rightPanel = document.querySelector(".gantt-right");
    if (leftPanel && rightPanel) {
      var syncing = false;
      leftPanel.addEventListener("scroll", function () {
        if (syncing) return;
        syncing = true;
        rightPanel.scrollTop = leftPanel.scrollTop;
        requestAnimationFrame(function () { syncing = false; });
      });
      rightPanel.addEventListener("scroll", function () {
        if (syncing) return;
        syncing = true;
        leftPanel.scrollTop = rightPanel.scrollTop;
        requestAnimationFrame(function () { syncing = false; });
      });
    }

    if (rightPanel && typeof window.__ontobdcGanttTodayX === "number") {
      if (!window.__ontobdcGanttCenteredOnce) {
        try {
          var todayX = window.__ontobdcGanttTodayX;
          var viewW = rightPanel.clientWidth || rightPanel.offsetWidth || 1000;
          var targetScroll = Math.max(0, todayX - viewW / 2);
          rightPanel.scrollLeft = targetScroll;
        } catch (e) { /* noop */ }
        window.__ontobdcGanttCenteredOnce = true;
      }
    }
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", runRender);
  } else {
    runRender();
  }

  // Exported alongside runParse: connecting a folder or refreshing from the
  // workbook replaces the graph, and the Page has to redraw from it. Without
  // this the new tasks are parsed and never reach the screen.
  rt.runRender = runRender;

  console.log("[gantt-view] state done: task_table_timeline_script_generated");
})(window);
"""

    def _dependency_arrows_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var rt = window.OntoBDCGanttViewRuntime = window.OntoBDCGanttViewRuntime || { state: {} };
  console.log("[gantt-view] state start: dependency_arrows_script_generated");

  function numOr(v, fallback) {
    var n = Number(v);
    if (isFinite(n) && !isNaN(n)) return n;
    var fb = Number(fallback);
    if (isFinite(fb) && !isNaN(fb)) return fb;
    return 0;
  }
  if (!rt.ROW_HEIGHT || !isFinite(Number(rt.ROW_HEIGHT))) rt.ROW_HEIGHT = 34;

  function resolveSequenceGlobalId(ref) {
    if (!ref) return "";
    var refId = rt.referenceId(ref);
    if (!refId) return "";
    var taskIndexByGlobalId = rt.state.taskIndexByGlobalId || {};
    if (taskIndexByGlobalId[refId] !== undefined) return refId;
    var gidKeys = Object.keys(taskIndexByGlobalId);
    for (var i = 0; i < gidKeys.length; i++) {
      if (refId.includes(gidKeys[i]) || refId.endsWith("#" + gidKeys[i]) || refId.endsWith("/" + gidKeys[i])) {
        return gidKeys[i];
      }
    }
    var enrichedTasks = rt.state.enrichedTasks || [];
    for (var j = 0; j < gidKeys.length; j++) {
      var task = enrichedTasks[taskIndexByGlobalId[gidKeys[j]]];
      if (task && task.raw && task.raw["@id"] === refId) {
        return gidKeys[j];
      }
    }
    var gidMatch = refId.match(/([A-Za-z0-9_]{5,})/);
    if (gidMatch && taskIndexByGlobalId[gidMatch[1]] !== undefined) {
      return gidMatch[1];
    }
    return "";
  }
  rt.resolveSequenceGlobalId = resolveSequenceGlobalId;

  function runArrows() {
    // Idem: as setas sao apensadas ao mesmo <svg>, entao uma segunda
    // passada duplicaria cada dependencia. runRender ja zerou o svg quando
    // roda antes desta — este guarda cobre a chamada isolada.
    var drawn = document.querySelectorAll("#gantt-svg .gantt-dependency-line");
    for (var d = 0; d < drawn.length; d++) drawn[d].parentNode.removeChild(drawn[d]);

    var svg = document.getElementById("gantt-svg");
    var rawSequences = rt.state.rawSequences || [];
    var taskIndexByGlobalId = rt.state.taskIndexByGlobalId || {};
    var taskBarData = rt.state.taskBarData || [];
    if (!svg || rawSequences.length === 0) return;

    var rh = Number(rt.ROW_HEIGHT);
    rh = isFinite(rh) && !isNaN(rh) ? rh : 40;

    for (var i = 0; i < rawSequences.length; i++) {
      var seq = rawSequences[i];
      var relatingRef = rt.anyLiteral(seq, ["RelatingProcess", "Predecessor"]) || rt.referenceId(seq["RelatingProcess"] || seq["relatingProcess"] || seq["Predecessor"] || seq["predecessor"] || null);
      var relatedRef = rt.anyLiteral(seq, ["RelatedProcess", "Successor"]) || rt.referenceId(seq["RelatedProcess"] || seq["relatedProcess"] || seq["Successor"] || seq["successor"] || null);
      var predGid = resolveSequenceGlobalId(relatingRef);
      var succGid = resolveSequenceGlobalId(relatedRef);
      if (!predGid || !succGid) continue;
      var predIdx = taskIndexByGlobalId[predGid];
      var succIdx = taskIndexByGlobalId[succGid];
      if (predIdx === undefined || succIdx === undefined) continue;
      var predData = taskBarData[predIdx];
      var succData = taskBarData[succIdx];
      if (!predData || !succData) continue;

      var xPredEnd = numOr(Number(predData.barX) + Number(predData.barWidth), 0);
      var yPred = numOr(Number(predData.rowY) + rh / 2, 0);
      var xSuccStart = numOr(succData.barX, 0);
      var ySucc = numOr(Number(succData.rowY) + rh / 2, 0);

      var gap = 15;
      var pathD;
      if (Math.abs(yPred - ySucc) < 2) {
        pathD = "M " + xPredEnd + " " + yPred + " L " + xSuccStart + " " + ySucc;
      } else {
        var xMid1 = xPredEnd + gap;
        var xMid2 = xSuccStart - 4;
        pathD = "M " + xPredEnd + " " + yPred + " H " + xMid1 + " V " + ySucc + " H " + xMid2;
      }
      var path = rt.createSvgEl("path", {
        d: pathD,
        class: "gantt-dependency-line"
      });
      svg.appendChild(path);
    }
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", runArrows);
  } else {
    runArrows();
  }

  // Exported alongside runParse: connecting a folder or refreshing from the
  // workbook replaces the graph, and the Page has to redraw from it. Without
  // this the new tasks are parsed and never reach the screen.
  rt.runArrows = runArrows;

  console.log("[gantt-view] state done: dependency_arrows_script_generated");
})(window);
"""
