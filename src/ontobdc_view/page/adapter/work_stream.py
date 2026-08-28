from __future__ import annotations

from .container import (
    WORK_STREAM_RUNTIME,
    chrome_controls_source,
    connection_state_source,
    container_connection_source,
)

from typing import Callable, Dict

from ontobdc_view.shared.adapter.vendor import (
    VENDOR_SHEET_JS_NAME,
    vendor_asset_source,
)
from ontobdc_view.shared.domain.port.work_stream_script import WorkStreamScriptPort


class WorkStreamScriptAdapter(WorkStreamScriptPort):
    """Returns the WorkStream Page's runtime JS, split by responsibility.

    Each of the 10 scripts below is one state of
    `WorkStreamScriptGenerationProcessState` (ontobdc-wip). The build-time
    Capability for that state calls `script_source(name)` to get the
    content, then writes it to
    `.__ontobdc__/asset/work_stream_view/<name>.js` inside the container —
    this class only builds text, it never touches a filesystem itself.

    All 10 attach their exports onto a single shared
    `window.OntoBDCWorkStreamViewRuntime` namespace object (functions
    directly, cross-cutting mutable state under `.state`), since none of
    these load as ES modules — matching the plain-`<script>`-tag loading
    every other Tile/Page asset in this project already uses. Each script
    logs a `console.log` line when it starts and finishes attaching its
    exports, so the browser console shows exactly which generation state a
    loaded page is running, at what point.
    """

    def script_source(self, name: str) -> str:
        builder = self._BUILDERS.get(name)
        if builder is None:
            raise ValueError(f"Unknown work stream script name: {name!r}")
        return builder(self)

    def _i18n_apply_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: i18n_script_generated");

  runtime.state = Object.assign(
    {
      graphNodes: [],
      selfNode: null,
      rawContainerHandle: null,
      datasetHandle: null,
      activeMountPath: null,
      liveWorkbookRecord: null,
      liveRelationships: null,
    },
    runtime.state || {},
  );

  runtime.WORK_STREAM_TYPE_NS = "http://datacenter.app.br/ontology/productivity/entity/work_stream/type.ttl#";
  runtime.OBDC_NS = "http://ontobdc.org/ontology/domain/ontobdc/ns.ttl#";
  runtime.DCTERMS_TITLE = "http://purl.org/dc/terms/title";
  runtime.DCTERMS_IDENTIFIER = "http://purl.org/dc/terms/identifier";
  runtime.DCTERMS_DESCRIPTION = "http://purl.org/dc/terms/description";
  runtime.FILE_TYPES = new Set([
    `${runtime.OBDC_NS}GenericFile`,
    `${runtime.OBDC_NS}ImageFile`,
    `${runtime.OBDC_NS}PdfFile`,
    `${runtime.OBDC_NS}CsvFile`,
  ]);
  runtime.WORKSTREAM_PAYLOAD = window.infoBimWorkStreamView || null;

  var I18N = JSON.parse(document.getElementById("ontobdc-i18n")?.textContent || "{}");

  function t(key, vars) {
    var locale = document.documentElement.lang || document.documentElement.dataset.language || "en";
    var table = I18N[locale] || I18N.en || {};
    var text = table[key] ?? key;
    if (vars) {
      for (const [name, value] of Object.entries(vars)) text = text.replaceAll(`{${name}}`, value);
    }
    return text;
  }

  function ensureConnectButtonInnerStatus() {
    const button = document.querySelector(".connect-btn");
    if (!button) return;

    // Inject layout overrides once so the pill-shaped button keeps
    // consistent breathing room between the status sphere, its text and
    // the button border even under theme swaps or font metrics variance.
    if (!document.getElementById("ontobdc-connect-btn-layout-style")) {
      const style = document.createElement("style");
      style.id = "ontobdc-connect-btn-layout-style";
      style.textContent = `
        .connect-btn {
          display: inline-flex !important;
          align-items: center !important;
          justify-content: center !important;
          gap: 8px !important;
          padding-block: 7px !important;
          padding-inline: 14px !important;
          line-height: 1;
        }
        .connect-btn .connection-status {
          inline-size: 9px;
          block-size: 9px;
          flex: none;
        }
        .connect-btn__label {
          display: inline-block;
          flex: none;
          max-inline-size: 32ch;
          overflow: hidden;
          text-overflow: ellipsis;
          white-space: nowrap;
        }
      `;
      document.head.appendChild(style);
    }

    // Move any external connection-status dot (the Jinja template
    // default positioned next to the button) inside the button itself.
    // When there is no external dot we build a fresh inline one so the
    // setConnectionStatus() contract keeps working unchanged.
    const external = document.querySelector(".header-actions > .connection-status");
    const existing = button.querySelector(":scope > .connection-status");
    let status = existing ?? external;
    if (status && external === status) {
      external.remove();
    }
    if (!status) {
      status = document.createElement("span");
      status.className = "connection-status";
      status.setAttribute("role", "status");
      status.setAttribute("aria-hidden", "true");
    }
    if (!existing) {
      status.dataset.status = status.dataset.status || "idle";
      button.prepend(status);
    }

    // Ensure the label wrapper is present and preserves the current
    // visible text (translated label or fallback) while keeping the
    // status dot element untouched during later .textContent swaps.
    let label = button.querySelector(":scope > .connect-btn__label");
    if (!label) {
      const previousText = (button.textContent || "").trim();
      label = document.createElement("span");
      label.className = "connect-btn__label";
      if (previousText) label.textContent = previousText;
      button.appendChild(label);
    }
  }

  function setConnectButtonLabel(labelText) {
    const button = document.querySelector(".connect-btn");
    if (!button) return;
    ensureConnectButtonInnerStatus();
    const label = button.querySelector(":scope > .connect-btn__label");
    if (label) label.textContent = labelText;
  }

  // Declarative translation for chrome that's static Jinja-rendered markup
  // (header/breadcrumb/buttons) rather than JS-built at runtime — walks
  // `data-i18n`/`data-i18n-title`/`data-i18n-aria-label` once per (re)render.
  function applyStaticI18n() {
    ensureConnectButtonInnerStatus();
    for (const el of document.querySelectorAll("[data-i18n]")) {
      if (el.classList.contains("connect-btn")) {
        // The connect-button label is connection-state driven (Connect /
        // Connected / Reconnect …), not static. Re-apply the *connected*
        // label when a folder is already connected so a re-render (e.g. from
        // refreshFromWorkbook) doesn't reset it back to "Connect folder".
        const connected = runtime.state.rawContainerHandle || runtime.state.datasetHandle;
        setConnectButtonLabel(t(connected ? "connectedFolder" : el.dataset.i18n));
      } else {
        el.textContent = t(el.dataset.i18n);
      }
    }
    for (const el of document.querySelectorAll("[data-i18n-title]")) {
      el.title = t(el.dataset.i18nTitle);
    }
    for (const el of document.querySelectorAll("[data-i18n-aria-label]")) {
      el.setAttribute("aria-label", t(el.dataset.i18nAriaLabel));
    }
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", ensureConnectButtonInnerStatus, { once: true });
  } else {
    ensureConnectButtonInnerStatus();
  }

  Object.assign(runtime, {
    t: t,
    applyStaticI18n: applyStaticI18n,
    setConnectButtonLabel: setConnectButtonLabel,
    ensureConnectButtonInnerStatus: ensureConnectButtonInnerStatus,
  });
  console.log("[work-stream-view] state end: i18n_script_generated");
}(window));
"""

    def _graph_reader_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: graph_reader_script_generated");
  var __ontobdcIdentityT = function __ontobdcIdentityT(key, vars) {
    var text = String(key || "");
    if (vars) {
      try {
        Object.keys(vars).forEach(function (k) {
          text = text.split("{" + k + "}").join(String(vars[k]));
        });
      } catch (e) { /* noop */ }
    }
    return text;
  };
  if (typeof runtime.t !== "function") runtime.t = __ontobdcIdentityT;
  var t = (typeof runtime.t === "function") ? runtime.t : __ontobdcIdentityT;


  var OBDC_NS = runtime.OBDC_NS;
  var DCTERMS_TITLE = runtime.DCTERMS_TITLE;
  var DCTERMS_IDENTIFIER = runtime.DCTERMS_IDENTIFIER;
  var DCTERMS_DESCRIPTION = runtime.DCTERMS_DESCRIPTION;
  var FILE_TYPES = runtime.FILE_TYPES;

  function loadGraph() {
    const script = document.getElementById("ontobdc-page-jsonld");
    if (!script) return [];
    try {
      const parsed = JSON.parse(script.textContent);
      return Array.isArray(parsed) ? parsed : [parsed];
    } catch {
      return [];
    }
  }

  // Maps canonical column names (workbook headers / ICDD linkset left-hand
  // side identifiers) onto the property URIs used on this page, so a live
  // workbook record can override a given JSON-LD property. Keys are the
  // exact string names produced by the openpyxl parse script (header row of
  // the WorkStream worksheet, lowercased on lookup); values are the
  // predicate URIs `literal()` normally reads from selfNode.
  const WORKBOOK_COLUMN_TO_PROPERTY = Object.freeze({
    Name: DCTERMS_TITLE,
    Description: DCTERMS_DESCRIPTION,
    What: `${runtime.WORK_STREAM_TYPE_NS}what`,
    Why: `${runtime.WORK_STREAM_TYPE_NS}why`,
    Who: `${runtime.WORK_STREAM_TYPE_NS}who`,
    Where: `${runtime.WORK_STREAM_TYPE_NS}where`,
    When: `${runtime.WORK_STREAM_TYPE_NS}when`,
    How: `${runtime.WORK_STREAM_TYPE_NS}how`,
    HowMuch: `${runtime.WORK_STREAM_TYPE_NS}howMuch`,
    "How much": `${runtime.WORK_STREAM_TYPE_NS}howMuch`,
  });

  function _workbookValueForProperty(propertyUri) {
    if (!runtime.state.liveWorkbookRecord || !propertyUri) return null;
    for (const [column, uri] of Object.entries(WORKBOOK_COLUMN_TO_PROPERTY)) {
      if (uri !== propertyUri) continue;
      const raw = runtime.state.liveWorkbookRecord[column] ?? runtime.state.liveWorkbookRecord[column.toLowerCase()];
      if (raw === null || raw === undefined) continue;
      return String(raw).trim();
    }
    return null;
  }

  function literal(node, property, lang) {
    const live = (node === runtime.state.selfNode || (node && node["@id"] === runtime.state.selfNode?.["@id"]))
      ? _workbookValueForProperty(property)
      : null;
    if (live !== null) return live;
    const values = node?.[property];
    if (!Array.isArray(values) || values.length === 0) return "";
    const localized = lang ? values.find((value) => value["@language"] === lang) : null;
    const picked = localized || values[0];
    return String(picked["@value"] ?? picked["@id"] ?? "").trim();
  }

  function nodeTypes(node) {
    const raw = node?.["@type"];
    if (Array.isArray(raw)) return raw;
    if (typeof raw === "string") return [raw];
    return [];
  }

  function isViewArtifact(filePath) {
    return typeof filePath === "string" && filePath.startsWith(".__ontobdc__/view/");
  }

  function resourceNodes() {
    return runtime.state.graphNodes.filter((node) => {
      const types = nodeTypes(node);
      if (!types.some((type) => FILE_TYPES.has(type))) return false;
      const filePath = literal(node, `${OBDC_NS}filePath`);
      return !isViewArtifact(filePath);
    });
  }

  function resourceLabel(node) {
    return literal(node, DCTERMS_TITLE) || literal(node, `${OBDC_NS}filePath`) || node["@id"];
  }

  function resourceMimeKind(node) {
    const types = nodeTypes(node);
    if (types.includes(`${OBDC_NS}ImageFile`)) return "image";
    if (types.includes(`${OBDC_NS}PdfFile`)) return "pdf";
    if (types.includes(`${OBDC_NS}CsvFile`)) return "csv";
    return "generic";
  }

  function renderHeader() {
    const lang = (document.documentElement.lang || "en").toLowerCase();
    const resourceId = document.querySelector(".onto-page")?.getAttribute("data-ontobdc-resource") || "";

    const identifier = runtime.state.selfNode
      ? literal(runtime.state.selfNode, DCTERMS_IDENTIFIER) || runtime.state.selfNode["@id"] || resourceId
      : resourceId;
    const name = runtime.state.selfNode
      ? literal(runtime.state.selfNode, DCTERMS_TITLE, lang) || literal(runtime.state.selfNode, DCTERMS_TITLE)
      : "";

    document.title = name || identifier || document.title;
    const nameHost = document.querySelector(".name");
    runtime.mountInlineEditor(nameHost, {
      value: name || identifier || t("breadcrumbWorkStream"),
      required: true,
      editLabel: t("editField", { field: t("name") }),
      saveLabel: t("saveField", { field: t("name") }),
      cancelLabel: t("cancelEdit"),
      requiredMessage: t("fieldRequired", { field: t("name") }),
      emptyLabel: t("emptyValue"),
      onSave: (value) => runtime.saveWorkStreamField("Name", value),
    });
    document.querySelector(".identifier").textContent = identifier;

    const description = literal(runtime.state.selfNode, DCTERMS_DESCRIPTION);
    const fieldsContainer = document.querySelector(".fields");
    fieldsContainer.innerHTML = "";
    const row = document.createElement("div");
    row.className = "field";
    const labelEl = document.createElement("div");
    labelEl.className = "field-label";
    labelEl.textContent = t("description");
    const valueEl = document.createElement("div");
    valueEl.className = "field-value";
    runtime.mountInlineEditor(valueEl, {
      value: description,
      multiline: true,
      rows: 3,
      editLabel: t("editField", { field: t("description") }),
      saveLabel: t("saveField", { field: t("description") }),
      cancelLabel: t("cancelEdit"),
      emptyLabel: t("emptyValue"),
      onSave: (value) => runtime.saveWorkStreamField("Description", value),
    });
    row.append(labelEl, valueEl);
    fieldsContainer.appendChild(row);

    return identifier;
  }


  Object.assign(runtime, {
    loadGraph: loadGraph,
    literal: literal,
    nodeTypes: nodeTypes,
    isViewArtifact: isViewArtifact,
    resourceNodes: resourceNodes,
    resourceLabel: resourceLabel,
    resourceMimeKind: resourceMimeKind,
    renderHeader: renderHeader
  });
  console.log("[work-stream-view] state end: graph_reader_script_generated");
}(window));
"""

    def _csv_preview_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: csv_preview_script_generated");
  var __ontobdcIdentityT = function __ontobdcIdentityT(key, vars) {
    var text = String(key || "");
    if (vars) {
      try {
        Object.keys(vars).forEach(function (k) {
          text = text.split("{" + k + "}").join(String(vars[k]));
        });
      } catch (e) { /* noop */ }
    }
    return text;
  };
  if (typeof runtime.t !== "function") runtime.t = __ontobdcIdentityT;
  var t = (typeof runtime.t === "function") ? runtime.t : __ontobdcIdentityT;


  // Minimal RFC4180-ish parser: handles quoted fields, escaped "" and commas
  // inside quotes. Ported from onto-csv-file-tile.js's #parseCsv
  // (component/asset/, the main Surface's CSV Tile) — this page's scripts
  // are loaded as plain non-module tags and there is no shared-module
  // mechanism between page/asset/*.js and component/asset/*.js, so this is
  // a deliberate second copy rather than an import.
  function parseCsv(text) {
    const rows = [];
    let row = [];
    let field = "";
    let inQuotes = false;

    for (let index = 0; index < text.length; index += 1) {
      const char = text[index];
      if (inQuotes) {
        if (char === '"' && text[index + 1] === '"') {
          field += '"';
          index += 1;
        } else if (char === '"') {
          inQuotes = false;
        } else {
          field += char;
        }
        continue;
      }

      if (char === '"') {
        inQuotes = true;
      } else if (char === ",") {
        row.push(field);
        field = "";
      } else if (char === "\n" || char === "\r") {
        if (char === "\r" && text[index + 1] === "\n") index += 1;
        row.push(field);
        field = "";
        rows.push(row);
        row = [];
      } else {
        field += char;
      }
    }
    if (field.length || row.length) {
      row.push(field);
      rows.push(row);
    }
    return rows.filter((cells) => cells.some((cell) => cell.trim() !== ""));
  }

  function renderCsvTable(rows) {
    const wrap = document.createElement("div");
    wrap.className = "csv-table-wrap";

    const scroll = document.createElement("div");
    scroll.className = "csv-table-scroll";
    const [header, ...body] = rows;
    const table = document.createElement("table");
    table.className = "csv-table";
    const thead = document.createElement("thead");
    const headRow = document.createElement("tr");
    for (const cell of header) {
      const th = document.createElement("th");
      th.textContent = cell;
      headRow.appendChild(th);
    }
    thead.appendChild(headRow);
    table.appendChild(thead);

    const tbody = document.createElement("tbody");
    for (const cells of body) {
      const tr = document.createElement("tr");
      for (const cell of cells) {
        const td = document.createElement("td");
        td.textContent = cell;
        tr.appendChild(td);
      }
      tbody.appendChild(tr);
    }
    table.appendChild(tbody);
    scroll.appendChild(table);
    wrap.appendChild(scroll);

    const count = document.createElement("div");
    count.className = "csv-table-count";
    count.textContent = body.length === 1 ? t("row", { count: body.length }) : t("rows", { count: body.length });
    wrap.appendChild(count);

    return wrap;
  }

  function renderCsvFallback(href) {
    const fallback = document.createElement("div");
    fallback.className = "csv-fallback";
    const message = document.createElement("span");
    message.textContent = t("previewUnavailable");
    const link = document.createElement("a");
    link.className = "resource-open-link";
    link.href = href;
    link.target = "_blank";
    link.rel = "noopener";
    link.textContent = t("openFile");
    link.addEventListener("click", (event) => event.stopPropagation());
    fallback.append(message, link);
    return fallback;
  }


  Object.assign(runtime, {
    parseCsv: parseCsv,
    renderCsvTable: renderCsvTable,
    renderCsvFallback: renderCsvFallback
  });
  console.log("[work-stream-view] state end: csv_preview_script_generated");
}(window));
"""

    def _container_connection_source(self) -> str:
        return container_connection_source(WORK_STREAM_RUNTIME)

    def _connection_state_source(self) -> str:
        return connection_state_source(WORK_STREAM_RUNTIME)

    def _annotation_bridge_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: annotation_bridge_script_generated");



  var __ontobdcIdentityT = function __ontobdcIdentityT(key, vars) {
    var text = String(key || "");
    if (vars) {
      try {
        Object.keys(vars).forEach(function (k) {
          text = text.split("{" + k + "}").join(String(vars[k]));
        });
      } catch (e) { /* noop */ }
    }
    return text;
  };
  if (typeof runtime.t !== "function") runtime.t = __ontobdcIdentityT;
  var t = (typeof runtime.t === "function") ? runtime.t : __ontobdcIdentityT;
  var literal = runtime.literal;
  var resourceLabel = runtime.resourceLabel;
  var OBDC_NS = runtime.OBDC_NS;

  let annotationRuntime = null;

  // The FileSystemAccess container handle that the annotation store receives
  // for a WorkStream page is already the *dataset folder* itself (the one
  // whose .__ontobdc__/datapackage.json describes this specific WorkStream
  // resource) — see acquireContainerHandle / resolveContainerHandle in
  // container_connection.js. Therefore:
  //   - metadataDirectory must be the dataset-relative metadata folder
  //     name: ".__ontobdc__" (NOT the dataset folder name repeated again).
  //   - datasetDirectory must be the dataset-relative TTL payload folder:
  //     "payload/triple" (as mandated by the EnrichmentAnnotation schema).
  // Injecting the dataset folder name as metadataDirectory here caused a
  // duplicated sub-path: <dataset>/<dataset>/payload/triple/EnrichmentAnnotation.ttl
  // instead of <dataset>/.__ontobdc__/payload/triple/EnrichmentAnnotation.ttl.
  function ensureAnnotationRuntime() {
    if (annotationRuntime || typeof OntoBDCAnnotations === "undefined") return annotationRuntime;
    annotationRuntime = OntoBDCAnnotations.createRuntime({
      normalizeContext: (raw) => raw,
      visual: { contract: OntoBDCAnnotationVisualContract },
      store: {
        metadataDirectory: ".__ontobdc__",
        datasetDirectory: "payload/triple",
        datasetFileName: "EnrichmentAnnotation.ttl",
      },
    });
    return annotationRuntime;
  }

  function openDialog(title) {
    const dialog = document.createElement("dialog");
    dialog.className = "onto-annotation-dialog";
    const header = document.createElement("div");
    header.className = "onto-annotation-dialog-header";
    const heading = document.createElement("strong");
    heading.textContent = title;
    const closeButton = document.createElement("button");
    closeButton.type = "button";
    closeButton.className = "icon-btn";
    closeButton.textContent = "×";
    closeButton.addEventListener("click", () => dialog.close());
    header.append(heading, closeButton);
    const body = document.createElement("div");
    body.className = "onto-annotation-dialog-body";
    dialog.append(header, body);
    document.body.appendChild(dialog);
    dialog.addEventListener("close", () => dialog.remove());
    dialog.showModal();
    return body;
  }

  async function openWorkspace() {
    const runtimeInstance = ensureAnnotationRuntime();
    if (!runtimeInstance || !runtime.state.datasetHandle) return;
    const body = openDialog(t("annotations"));
    await runtimeInstance.openWorkspace(body, { containerHandle: runtime.state.datasetHandle });
  }

  async function openSubjectPage() {
    const runtimeInstance = ensureAnnotationRuntime();
    if (!runtimeInstance || !runtime.state.datasetHandle) return;
    const body = openDialog(t("subjects"));
    await runtimeInstance.openSubjectPage(body, { containerHandle: runtime.state.datasetHandle }, null);
  }

  // obdc:filePath (data_gathered.py, ontobdc-wip) is written relative to
  // the container root the CLI ran against — which is runtime.state.rawContainerHandle
  // (exactly what the user picked), NOT runtime.state.datasetHandle (the
  // BFS-descended work_stream dataset folder — see acquireContainerHandle()'s
  // comment in container_connection.js). The two are usually the same
  // handle, but differ whenever the user selects a shared parent folder, so
  // this tries the root first and falls back to the dataset folder rather
  // than assuming one convention.
  async function resolveFileUnder(rootHandle, filePath) {
    const segments = filePath.split("/").filter(Boolean);
    let directory = rootHandle;
    let walked = [];
    for (const segment of segments.slice(0, -1)) {
      let entryNames = null;
      try {
        entryNames = [];
        for await (const entry of directory.values()) entryNames.push(entry.name);
      } catch {
        // listing is only for the error message below — never blocks resolution
      }
      try {
        directory = await directory.getDirectoryHandle(segment);
        walked.push(segment);
      } catch (error) {
        error.ontobdcContext = {
          filePath, walked, failedSegment: segment,
          availableAtFailure: entryNames,
        };
        throw error;
      }
    }
    const fileName = segments[segments.length - 1];
    try {
      const fileHandle = await directory.getFileHandle(fileName);
      return fileHandle.getFile();
    } catch (error) {
      let entryNames = null;
      try {
        entryNames = [];
        for await (const entry of directory.values()) entryNames.push(entry.name);
      } catch {
      }
      error.ontobdcContext = {
        filePath, walked, failedSegment: fileName,
        availableAtFailure: entryNames,
      };
      throw error;
    }
  }

  async function resolveFile(filePath) {
    const roots = [];
    if (runtime.state.rawContainerHandle) roots.push(["container root", runtime.state.rawContainerHandle]);
    if (
      runtime.state.datasetHandle
      && runtime.state.datasetHandle !== runtime.state.rawContainerHandle
    ) {
      roots.push(["work_stream dataset folder", runtime.state.datasetHandle]);
    }
    const attempts = [];
    for (const [rootLabel, rootHandle] of roots) {
      try {
        return await resolveFileUnder(rootHandle, filePath);
      } catch (error) {
        if (error && error.name !== "NotFoundError") throw error;
        attempts.push({ root: rootLabel, ...(error.ontobdcContext || {}) });
      }
    }
    const detail = attempts
      .map((a) => `[${a.root}] failed at "${a.failedSegment}" (walked: ${JSON.stringify(a.walked)}; available: ${JSON.stringify(a.availableAtFailure)})`)
      .join(" | ");
    throw new Error(`Could not resolve "${filePath}": ${detail || "no connected folder to search"}`);
  }

  async function annotateResource(node, dimensionUri, button) {
    const runtimeInstance = ensureAnnotationRuntime();
    if (!runtimeInstance || !runtime.state.datasetHandle || !node) return;
    const filePath = literal(node, `${OBDC_NS}filePath`);
    if (!filePath) return;
    const originalTitle = button.title;
    button.disabled = true;
    button.title = t("opening");
    try {
      const file = await resolveFile(filePath);
      await runtimeInstance.open({
        containerHandle: runtime.state.datasetHandle,
        file,
        mediaType: file.type,
        logicalSource: node["@id"],
        representationSource: node["@id"],
        representationName: resourceLabel(node),
        dimension: dimensionUri,
      });
      button.disabled = false;
      button.title = originalTitle;
    } catch (error) {
      console.error(error);
      const message = (error && error.message) || String(error);
      button.title = message;
      const isUnsupported = /annotatable representation/i.test(message);
      if (!isUnsupported) {
        window.alert(`Could not open the annotation editor: ${message}`);
      }
      setTimeout(() => {
        button.disabled = !runtime.state.datasetHandle;
        button.title = originalTitle;
      }, 3000);
    }
  }

  function wireAnnotationControls() {
    // The Subjects/Threads and Workspace buttons are wired exactly once by
    // container.py's wireChromeControls() (runs on DOMContentLoaded and is
    // idempotent). Doing it again here is what made the Threads dialog open
    // twice per click, so this render step now only guarantees that wiring
    // has happened rather than adding its own duplicate listeners.
    if (typeof runtime.wireChromeControls === "function") {
      runtime.wireChromeControls();
    }
  }


  function listAnnotations() {
    const instance = ensureAnnotationRuntime();
    if (!instance || typeof instance.listAnnotations !== "function") return [];
    try {
      return instance.listAnnotations() || [];
    } catch {
      return [];
    }
  }

  Object.assign(runtime, {
    ensureAnnotationRuntime: ensureAnnotationRuntime,
    listAnnotations: listAnnotations,
    openDialog: openDialog,
    openWorkspace: openWorkspace,
    openSubjectPage: openSubjectPage,
    resolveFile: resolveFile,
    annotateResource: annotateResource,
    wireAnnotationControls: wireAnnotationControls
  });
  console.log("[work-stream-view] state end: annotation_bridge_script_generated");
}(window));
"""

    def _pyodide_runtime_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: pyodide_runtime_script_generated");



  var __ontobdcIdentityT = function __ontobdcIdentityT(key, vars) {
    var text = String(key || "");
    if (vars) {
      try {
        Object.keys(vars).forEach(function (k) {
          text = text.split("{" + k + "}").join(String(vars[k]));
        });
      } catch (e) { /* noop */ }
    }
    return text;
  };
  if (typeof runtime.t !== "function") runtime.t = __ontobdcIdentityT;
  var t = (typeof runtime.t === "function") ? runtime.t : __ontobdcIdentityT;
  var WORKSTREAM_PAYLOAD = runtime.WORKSTREAM_PAYLOAD;
  var resolveContainerHandle = runtime.resolveContainerHandle;
  var renderHeader = runtime.renderHeader;
  var applyStaticI18n = runtime.applyStaticI18n;

  const PYODIDE_CDN_URL = "https://cdn.jsdelivr.net/pyodide/v0.27.2/full/pyodide.js";

  let pyodideInstance = null;
  let pyodideLoadPromise = null;
  let pyodideHasOpenpyxl = false;

  // pyodide.globals is ONE shared namespace across every caller — there is
  // no per-call isolation. openContainer()/openContainerFromHandle()
  // (this file), runLinksetOperation() (linkset_operations.js) and
  // loadFileDisplayProfiles() (file_category.js) all do
  // `pyodide.globals.set(...)` immediately followed by
  // `await pyodide.runPythonAsync(...)`. With up to 7 dimension cards each
  // firing related+suggested linkset reads concurrently on connect
  // (notifyDimensionCardsConnected() never awaits between cards), one
  // call's globals.set() can land in the gap between another call's own
  // globals.set() and its runPythonAsync() actually executing — so that
  // second call's Python script silently reads the FIRST call's globals
  // instead of its own. Every pyodide.globals.set()+runPythonAsync() pair
  // in this codebase must go through withPyodideLock() so only one such
  // critical section ever runs at a time.
  let pyodideOperationQueue = Promise.resolve();

  function withPyodideLock(taskFn) {
    const settled = pyodideOperationQueue.then(taskFn, taskFn);
    pyodideOperationQueue = settled.then(() => {}, () => {});
    return settled;
  }

  function loadScriptTag(src) {
    return new Promise((resolve, reject) => {
      const script = document.createElement("script");
      script.src = src;
      script.onload = resolve;
      script.onerror = () => reject(new Error(`Failed to load ${src}`));
      document.head.appendChild(script);
    });
  }

  async function ensurePyodide(options) {
    const withOpenpyxl = Boolean(options && options.withOpenpyxl);
    if (pyodideInstance && (!withOpenpyxl || pyodideHasOpenpyxl)) {
      return pyodideInstance;
    }
    if (!pyodideLoadPromise) {
      pyodideLoadPromise = (async () => {
        const isFileProtocol =
          typeof location !== "undefined" &&
          location &&
          typeof location.protocol === "string" &&
          location.protocol === "file:";
        if (typeof loadPyodide !== "function") {
          if (isFileProtocol) {
            throw new Error(
              "Pyodide unavailable on file:// protocol. If SheetJS " +
              "workbook parsing also fails, open this page via HTTP " +
              "(python -m http.server 8765) instead."
            );
          }
          await loadScriptTag(PYODIDE_CDN_URL);
        }
        const instance = await loadPyodide();
        await instance.loadPackage("micropip");
        await instance.runPythonAsync("import micropip\nawait micropip.install('rdflib')");
        return instance;
      })();
    }
    try {
      pyodideInstance = await pyodideLoadPromise;
    } catch (error) {
      pyodideLoadPromise = null;
      throw error;
    }
    if (withOpenpyxl && !pyodideHasOpenpyxl) {
      await pyodideInstance.runPythonAsync("import micropip\nawait micropip.install('openpyxl')");
      pyodideHasOpenpyxl = true;
    }
    return pyodideInstance;
  }

  // Matches the canonical reference runtime (techcenter-doc workstream_5w2h.js
  // openContainer() Python block). Reads the dataset's `datapackage.json`,
  // locates the "work_stream" resource Excel file, parses the worksheet with
  // openpyxl, finds the row whose GlobalId matches the current elementId,
  // then re-parses the WorkStream.ttl ICDD linkset (column <-> property URI
  // mapping) and the WorkStreamResource.ttl dimension <-> resource relation
  // linkset, plus the RO-Crate manifest for file display categories. Returns
  // exactly the same shape as the reference runtime so callers downstream
  // don't have to branch.
  const WORKBOOK_PARSE_SCRIPT = `
import json
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook
from rdflib import Graph, Literal, Namespace, URIRef
from rdflib.namespace import RDF

try:
    import ontobdc_trace as _trace
    _trace_p = lambda msg: _trace.log(str(msg))
except Exception:
    _trace_p = lambda msg: None

def _resolve_candidates(payload_key: str, *extra_candidates):
    raw = payload.get(payload_key) or ""
    candidates = []
    if raw:
        candidates.append(container / raw)
    for entry in extra_candidates:
        if isinstance(entry, str):
            candidate_path = container / entry
        else:
            candidate_path = Path(entry)
        if not any(str(c) == str(candidate_path) for c in candidates):
            candidates.append(candidate_path)
    chosen = None
    for candidate in candidates:
        try:
            exists_p = candidate.exists()
        except Exception:
            exists_p = False
        _trace_p("CANDIDATE_" + payload_key + "=" + str(candidate) + " EX=" + str(exists_p))
        if exists_p and chosen is None:
            chosen = candidate
    if chosen is None and candidates:
        chosen = candidates[0]
    return chosen

payload = json.loads(view_payload_json)
container = Path(container_mount_path)
_trace_p("CONTAINER=" + str(container) + " EX=" + str(container.exists()))

datapackage_candidates = [
    payload.get("datapackagePath") or "",
    ".__ontobdc__/datapackage.json",
    "__ontobdc__/datapackage.json",
    "datapackage.json",
]
_dp_seen = set()
_dp_ordered = []
for _c in datapackage_candidates:
    _norm = str(container / _c) if _c and not str(_c).startswith(str(container)) else str(_c)
    if not _c or _norm in _dp_seen:
        continue
    _dp_seen.add(_norm)
    _dp_ordered.append(_c)
datapackage_path = _resolve_candidates("datapackagePath", *_dp_ordered)
_trace_p("TRY_datapackage_path=" + str(datapackage_path) + " EX=" + str(datapackage_path.exists()))
datapackage_raw = datapackage_path.read_text(encoding="utf-8")
datapackage = json.loads(datapackage_raw)
_trace_p(
    "DATAPACKAGE_RESOURCES_PREVIEW="
    + json.dumps(datapackage.get("resources", []), ensure_ascii=False)[:1000]
)
resource = next(
    (item for item in datapackage.get("resources", []) if item.get("name") == work_stream_resource_name),
    None,
)
if not resource:
    raise ValueError(
        "datapackage.json does not contain a resource named \\"" + work_stream_resource_name + "\\".\\n"
        + "datapackage_path=" + str(datapackage_path) + "\\n"
        + "container_root_children="
        + json.dumps(
            [p.name for p in container.iterdir()][:50],
            ensure_ascii=False,
        )[:800]
        + "\\n"
        + "resources="
        + json.dumps(datapackage.get("resources", []), ensure_ascii=False)[:1200]
    )
workbook_path = (datapackage_path.parent / resource["path"]).resolve()
_trace_p("WORKBOOK_PATH=" + str(workbook_path) + " EX=" + str(workbook_path.exists()))
try:
    workbook_relative_path = workbook_path.relative_to(container).as_posix()
except ValueError:
    workbook_relative_path = ""
worksheet_name = (
    resource.get("dialect", {})
    .get("excel", {})
    .get("sheet", "WorkStream")
)

entity_id = resource.get("entityIdentifier") or work_stream_resource_name
linkset_dir = datapackage_path.parent / "linkset"
facade_candidates = [
    payload.get("linksetPath") or "",
    f".__ontobdc__/linkset/WorkStream.ttl",
    f"{linkset_dir.name}/{entity_id}_linkset.ttl",
    f"{linkset_dir.name}/WorkStream.ttl",
    f"{linkset_dir.name}/facade.ttl",
]
# Dedupe while preserving order
_facade_seen = set()
_facade_ordered = []
for _c in facade_candidates:
    _norm = str(container / _c) if _c and not str(_c).startswith(str(container)) else str(_c)
    if not _c or _norm in _facade_seen:
        continue
    _facade_seen.add(_norm)
    _facade_ordered.append(_c)
linkset_path = _resolve_candidates("linksetPath", *_facade_ordered)

resource_candidates = [
    payload.get("resourceLinksetPath") or "",
    ".__ontobdc__/linkset/WorkStreamResource.ttl",
    f"{linkset_dir.name}/WorkStreamResource.ttl",
    f"{linkset_dir.name}/resource_linkset.ttl",
]
_resource_seen = set()
_resource_ordered = []
for _c in resource_candidates:
    _norm = str(container / _c) if _c and not str(_c).startswith(str(container)) else str(_c)
    if not _c or _norm in _resource_seen:
        continue
    _resource_seen.add(_norm)
    _resource_ordered.append(_c)
resource_linkset_path = _resolve_candidates("resourceLinksetPath", *_resource_ordered)

# RO-Crate: dataset level first; if absent, container level (dataset folder's
# sibling .__ontobdc__) because many generated containers ship a single shared
# ro-crate-metadata.json at the container root, not duplicated per dataset.
ro_crate_candidates = [
    payload.get("roCratePath") or "",
    ".__ontobdc__/ro-crate-metadata.json",
]
parent_meta = (container.parent / ".__ontobdc__" / "ro-crate-metadata.json")
try:
    if parent_meta.exists():
        ro_crate_candidates.append(str(parent_meta))
except Exception:
    pass
ro_crate_path = _resolve_candidates("roCratePath", *ro_crate_candidates)

# File display ontology: dataset level first, then the shared container-level
# asset path. The UI's own FILE_DISPLAY_PARSE_SCRIPT also has candidate paths
# file.ttl / file_display.ttl, so we keep the same convention here.
file_display_candidates = [
    payload.get("fileDisplayOntologyPath") or "",
    ".__ontobdc__/ontology/file_display.ttl",
    ".__ontobdc__/ontology/file.ttl",
]
file_parent_candidates = [
    container.parent / ".__ontobdc__" / "ontology" / "file_display.ttl",
    container.parent / ".__ontobdc__" / "asset" / "infobim-view" / "ontology" / "file_display.ttl",
    container.parent / ".__ontobdc__" / "ontology" / "file.ttl",
    container.parent / ".__ontobdc__" / "asset" / "infobim-view" / "ontology" / "file.ttl",
]
for _c in file_parent_candidates:
    try:
        if _c.exists():
            file_display_candidates.append(str(_c))
    except Exception:
        pass
file_display_ontology_path = _resolve_candidates("fileDisplayOntologyPath", *file_display_candidates)

dataset_prefix = str(payload.get("datasetPath") or "").strip("/")

LS = Namespace("https://standards.iso.org/iso/21597/-1/ed-1/en/Linkset#")
SCHEMA = Namespace("http://schema.org/")
linkset = Graph()
mappings = {}
_trace_p("USING_linkset_path=" + str(linkset_path) + " EX=" + str(linkset_path.exists() if linkset_path is not None else "none"))
icdd_mode = False
if linkset_path is not None and linkset_path.exists():
    linkset.parse(linkset_path, format="turtle")
    icdd_links = list(linkset.subjects(RDF.type, LS.DirectedBinaryLink))
    _trace_p("ICDD_links=" + str(len(icdd_links)))
    if icdd_links:
        icdd_mode = True
        for link in icdd_links:
            from_element = linkset.value(link, LS.hasFromLinkElement)
            to_element = linkset.value(link, LS.hasToLinkElement)
            from_identifier = linkset.value(from_element, LS.hasIdentifier)
            to_identifier = linkset.value(to_element, LS.hasIdentifier)
            column = linkset.value(from_identifier, LS.identifier)
            facade_field = linkset.value(to_identifier, LS.uri)
            if isinstance(column, Literal) and facade_field is not None:
                mappings[str(column)] = str(facade_field)
if not icdd_mode and linkset_path is not None and linkset_path.exists():
    # No ICDD DirectedBinaryLink found. This dataset ships a facade.ttl
    # (schema:identifier based FacadeField registry) instead. Derive the
    # same column->uri mappings from those so downstream callers get an
    # identical mappings dict shape either way.
    FACADE = Namespace(str(linkset.identifier) + "#") if linkset.identifier else None
    facade_field_type = URIRef("http://datacenter.app.br/ontology/productivity/entity/work_stream/facade.ttl#FacadeField")
    for field in linkset.subjects(RDF.type, facade_field_type):
        label_bearer = linkset.value(field, SCHEMA.identifier)
        maps_to = linkset.value(field, URIRef("http://datacenter.app.br/ontology/productivity/entity/work_stream/facade.ttl#mapsToProperty"))
        if label_bearer is not None and maps_to is not None:
            mappings[str(label_bearer)] = str(maps_to)
    _trace_p("FACADE_mappings_count=" + str(len(mappings)))

workbook = load_workbook(workbook_path, data_only=True, read_only=True)
worksheet = workbook[worksheet_name]
headers = [str(cell.value or "").strip() for cell in worksheet[1]]
record = None
for values in worksheet.iter_rows(min_row=2, values_only=True):
    candidate = dict(zip(headers, values))
    if str(candidate.get("GlobalId") or "").strip() == payload["elementId"]:
        record = candidate
        break
workbook.close()

if record is None:
    raise ValueError(
        f"WorkStream not found in workbook: {payload['elementId']}"
    )
# Only enforce ICDD-linkset header coverage when the container actually
# shipped an ICDD DirectedBinaryLink linkset. Facade.ttl-only containers
# (and techcenter-doc style legacy containers where mappings is empty by
# design) proceed with header -> WORKBOOK_COLUMN_TO_PROPERTY resolution.
if icdd_mode and mappings and any(field not in mappings for field in headers):
    missing = sorted(field for field in headers if field not in mappings)
    raise ValueError(
        "The ICDD linkset does not map workbook fields: " + ", ".join(missing)
    )

def values(item, key):
    value = item.get(key)
    if value is None:
        return []
    return value if isinstance(value, list) else [value]

def text_values(item, *keys):
    result = []
    for key in keys:
        for value in values(item, key):
            if isinstance(value, dict):
                value = value.get("@id") or value.get("@value")
            if value is not None:
                result.append(str(value))
    return result

FILE_DISPLAY = Namespace(
    "https://w3id.org/ontobdc/ontology/file-display#"
)
file_display_graph = Graph()
_trace_p("USING_file_display_ontology=" + str(file_display_ontology_path) + " EX=" + str(file_display_ontology_path.exists() if file_display_ontology_path is not None else "none"))
if file_display_ontology_path is not None and file_display_ontology_path.exists():
    file_display_graph.parse(file_display_ontology_path, format="turtle")
display_profiles = []
for profile in file_display_graph.subjects(
    RDF.type,
    FILE_DISPLAY.FileDisplayProfile,
):
    category = file_display_graph.value(
        profile,
        FILE_DISPLAY.displayCategory,
    )
    if category is None:
        continue
    display_profiles.append({
        "category": str(category),
        "requiredSemanticTypes": {
            str(value).strip()
            for value in file_display_graph.objects(
                profile,
                FILE_DISPLAY.requiredSemanticType,
            )
            if str(value).strip()
        },
        "mimeTypes": {
            str(value).strip().lower()
            for value in file_display_graph.objects(
                profile,
                FILE_DISPLAY.acceptedMimeType,
            )
            if str(value).strip()
        },
        "extensions": {
            str(value).strip().lower()
            for value in file_display_graph.objects(
                profile,
                FILE_DISPLAY.acceptedExtension,
            )
            if str(value).strip()
        },
    })

def category_for(item, resource_id):
    formats = {
        value.strip().lower()
        for value in text_values(item, "encodingFormat")
        if value.strip()
    }
    semantic_types = {
        value.strip()
        for value in text_values(item, "@type", "additionalType")
        if value.strip()
    }
    extension = Path(unquote(resource_id)).suffix.lower()

    def semantic_name(value):
        return value.rsplit("#", 1)[-1].rsplit("/", 1)[-1]

    semantic_names = {semantic_name(value) for value in semantic_types}
    for profile in display_profiles:
        required_types = profile["requiredSemanticTypes"]
        required_names = {semantic_name(value) for value in required_types}
        if required_types and not (
            semantic_types.intersection(required_types)
            or semantic_names.intersection(required_names)
        ):
            continue
        if (
            formats.intersection(profile["mimeTypes"])
            or extension in profile["extensions"]
        ):
            return profile["category"]
    return None

_trace_p("USING_ro_crate_path=" + str(ro_crate_path) + " EX=" + str(ro_crate_path.exists() if ro_crate_path is not None else "none"))
try:
    if ro_crate_path is None or not ro_crate_path.exists():
        ro_crate = {"@graph": []}
    else:
        ro_crate_text = ro_crate_path.read_text(encoding="utf-8").strip()
        ro_crate = json.loads(ro_crate_text) if ro_crate_text else {"@graph": []}
except FileNotFoundError:
    ro_crate = {"@graph": []}

catalog_resources = []
resources = []
for item in ro_crate.get("@graph", []):
    source_resource_id = item.get("@id")
    types = set(text_values(item, "@type"))
    if not source_resource_id or source_resource_id in (".", "./"):
        continue
    is_external_resource = (
        "://" in source_resource_id
        or source_resource_id.startswith("urn:")
    )
    resource_id = (
        source_resource_id
        if is_external_resource or not dataset_prefix
        else f"{dataset_prefix}/{source_resource_id.lstrip('./')}"
    )
    if not types.intersection({
        "File", "MediaObject", "DigitalDocument",
        "CreativeWork", "Message", "EmailMessage",
    }):
        continue
    names = text_values(item, "name")
    formats = text_values(item, "encodingFormat")
    category = category_for(item, source_resource_id)
    display_parts = [
        unquote(part)
        for part in resource_id.split("/")
        if part
    ]
    catalog_resource = {
        "id": resource_id,
        "sourceId": source_resource_id,
        "name": (
            unquote(names[0])
            if names
            else unquote(Path(resource_id).name)
        ),
        "displayParts": display_parts,
        "encodingFormat": formats[0] if formats else "",
    }
    catalog_resources.append(catalog_resource)
    if category is not None:
        resources.append({
            **catalog_resource,
            "category": category,
        })

relationships = {}
_trace_p("USING_resource_linkset=" + str(resource_linkset_path) + " EX=" + str(resource_linkset_path.exists() if resource_linkset_path is not None else "none"))
if resource_linkset_path is not None and resource_linkset_path.exists():
    resource_linkset = Graph()
    resource_linkset.parse(resource_linkset_path, format="turtle")
    resource_id_by_endpoint = {
        key: item["id"]
        for item in resources
        for key in (item["id"], item["sourceId"])
    }

    def endpoint_value(element):
        identifier = resource_linkset.value(element, LS.hasIdentifier)
        return (
            resource_linkset.value(identifier, LS.uri)
            or resource_linkset.value(identifier, LS.identifier)
        )

    for link in resource_linkset.subjects(RDF.type, LS.DirectedBinaryLink):
        endpoints = [
            endpoint_value(resource_linkset.value(link, LS.hasFromLinkElement)),
            endpoint_value(resource_linkset.value(link, LS.hasToLinkElement)),
        ]
        endpoint_strings = [str(value) for value in endpoints if value is not None]
        dimension_uri = next(
            (
                value for value in endpoint_strings
                if value.startswith(payload["dimensionBaseUri"] + "/")
            ),
            None,
        )
        resource_id = next(
            (
                resource_id_by_endpoint[value]
                for value in endpoint_strings
                if value in resource_id_by_endpoint
            ),
            None,
        )
        if dimension_uri and resource_id:
            relationships.setdefault(dimension_uri, []).append(resource_id)

json.dumps(
    {
        "record": record,
        "mappings": mappings,
        "resources": resources,
        "catalogResources": catalog_resources,
        "relationships": relationships,
        "workbookPath": str(workbook_path),
        "workbookRelativePath": workbook_relative_path,
        "worksheetName": worksheet_name,
        "linksetPath": str(linkset_path),
    },
    ensure_ascii=False,
)
`;

  const WORKBOOK_WRITE_SCRIPT = `
import json
from openpyxl import load_workbook

columns = json.loads(work_stream_column_candidates_json)
workbook = load_workbook(workbook_path)
try:
    if worksheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet {worksheet_name} was not found")
    sheet = workbook[worksheet_name]
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }
    global_id_column = headers.get("GlobalId")
    if global_id_column is None:
        raise ValueError("WorkStream worksheet does not have a GlobalId column")
    value_column = next((headers[name] for name in columns if name in headers), None)
    if value_column is None:
        value_column = sheet.max_column + 1
        sheet.cell(row=1, column=value_column, value=columns[0])
    target_row = next(
        (
            row
            for row in range(2, sheet.max_row + 1)
            if str(sheet.cell(row=row, column=global_id_column).value or "").strip()
            == str(work_stream_global_id).strip()
        ),
        None,
    )
    if target_row is None:
        raise ValueError(f"WorkStream row not found (GlobalId {work_stream_global_id})")
    sheet.cell(row=target_row, column=value_column, value=work_stream_value)
    workbook.save(workbook_path)
finally:
    workbook.close()
json.dumps({"ok": True})
`;

  function syncMountedFilesystem(pyodide) {
    return new Promise((resolve, reject) => {
      pyodide.FS.syncfs(false, (error) => error ? reject(error) : resolve());
    });
  }

  // ── SheetJS-native workbook access ──────────────────────────────────────
  // Read and write the WorkStream .xlsx straight in the browser with the
  // vendored SheetJS build, so connecting and saving work on file:// too
  // (Pyodide never boots offline — no CDN, no vendored wheels). The Python
  // WORKBOOK_PARSE_SCRIPT / WORKBOOK_WRITE_SCRIPT stay as the fallback.
  function hasSheetJS() {
    return typeof window.XLSX !== "undefined"
      && !!window.XLSX
      && typeof window.XLSX.read === "function";
  }

  function _wsNavigateSegments(baseDir, relPath) {
    const stack = String(baseDir || "").split("/").filter(Boolean);
    String(relPath || "").split("/").forEach((segment) => {
      if (!segment || segment === ".") return;
      if (segment === "..") stack.pop();
      else stack.push(segment);
    });
    return stack;
  }

  async function _wsFileHandleBySegments(rootHandle, segments) {
    let dir = rootHandle;
    for (let i = 0; i < segments.length - 1; i++) {
      dir = await dir.getDirectoryHandle(segments[i]);
    }
    return dir.getFileHandle(segments[segments.length - 1]);
  }

  async function sheetJsRefreshFromWorkbook() {
    if (!hasSheetJS()) throw new Error("SheetJS is not available on this page.");
    const payload = WORKSTREAM_PAYLOAD || runtime.WORKSTREAM_PAYLOAD || window.infoBimWorkStreamView;
    if (!payload) throw new Error("No WorkStream context payload present on this page.");
    const elementId = String(payload.elementId || "").trim();
    const datasetHandle = runtime.state.datasetHandle || runtime.state.rawContainerHandle;
    if (!datasetHandle) throw new Error(t("connectFolderFirst"));

    const dpSegments = String(payload.datapackagePath || ".__ontobdc__/datapackage.json")
      .split("/").filter(Boolean);
    const dpFile = await _wsFileHandleBySegments(datasetHandle, dpSegments);
    const datapackage = JSON.parse(await (await dpFile.getFile()).text());
    const resources = Array.isArray(datapackage.resources) ? datapackage.resources : [];
    const resourceName = runtime.WORK_STREAM_RESOURCE_NAME || "work_stream";
    const resource = resources.find((r) => r && r.name === resourceName)
      || resources.find((r) => r && typeof r.path === "string" && /\.xlsx$/i.test(r.path));
    if (!resource || !resource.path) {
      throw new Error('datapackage.json has no "' + resourceName + '" resource.');
    }
    let sheetName = (((resource.dialect || {}).excel || {}).sheet) || "WorkStream";

    const dpDir = dpSegments.slice(0, -1).join("/");
    const xlsxSegments = _wsNavigateSegments(dpDir, resource.path);
    const xlsxHandle = await _wsFileHandleBySegments(datasetHandle, xlsxSegments);
    const buffer = await (await xlsxHandle.getFile()).arrayBuffer();
    const workbook = window.XLSX.read(buffer, { type: "array" });
    if (workbook.SheetNames.indexOf(sheetName) === -1) sheetName = workbook.SheetNames[0];
    const rows = window.XLSX.utils.sheet_to_json(workbook.Sheets[sheetName], { defval: "" });
    const record = rows.find(
      (row) => String(row.GlobalId || "").trim() === elementId
    ) || null;

    const relPath = xlsxSegments.join("/");
    runtime.state.workbookFileHandle = xlsxHandle;
    runtime.state.workbookSheetJs = workbook;
    runtime.state.workbookRelPath = relPath;
    runtime.state.liveWorkbookPath = relPath;
    runtime.state.liveWorksheetName = sheetName;

    return {
      nodes: [{
        "@id": String(payload.workstreamUri || elementId),
        "@type": [(runtime.WORK_STREAM_TYPE_NS || "") + "WorkStream"],
      }],
      record: record,
      workbookPath: relPath,
      workbookRelativePath: relPath,
      worksheetName: sheetName,
      relationships: null,
    };
  }

  async function _sheetJsWriteWorkStreamCell(columnCandidates, value, globalId) {
    const XLSX = window.XLSX;
    const wb = runtime.state.workbookSheetJs;
    const fileHandle = runtime.state.workbookFileHandle;
    let sheetName = runtime.state.liveWorksheetName || "WorkStream";
    if (wb.SheetNames.indexOf(sheetName) === -1) sheetName = wb.SheetNames[0];
    const ws = wb.Sheets[sheetName];
    const range = XLSX.utils.decode_range(ws["!ref"]);
    const headerRow = range.s.r;
    const colByName = {};
    for (let c = range.s.c; c <= range.e.c; c++) {
      const hc = ws[XLSX.utils.encode_cell({ r: headerRow, c: c })];
      const name = (hc && hc.v != null) ? String(hc.v).trim() : "";
      if (name) colByName[name] = c;
    }
    const gidCol = colByName["GlobalId"];
    if (gidCol === undefined) throw new Error("WorkStream sheet has no GlobalId column.");
    let valueCol;
    for (let i = 0; i < columnCandidates.length; i++) {
      if (colByName[columnCandidates[i]] !== undefined) { valueCol = colByName[columnCandidates[i]]; break; }
    }
    if (valueCol === undefined) {
      valueCol = range.e.c + 1;
      ws[XLSX.utils.encode_cell({ r: headerRow, c: valueCol })] = { t: "s", v: columnCandidates[0] };
      range.e.c = valueCol;
      ws["!ref"] = XLSX.utils.encode_range(range);
    }
    let targetRow = -1;
    for (let r = headerRow + 1; r <= range.e.r; r++) {
      const gc = ws[XLSX.utils.encode_cell({ r: r, c: gidCol })];
      if (gc && String(gc.v).trim() === globalId) { targetRow = r; break; }
    }
    if (targetRow === -1) throw new Error("WorkStream row not found (GlobalId " + globalId + ").");
    ws[XLSX.utils.encode_cell({ r: targetRow, c: valueCol })] = { t: "s", v: value };
    const out = XLSX.write(wb, { type: "array", bookType: "xlsx" });
    const writable = await fileHandle.createWritable();
    await writable.write(out);
    await writable.close();
  }

  async function saveWorkStreamField(column, value) {
    const candidatesByColumn = {
      Name: ["Name"], Description: ["Description"], What: ["What"],
      Why: ["Why"], Who: ["Who"], Where: ["Where"], When: ["When"],
      How: ["How"], HowMuch: ["HowMuch", "How much"],
    };
    const candidates = candidatesByColumn[column];
    if (!candidates) throw new Error(`Unsupported WorkStream column: ${column}`);
    const globalId = String(WORKSTREAM_PAYLOAD?.elementId || "").trim();
    if (!globalId) throw new Error(t("noWorkStreamContext"));

    if (hasSheetJS() && runtime.state.workbookFileHandle && runtime.state.workbookSheetJs) {
      await _sheetJsWriteWorkStreamCell(candidates, value, globalId);
      runtime.state.liveWorkbookRecord = runtime.state.liveWorkbookRecord || {};
      runtime.state.liveWorkbookRecord[column] = value;
      if (runtime.scheduleSurfaceRegeneration) {
        runtime.scheduleSurfaceRegeneration("workstream_field:" + column);
      }
      return;
    }

    if (!runtime.state.activeMountPath || !runtime.state.liveWorkbookPath) {
      throw new Error(t("connectFolderFirst"));
    }
    const pyodide = await ensurePyodide({ withOpenpyxl: true });
    await withPyodideLock(async () => {
      pyodide.globals.set("workbook_path", runtime.state.liveWorkbookPath);
      pyodide.globals.set("worksheet_name", runtime.state.liveWorksheetName || "WorkStream");
      pyodide.globals.set("work_stream_global_id", globalId);
      pyodide.globals.set("work_stream_column_candidates_json", JSON.stringify(candidates));
      pyodide.globals.set("work_stream_value", value);
      const proxy = await pyodide.runPythonAsync(WORKBOOK_WRITE_SCRIPT);
      if (proxy && typeof proxy.destroy === "function") proxy.destroy();
      await syncMountedFilesystem(pyodide);
    });
    runtime.state.liveWorkbookRecord = runtime.state.liveWorkbookRecord || {};
    runtime.state.liveWorkbookRecord[column] = value;
    if (runtime.scheduleSurfaceRegeneration) {
      runtime.scheduleSurfaceRegeneration("workstream_field:" + column);
    }
  }

  async function openContainerFromHandle(handle) {
    if (!WORKSTREAM_PAYLOAD) {
      throw new Error("No WorkStream context payload present on this page.");
    }
    const resolved = await resolveContainerHandle(handle);
    const pyodide = await ensurePyodide({ withOpenpyxl: true });
    if (runtime.state.activeMountPath) {
      try {
        pyodide.FS.unmount(runtime.state.activeMountPath);
      } catch {
      }
      runtime.state.activeMountPath = null;
    }
    const mountPath = `/container_${Date.now()}`;
    pyodide.FS.mkdirTree(mountPath);
    await pyodide.mountNativeFS(mountPath, resolved);
    runtime.state.activeMountPath = mountPath;

    try {
      pyodide.registerJsModule("ontobdc_trace", {
        log: (msg) => {
          console.log("[ontobdc-pyodide]", String(msg));
        },
      });
    } catch {
      // already registered from a previous openContainer()/refresh call
    }

    const resultProxy = await withPyodideLock(() => {
      pyodide.globals.set("view_payload_json", JSON.stringify(WORKSTREAM_PAYLOAD));
      pyodide.globals.set("container_mount_path", mountPath);
      pyodide.globals.set(
        "work_stream_resource_name",
        runtime.WORK_STREAM_RESOURCE_NAME || "work_stream"
      );
      return pyodide.runPythonAsync(WORKBOOK_PARSE_SCRIPT);
    });
    const result = JSON.parse(String(resultProxy));
    if (resultProxy && typeof resultProxy.destroy === "function") {
      try { resultProxy.destroy(); } catch { /* no-op */ }
    }
    return result;
  }

  async function refreshFromWorkbook() {
    if (!runtime.state.rawContainerHandle && !runtime.state.datasetHandle) return;
    // Refresh only re-reads the workbook and re-renders — it must NOT touch
    // the connection status or the connect-button label (that state is owned
    // by the connect flow). This matches the IfcWorkSchedule Page's refresh,
    // which just re-parses and never repaints "Connecting…"/the status dot.
    const refreshBtn = document.querySelector(".refresh-btn");
    const originalTitle = refreshBtn?.title || "";
    try {
      if (refreshBtn) {
        refreshBtn.disabled = true;
        refreshBtn.title = t("refreshingFromWorkbook");
      }
      let result = null;
      if (hasSheetJS()) {
        try {
          result = await sheetJsRefreshFromWorkbook();
        } catch (sheetError) {
          console.warn("SheetJS refresh failed; falling back to Pyodide:", sheetError);
          result = null;
        }
      }
      if (!result || !result.nodes || !result.nodes.length) {
        result = await openContainerFromHandle(runtime.state.datasetHandle || runtime.state.rawContainerHandle);
      }
      applyLiveWorkbookResult(result);
    } catch (error) {
      console.error(error);
    } finally {
      if (refreshBtn) {
        refreshBtn.disabled = !runtime.state.rawContainerHandle && !runtime.state.datasetHandle;
        refreshBtn.title = originalTitle || t("refreshFromWorkbookTitle");
      }
    }
  }

  function applyLiveWorkbookResult(result, options) {
    if (!result) return;
    const fireCardOnConnected = !(options && options.fireCardOnConnected === false);
    runtime.state.liveWorkbookRecord = result.record || null;
    runtime.state.liveWorkbookPath = result.workbookPath || "";
    runtime.state.workbookRelPath = result.workbookRelativePath || "";
    runtime.state.liveWorksheetName = result.worksheetName || "WorkStream";
    runtime.state.liveRelationships = result.relationships || null;
    const openWorkbookBtn = document.querySelector(".workstream-open-workbook-btn");
    if (openWorkbookBtn) {
      openWorkbookBtn.disabled = !runtime.state.workbookRelPath;
    }
    renderHeader();
    // Re-mount every dimension card. The live record may have added a
    // dimension value that the static JSON-LD left blank (so the card
    // didn't even exist before this refresh), or removed one — a full
    // re-render is simpler and safer than trying to patch individual
    // cards' .dimension-value text nodes in-place.
    (runtime.renderDimensionCards || function () {})();
    applyStaticI18n();
    if (fireCardOnConnected) {
      (runtime.notifyDimensionCardsConnected || function () {})();
    }
  }


  Object.assign(runtime, {
    loadScriptTag: loadScriptTag,
    ensurePyodide: ensurePyodide,
    withPyodideLock: withPyodideLock,
    WORKBOOK_PARSE_SCRIPT: WORKBOOK_PARSE_SCRIPT,
    WORKBOOK_WRITE_SCRIPT: WORKBOOK_WRITE_SCRIPT,
    openContainerFromHandle: openContainerFromHandle,
    refreshFromWorkbook: refreshFromWorkbook,
    applyLiveWorkbookResult: applyLiveWorkbookResult,
    saveWorkStreamField: saveWorkStreamField,
    hasSheetJS: hasSheetJS,
    sheetJsRefreshFromWorkbook: sheetJsRefreshFromWorkbook
  });
  console.log("[work-stream-view] state end: pyodide_runtime_script_generated");
}(window));
"""

    def _linkset_operations_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: linkset_operations_script_generated");



  var ensurePyodide = runtime.ensurePyodide;

  // --- Linkset editing (pyodide + rdflib, loaded lazily on first use) ---
  //
  // Generalized from InfoBIM's WorkStream<->resource linkset: same ISO
  // 21597 (DirectedBinaryLink) shape, but with an OntoBDC URN scheme
  // instead of `urn:infobim:...`, and no dependency on openpyxl/workbook
  // parsing — 5W2H data is already resolved server-side, pyodide here only
  // ever reads/writes the dimension<->resource linksets themselves. A link
  // binds a *dimension* URI (not the WorkStream itself) to a resource URI.
  // Two independent linkset files exist under .__ontobdc__/linkset/:
  //   * WorkStreamResource.ttl  — "Related" (curated, confirmed relations;
  //                                no lifecycle — a link exists or not)
  //   * WorkStreamSuggested.ttl — "Suggested" (candidates with a lifecycle
  //                                state: Proposed -> Rejected. Rejected
  //                                links are kept for audit but are never
  //                                surfaced on the Suggested tab).
  // Both share the ISO 21597 DirectedBinaryLink structure.

  const LINKSET_FILES = Object.freeze({
    related: "WorkStreamResource.ttl",
    suggested: "WorkStreamSuggested.ttl",
  });

  const LINKSET_NS_PREFIXES = Object.freeze({
    related: "urn:ontobdc:linkset:workstream-resource",
    suggested: "urn:ontobdc:linkset:workstream-suggested",
  });

  const SUGGESTION_STATUS_NS = "http://ontobdc.org/ontology/domain/ontobdc/ns.ttl#";

  const SUGGESTION_STATUS = Object.freeze({
    PROPOSED: `${SUGGESTION_STATUS_NS}Proposed`,
    REJECTED: `${SUGGESTION_STATUS_NS}Rejected`,
  });

  const SUGGESTION_STATUS_PREDICATE = `${SUGGESTION_STATUS_NS}suggestionStatus`;
  const SUGGESTION_MODIFIED_AT_PREDICATE = `${SUGGESTION_STATUS_NS}suggestionModifiedAt`;

  const LINKSET_ACTIONS = Object.freeze({
    related: { add: "relate", remove: "unrelate" },
    suggested: { add: "suggest", remove: "unsuggest" },
  });

  async function linksetFileHandle(kind, create) {
    const fileName = LINKSET_FILES[kind];
    if (!fileName) throw new Error(`Unknown linkset kind: ${kind}`);
    const metadata = await runtime.state.datasetHandle.getDirectoryHandle(".__ontobdc__", { create });
    const linksetDir = await metadata.getDirectoryHandle("linkset", { create });
    return linksetDir.getFileHandle(fileName, { create });
  }

  async function readLinksetText(kind) {
    try {
      const handle = await linksetFileHandle(kind, false);
      return await (await handle.getFile()).text();
    } catch {
      return "";
    }
  }

  async function writeLinksetText(kind, text) {
    const handle = await linksetFileHandle(kind, true);
    const writable = await handle.createWritable();
    await writable.write(text);
    await writable.close();
  }

  const LINKSET_PYTHON_SCRIPT = `
import hashlib
import json
from datetime import datetime, timezone
from rdflib import BNode, Graph, Literal, Namespace, URIRef
from rdflib.namespace import RDF, XSD

LS = Namespace("https://standards.iso.org/iso/21597/-1/ed-1/en/Linkset#")
OBDC = Namespace(status_ns)

STATUS_PROPOSED = URIRef(status_proposed)
STATUS_REJECTED = URIRef(status_rejected)
PREDICATE_STATUS = URIRef(status_predicate)
PREDICATE_MODIFIED_AT = URIRef(status_modified_at_predicate)

graph = Graph()
graph.bind("obdc", OBDC, override=False)
graph.bind("ls", LS, override=False)
if existing_ttl.strip():
    graph.parse(data=existing_ttl, format="turtle")

def link_element_uri(element):
    if element is None:
        return None
    identifier = graph.value(element, LS.hasIdentifier)
    if identifier is None:
        return None
    value = graph.value(identifier, LS.uri)
    return str(value) if value is not None else None

def find_link(from_uri, to_uri):
    for link in graph.subjects(RDF.type, LS.DirectedBinaryLink):
        from_el = graph.value(link, LS.hasFromLinkElement)
        to_el = graph.value(link, LS.hasToLinkElement)
        if link_element_uri(from_el) == from_uri and link_element_uri(to_el) == to_uri:
            return link
    return None

def remove_link(link):
    for predicate in (LS.hasFromLinkElement, LS.hasToLinkElement):
        element = graph.value(link, predicate)
        if element is None:
            continue
        identifier = graph.value(element, LS.hasIdentifier)
        if identifier is not None:
            graph.remove((identifier, None, None))
        graph.remove((element, None, None))
    graph.remove((link, None, None))

def make_element(uri):
    element = BNode()
    identifier = BNode()
    graph.add((element, LS.hasIdentifier, identifier))
    graph.add((identifier, RDF.type, LS.URIBasedIdentifier))
    graph.add((identifier, LS.uri, Literal(uri, datatype=XSD.anyURI)))
    return element

def set_link_status(link, status_uri):
    graph.set((link, PREDICATE_STATUS, status_uri))
    now = datetime.now(timezone.utc).isoformat()
    graph.set((link, PREDICATE_MODIFIED_AT, Literal(now, datatype=XSD.dateTimeStamp)))

def link_status(link):
    value = graph.value(link, PREDICATE_STATUS)
    return str(value) if value is not None else None

action_add = ns_prefix + ":add"
action_remove = ns_prefix + ":remove"
has_status_lifecycle = has_status_lifecycle == "true"

existing_link = find_link(dimension_uri, resource_uri) if resource_uri else None

if action == action_remove and existing_link is not None:
    if has_status_lifecycle:
        set_link_status(existing_link, STATUS_REJECTED)
    else:
        remove_link(existing_link)
elif action == action_add:
    if existing_link is None:
        digest = hashlib.sha256((dimension_uri + "|" + resource_uri).encode("utf-8")).hexdigest()[:16]
        link = URIRef(f"{ns_prefix}:{digest}")
        graph.add((link, RDF.type, LS.DirectedBinaryLink))
        graph.add((link, LS.hasFromLinkElement, make_element(dimension_uri)))
        graph.add((link, LS.hasToLinkElement, make_element(resource_uri)))
        if has_status_lifecycle:
            set_link_status(link, STATUS_PROPOSED)
    elif has_status_lifecycle:
        current_status = link_status(existing_link)
        if current_status == str(STATUS_REJECTED):
            set_link_status(existing_link, STATUS_PROPOSED)

active_entries = {}
all_status_entries = {}
for link in graph.subjects(RDF.type, LS.DirectedBinaryLink):
    from_el = graph.value(link, LS.hasFromLinkElement)
    to_el = graph.value(link, LS.hasToLinkElement)
    from_uri = link_element_uri(from_el)
    to_uri = link_element_uri(to_el)
    if not from_uri or not to_uri:
        continue
    status = link_status(link)
    all_status_entries.setdefault(from_uri, {})[to_uri] = status
    is_rejected = has_status_lifecycle and status == str(STATUS_REJECTED)
    if not is_rejected:
        active_entries.setdefault(from_uri, []).append(to_uri)

json.dumps({
    "ttl": graph.serialize(format="turtle"),
    "entries": active_entries,
    "allStatus": all_status_entries,
})
`;

  // ── Pure-JS linkset editing ────────────────────────────────────────────
  // The ISO 21597 linkset files this reads/writes are generated by OntoBDC
  // itself with a fixed shape (one DirectedBinaryLink per resource, each with
  // a nested from/to URIBasedIdentifier and — for the suggested kind — a
  // status/modifiedAt pair). That regularity means the read-modify-write can
  // run entirely in the browser, so relate/suggest work on file:// too:
  // Pyodide + rdflib were only ever needed here for the Turtle round-trip and
  // never boot offline (no vendored wheels). The Python side still parses
  // these files with rdflib, so the output stays standard Turtle.

  const _LINKSET_TTL_PREFIXES = {
    related:
      '@prefix ls: <https://standards.iso.org/iso/21597/-1/ed-1/en/Linkset#> .\n'
      + '@prefix xsd: <http://www.w3.org/2001/XMLSchema#> .\n',
    suggested:
      '@prefix ls: <https://standards.iso.org/iso/21597/-1/ed-1/en/Linkset#> .\n'
      + '@prefix obdc: <http://ontobdc.org/ontology/domain/ontobdc/ns.ttl#> .\n'
      + '@prefix xsd: <http://www.w3.org/2001/XMLSchema#> .\n',
  };

  function _ttlEscape(value) {
    return String(value)
      .replace(/\\/g, "\\\\")
      .replace(/"/g, '\\"')
      .replace(/\n/g, "\\n")
      .replace(/\r/g, "\\r")
      .replace(/\t/g, "\\t");
  }
  function _ttlUnescape(value) {
    return String(value).replace(/\\(["\\ntr])/g, function (_, ch) {
      return ch === "n" ? "\n" : ch === "t" ? "\t" : ch === "r" ? "\r" : ch;
    });
  }

  function _statusToUri(local) {
    return local ? SUGGESTION_STATUS_NS + local : null;
  }
  function _statusLocalName(uri) {
    return uri ? String(uri).split("#").pop() : null;
  }

  // Parse a linkset TTL into [{ linkUri, from, to, status, modifiedAt }].
  function parseLinksetLinks(ttl) {
    const links = [];
    if (!ttl || !ttl.trim()) return links;
    // rdflib serializes one subject per blank-line-separated paragraph.
    const chunks = ttl.split(/\r?\n\s*\r?\n/);
    for (const chunk of chunks) {
      if (!chunk || chunk.indexOf("ls:DirectedBinaryLink") === -1) continue;
      const linkMatch = chunk.match(/<([^>]+)>\s+a\s+ls:DirectedBinaryLink/);
      if (!linkMatch) continue;
      const uris = [];
      const uriRe = /ls:uri\s+"((?:[^"\\]|\\.)*)"\s*\^\^\s*xsd:anyURI/g;
      let u;
      while ((u = uriRe.exec(chunk)) !== null) uris.push(_ttlUnescape(u[1]));
      if (uris.length < 2) continue;
      let from = uris[0];
      let to = uris[1];
      const fromKw = chunk.indexOf("hasFromLinkElement");
      const toKw = chunk.indexOf("hasToLinkElement");
      if (fromKw !== -1 && toKw !== -1 && toKw < fromKw) {
        from = uris[1];
        to = uris[0];
      }
      const statusMatch = chunk.match(
        /suggestionStatus\s+(?:obdc:(\w+)|<([^>]+)>)/
      );
      const status = statusMatch
        ? (statusMatch[1] ? _statusToUri(statusMatch[1]) : statusMatch[2])
        : null;
      const modifiedMatch = chunk.match(
        /suggestionModifiedAt\s+"((?:[^"\\]|\\.)*)"/
      );
      links.push({
        linkUri: linkMatch[1],
        from: from,
        to: to,
        status: status,
        modifiedAt: modifiedMatch ? modifiedMatch[1] : null,
      });
    }
    return links;
  }

  function serializeLinksetLinks(kind, links) {
    const parts = [_LINKSET_TTL_PREFIXES[kind] || _LINKSET_TTL_PREFIXES.related, ""];
    for (const link of links) {
      let block = `<${link.linkUri}> a ls:DirectedBinaryLink`;
      if (kind === "suggested" && link.status) {
        if (link.modifiedAt) {
          block += ` ;\n    obdc:suggestionModifiedAt "${_ttlEscape(link.modifiedAt)}"^^xsd:dateTimeStamp`;
        }
        block += ` ;\n    obdc:suggestionStatus obdc:${_statusLocalName(link.status)}`;
      }
      block +=
        ` ;\n    ls:hasFromLinkElement [ ls:hasIdentifier [ a ls:URIBasedIdentifier ;\n`
        + `                    ls:uri "${_ttlEscape(link.from)}"^^xsd:anyURI ] ] ;\n`
        + `    ls:hasToLinkElement [ ls:hasIdentifier [ a ls:URIBasedIdentifier ;\n`
        + `                    ls:uri "${_ttlEscape(link.to)}"^^xsd:anyURI ] ] .`;
      parts.push(block, "");
    }
    return parts.join("\n");
  }

  async function _sha256Hex16(text) {
    const digest = await crypto.subtle.digest(
      "SHA-256",
      new TextEncoder().encode(text)
    );
    let hex = "";
    const bytes = new Uint8Array(digest);
    for (let i = 0; i < bytes.length; i++) {
      hex += bytes[i].toString(16).padStart(2, "0");
    }
    return hex.slice(0, 16);
  }

  // Serializes every linkset read-modify-write through one queue so two
  // near-simultaneous button clicks (or a burst of dimension-card reads on
  // connect) can't clobber the same file.
  let _linksetQueue = Promise.resolve();
  function _withLinksetLock(taskFn) {
    const settled = _linksetQueue.then(taskFn, taskFn);
    _linksetQueue = settled.then(() => {}, () => {});
    return settled;
  }

  // Runs one read/add/remove op on a linkset kind (related/suggested)
  // and returns { entries, allStatus } where:
  //   * `entries`     — dimension URI -> resource URI list (only "active"
  //                     links: every related link, plus Proposed suggestions;
  //                     Rejected suggestions are excluded on purpose).
  //   * `allStatus`   — dimension URI -> resource URI -> status URI string.
  async function runLinksetOperation(kind, action, dimensionUri, resourceId) {
    const nsPrefix = LINKSET_NS_PREFIXES[kind];
    if (!nsPrefix) throw new Error(`Unknown linkset kind: ${kind}`);
    const actionCodes = LINKSET_ACTIONS[kind];
    if (!actionCodes) throw new Error(`Unknown linkset kind: ${kind}`);

    const codeMap = {
      [actionCodes.add]: "add",
      [actionCodes.remove]: "remove",
      read: "read",
    };
    const codedAction = codeMap[action];
    if (codedAction === undefined) throw new Error(`Unknown action ${action} for kind ${kind}`);

    const hasStatusLifecycle = kind === "suggested";
    const REJECTED = SUGGESTION_STATUS.REJECTED;
    const PROPOSED = SUGGESTION_STATUS.PROPOSED;

    return _withLinksetLock(async () => {
      const links = parseLinksetLinks(await readLinksetText(kind));

      if (codedAction !== "read" && dimensionUri && resourceId) {
        const index = links.findIndex(
          (link) => link.from === dimensionUri && link.to === resourceId
        );
        if (codedAction === "remove" && index !== -1) {
          if (hasStatusLifecycle) {
            links[index].status = REJECTED;
            links[index].modifiedAt = new Date().toISOString();
          } else {
            links.splice(index, 1);
          }
        } else if (codedAction === "add") {
          if (index === -1) {
            const digest = await _sha256Hex16(dimensionUri + "|" + resourceId);
            const link = {
              linkUri: `${nsPrefix}:${digest}`,
              from: dimensionUri,
              to: resourceId,
              status: hasStatusLifecycle ? PROPOSED : null,
              modifiedAt: hasStatusLifecycle ? new Date().toISOString() : null,
            };
            links.push(link);
          } else if (hasStatusLifecycle && links[index].status === REJECTED) {
            links[index].status = PROPOSED;
            links[index].modifiedAt = new Date().toISOString();
          }
        }
      }

      if (codedAction !== "read") {
        await writeLinksetText(kind, serializeLinksetLinks(kind, links));
      }

      const entries = {};
      const allStatus = {};
      for (const link of links) {
        (allStatus[link.from] = allStatus[link.from] || {})[link.to] = link.status || null;
        const isRejected = hasStatusLifecycle && link.status === REJECTED;
        if (!isRejected) {
          (entries[link.from] = entries[link.from] || []).push(link.to);
        }
      }
      return { entries, allStatus };
    });
  }

  async function loadAllLinks(kind) {
    if (!runtime.state.datasetHandle) return { entries: {}, allStatus: {} };
    try {
      return await runLinksetOperation(kind, "read", null, null);
    } catch (error) {
      console.error(error);
      return { entries: {}, allStatus: {} };
    }
  }


  Object.assign(runtime, {
    linksetFileHandle: linksetFileHandle,
    readLinksetText: readLinksetText,
    writeLinksetText: writeLinksetText,
    LINKSET_PYTHON_SCRIPT: LINKSET_PYTHON_SCRIPT,
    runLinksetOperation: runLinksetOperation,
    loadAllLinks: loadAllLinks
  });
  console.log("[work-stream-view] state end: linkset_operations_script_generated");
}(window));
"""

    def _file_category_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: file_category_script_generated");



  var literal = runtime.literal;
  var nodeTypes = runtime.nodeTypes;
  var ensurePyodide = runtime.ensurePyodide;

  // --- Declarative file display profiles (brasidatacenter/ontology/ontobdc/domain/file.ttl) ---
  //
  // Mirror of the InfoBIM reference's file_display.ttl engine: when a
  // container ships a file.ttl (compatible with the canonical OntoBDC file
  // ontology) under .__ontobdc__/ontology/file.ttl (or the legacy
  // file_display.ttl path for coexistence with existing generated datasets)
  // we parse its :FileDisplayProfile individuals in pyodide/rdflib and
  // return both (a) the ordered list of display categories for rendering
  // category tiles and (b) a helper `category_for(resource_node)` that
  // matches a JSON-LD resource (its type, mime, filePath extension) against
  // each profile following the canonical precedence chain:
  //   1. profile requires a semantic type and the resource has that type
  //   2. profile has an acceptedMimeType equal to the resource's inferred mime
  //   3. profile has an acceptedExtension equal to the resource's lowercased
  //      filename extension (including the leading full stop)
  //
  // When no such ontology file exists in the container the engine returns
  // an empty list of categories and every call to `category_for` returns
  // null so the UI gracefully degrades to its legacy no-tile, "show all the
  // catalogued files" behaviour.

  const FILE_DISPLAY_ONTOLOGY_CANDIDATES = Object.freeze([
    "file.ttl",
    "file_display.ttl",
  ]);

  const FILE_DISPLAY_PARSE_SCRIPT = `
import json
from rdflib import Graph, Namespace, RDF, RDFS, URIRef

FILE = Namespace("https://w3id.org/ontobdc/ontology/file#")
OBDC = Namespace("http://ontobdc.org/ontology/domain/ontobdc/ns.ttl#")
SIG = Namespace("http://ontobdc.org/ontology/domain/signature.ttl#")

g = Graph()
if ontology_ttl.strip():
    g.parse(data=ontology_ttl, format="turtle")

def scalar(uri, predicate):
    value = g.value(uri, predicate)
    if value is None:
        return None
    try:
        return str(value.toPython())
    except Exception:
        return str(value)

def all_values(uri, predicate):
    return [
        str(v) if not isinstance(v, URIRef) else str(v)
        for v in g.objects(uri, predicate)
    ]

profiles = []
for profile_uri in sorted(g.subjects(RDF.type, FILE.FileDisplayProfile), key=lambda u: str(u)):
    category = scalar(profile_uri, FILE.displayCategory)
    if not category:
        continue
    profiles.append({
        "id": category,
        "category": category,
        "label": scalar(profile_uri, FILE.categoryLabel) or category,
        "requiredSemanticType": str(scalar(profile_uri, FILE.requiredSemanticType))
            if scalar(profile_uri, FILE.requiredSemanticType) else None,
        "acceptedMimeTypes": sorted(set(all_values(profile_uri, FILE.acceptedMimeType))),
        "acceptedExtensions": sorted(set(all_values(profile_uri, FILE.acceptedExtension))),
    })

# Preserve the canonical InfoBIM ordering users already expect when all four
# legacy categories are present. Any custom categories added to the ontology
# file keep their sorted-by-URI position after the known ones.
canonical_order = {"drawings": 10, "documents": 20, "photos": 30, "messages": 40}
profiles.sort(key=lambda p: (canonical_order.get(p["id"], 99), p["label"], p["id"]))

json.dumps({"profiles": profiles})
`;

  const FILE_DISPLAY_STATUS_NS = "http://ontobdc.org/ontology/domain/ontobdc/ns.ttl#";

  function extensionOfNode(node) {
    const filePath = literal(node, `${runtime.OBDC_NS}filePath`);
    if (!filePath) return null;
    const fileName = filePath.split("/").pop();
    if (!fileName || !fileName.includes(".")) return null;
    const index = fileName.lastIndexOf(".");
    return fileName.slice(index).toLowerCase();
  }

  function mimeOfNode(node) {
    return literal(node, `${FILE_DISPLAY_STATUS_NS}mimeType`) || literal(node, "https://schema.org/encodingFormat");
  }

  function typedAs(node, requiredTypeUri) {
    if (!requiredTypeUri) return true;
    const types = nodeTypes(node);
    if (types.some((type) => type === requiredTypeUri)) return true;
    const short = requiredTypeUri.split("#").slice(-1)[0] || requiredTypeUri;
    return types.some((type) => type.endsWith(`#${short}`));
  }

  function matchesProfile(node, profile, inferredExtension, inferredMime) {
    const required = profile.requiredSemanticType;
    const extMatches = inferredExtension
      ? profile.acceptedExtensions.includes(inferredExtension)
      : false;
    const mimeMatches = inferredMime
      ? profile.acceptedMimeTypes.includes(inferredMime)
      : false;
    if (required) {
      if (!typedAs(node, required)) return false;
      return extMatches || mimeMatches || profile.acceptedExtensions.length + profile.acceptedMimeTypes.length === 0;
    }
    return extMatches || mimeMatches;
  }

  let cachedFileDisplayPromise = null;

  async function readFileDisplayOntologyText() {
    if (!runtime.state.datasetHandle) return null;
    try {
      const metadata = await runtime.state.datasetHandle.getDirectoryHandle(".__ontobdc__", { create: false });
      const ontologyDir = await metadata.getDirectoryHandle("ontology", { create: false });
      let handle = null;
      for (const fileName of FILE_DISPLAY_ONTOLOGY_CANDIDATES) {
        try {
          handle = await ontologyDir.getFileHandle(fileName, { create: false });
          break;
        } catch {
          handle = null;
        }
      }
      if (!handle) return null;
      return await (await handle.getFile()).text();
    } catch {
      return null;
    }
  }

  // Returns { profiles: Array<DisplayProfile> }
  async function loadFileDisplayProfiles() {
    if (cachedFileDisplayPromise) return cachedFileDisplayPromise;
    cachedFileDisplayPromise = (async () => {
      try {
        const [ontologyText, pyodide] = await Promise.all([
          readFileDisplayOntologyText(),
          ensurePyodide(),
        ]);
        if (!ontologyText) return { profiles: [] };
        const resultJson = await runtime.withPyodideLock(() => {
          pyodide.globals.set("ontology_ttl", ontologyText);
          return pyodide.runPythonAsync(FILE_DISPLAY_PARSE_SCRIPT);
        });
        const result = JSON.parse(resultJson);
        return { profiles: Array.isArray(result.profiles) ? result.profiles : [] };
      } catch (error) {
        console.error(error);
        return { profiles: [] };
      }
    })();
    return cachedFileDisplayPromise;
  }

  function categoryForNode(node, profiles) {
    if (!node || !profiles || profiles.length === 0) return null;
    const inferredExtension = extensionOfNode(node);
    const inferredMime = mimeOfNode(node);
    for (const profile of profiles) {
      if (matchesProfile(node, profile, inferredExtension, inferredMime)) {
        return profile.id;
      }
    }
    return null;
  }


  Object.assign(runtime, {
    extensionOfNode: extensionOfNode,
    mimeOfNode: mimeOfNode,
    typedAs: typedAs,
    matchesProfile: matchesProfile,
    readFileDisplayOntologyText: readFileDisplayOntologyText,
    loadFileDisplayProfiles: loadFileDisplayProfiles,
    categoryForNode: categoryForNode
  });
  console.log("[work-stream-view] state end: file_category_script_generated");
}(window));
"""

    def _dimension_card_source(self) -> str:
        return r"""(function (window) {
  "use strict";
  var runtime = window.OntoBDCWorkStreamViewRuntime = window.OntoBDCWorkStreamViewRuntime || { state: {} };
  console.log("[work-stream-view] state start: dimension_card_script_generated");



  var __ontobdcIdentityT = function __ontobdcIdentityT(key, vars) {
    var text = String(key || "");
    if (vars) {
      try {
        Object.keys(vars).forEach(function (k) {
          text = text.split("{" + k + "}").join(String(vars[k]));
        });
      } catch (e) { /* noop */ }
    }
    return text;
  };
  if (typeof runtime.t !== "function") runtime.t = __ontobdcIdentityT;
  var t = (typeof runtime.t === "function") ? runtime.t : __ontobdcIdentityT;
  var literal = runtime.literal;
  var loadGraph = runtime.loadGraph;
  var renderHeader = runtime.renderHeader;
  var applyStaticI18n = runtime.applyStaticI18n;
  var resourceNodes = runtime.resourceNodes;
  var resourceLabel = runtime.resourceLabel;
  var resourceMimeKind = runtime.resourceMimeKind;
  var parseCsv = runtime.parseCsv;
  var renderCsvTable = runtime.renderCsvTable;
  var renderCsvFallback = runtime.renderCsvFallback;
  var tryReconnectSilently = runtime.tryReconnectSilently;
  var wireAnnotationControls = runtime.wireAnnotationControls;
  var ensureAnnotationRuntime = runtime.ensureAnnotationRuntime;
  var RESOURCE_NODE_ICONS = Object.freeze({
    file: '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8Z"/><polyline points="14 2 14 8 20 8"/></svg>',
    folder: '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M3 6a2 2 0 0 1 2-2h5l2 3h7a2 2 0 0 1 2 2v9a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2Z"/></svg>',
    folderOpen: '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M3 7V6a2 2 0 0 1 2-2h5l2 3h7a2 2 0 0 1 2 2v1"/><path d="M3.4 10h17.2a1 1 0 0 1 .95 1.32l-2.34 7A2 2 0 0 1 17.31 20H5.69a2 2 0 0 1-1.9-1.37l-2.34-7A1 1 0 0 1 2.4 10Z"/></svg>',
  });
  var annotateResource = runtime.annotateResource;
  var loadAllLinks = runtime.loadAllLinks;
  var runLinksetOperation = runtime.runLinksetOperation;
  var loadFileDisplayProfiles = runtime.loadFileDisplayProfiles;
  var categoryForNode = runtime.categoryForNode;
  var OBDC_NS = runtime.OBDC_NS;

  // The seven WorkStreamDimensions (type.ttl :DimensionKind individuals),
  // each with its own Related/Found resource tree — a relate action links
  // a resource to exactly one dimension, never to the WorkStream itself.
  // `slug` matches each DimensionKind's schema:identifier exactly (used to
  // build the dimension's URI via OntoBDCWorkStreamContext.dimensionUri).
  const DIMENSIONS = [
    { key: "what", column: "What", slug: "what", property: `${runtime.WORK_STREAM_TYPE_NS}what` },
    { key: "why", column: "Why", slug: "why", property: `${runtime.WORK_STREAM_TYPE_NS}why` },
    { key: "who", column: "Who", slug: "who", property: `${runtime.WORK_STREAM_TYPE_NS}who` },
    { key: "where", column: "Where", slug: "where", property: `${runtime.WORK_STREAM_TYPE_NS}where` },
    { key: "when", column: "When", slug: "when", property: `${runtime.WORK_STREAM_TYPE_NS}when` },
    { key: "how", column: "How", slug: "how", property: `${runtime.WORK_STREAM_TYPE_NS}how` },
    { key: "howMuch", column: "HowMuch", slug: "how-much", property: `${runtime.WORK_STREAM_TYPE_NS}howMuch` },
  ];

  let workStreamContext = null;
  // Each card registers a reload() (re-check relations) and a
  // setConnected()/setDisconnected() callback here, invoked from the
  // shared folder-connection flow (pyodide_runtime.js's applyLiveWorkbookResult).
  const dimensionCards = [];

  function createDimensionCard(dimension, value) {
    const dimensionUri = workStreamContext ? workStreamContext.dimensionUri(dimension.slug) : "";

    const card = document.createElement("section");
    card.className = "tile dimension-card";
    card.innerHTML = `
      <div class="dimension-card-head">
        <div class="dimension-label">${t(dimension.key)}</div>
      </div>
      <div class="dimension-value"></div>
      <div class="resource-tabs" role="tablist">
        <button type="button" class="resource-tab is-active" data-tab="related" role="tab" aria-selected="true">${t("related")}</button>
        <button type="button" class="resource-tab" data-tab="suggested" role="tab" aria-selected="false">${t("suggested")}</button>
        <button type="button" class="resource-tab" data-tab="found" role="tab" aria-selected="false">${t("found")}</button>
      </div>
      <div class="resource-body">
        <div class="resource-list">
          <nav class="resource-categories" aria-label="${t("resourceCategories")}" data-resource-categories hidden></nav>
          <div class="resource-tree-container"></div>
        </div>
        <div class="resource-preview-column">
          <div class="preview-tabs" role="tablist"></div>
          <div class="resource-preview"></div>
          <div class="inline-annotations">
            <span class="inline-annotations-label">${t("annotations")}</span>
            <div class="inline-annotations-actions">
              <button type="button" class="icon-btn view-annotations-btn" disabled hidden title="${t("viewAnnotations")}" aria-label="${t("viewAnnotations")}">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M1 12s4-7 11-7 11 7 11 7-4 7-11 7-11-7-11-7Z"/><circle cx="12" cy="12" r="3"/></svg>
              </button>
              <button type="button" class="icon-btn create-annotation-btn" disabled hidden title="${t("createAnnotation")}" aria-label="${t("createAnnotation")}">
                <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 20h9"/><path d="M16.5 3.5a2.121 2.121 0 0 1 3 3L7 19l-4 1 1-4Z"/></svg>
              </button>
            </div>
          </div>
        </div>
      </div>
    `;
    runtime.mountInlineEditor(card.querySelector(".dimension-value"), {
      value: value,
      multiline: true,
      rows: 3,
      editLabel: t("editField", { field: t(dimension.key) }),
      saveLabel: t("saveField", { field: t(dimension.key) }),
      cancelLabel: t("cancelEdit"),
      emptyLabel: t("emptyValue"),
      onSave: (nextValue) => runtime.saveWorkStreamField(dimension.column, nextValue),
    });

    const previewEl = card.querySelector(".resource-preview");
    const tabStripEl = card.querySelector(".preview-tabs");
    const emptyPane = document.createElement("div");
    emptyPane.className = "preview-pane preview-pane-empty is-active";
    emptyPane.innerHTML = '<div class="resource-preview-empty"></div>';
    emptyPane.querySelector(".resource-preview-empty").textContent = t("selectResourceToPreview");
    previewEl.appendChild(emptyPane);

    let activeTab = "related";
    // Multi-tab preview state: `openTabs` is the ordered list of currently
    // open resource ids (tab strip order); `activeTabId` is whichever one
    // is visible right now (null = nothing open, the empty pane shows);
    // `previewPanes` caches each open tab's mounted DOM subtree so
    // switching tabs never re-fetches/re-parses a CSV or reloads a PDF
    // iframe. This is deliberately independent of the three filter axes
    // below — flipping a filter must never touch which preview tabs are
    // open.
    let openTabs = [];
    let activeTabId = null;
    const previewPanes = new Map();
    let relatedResourceIds = new Set();
    let suggestedResourceIds = new Set();
    let displayProfiles = [];
    let activeCategory = null;
    let expandedPaths = new Set();
    let connected = false;
    let annotationsTile = null;
    let annotationsTileResourceId = null;

    function currentTabNodes() {
      const resources = resourceNodes();
      const withStatus = resources.filter((node) => {
        const nodeId = node["@id"];
        if (activeTab === "related") return relatedResourceIds.has(nodeId);
        if (activeTab === "suggested") {
          return suggestedResourceIds.has(nodeId) && !relatedResourceIds.has(nodeId);
        }
        return true;
      });
      if (!activeCategory) return withStatus;
      return withStatus.filter((node) => categoryForNode(node, displayProfiles) === activeCategory);
    }

    function _dedupSuggestedVsRelated() {
      const ids = Array.from(suggestedResourceIds);
      for (let i = 0; i < ids.length; i++) {
        if (relatedResourceIds.has(ids[i])) suggestedResourceIds.delete(ids[i]);
      }
    }

    function renderCategoryTiles() {
      const container = card.querySelector("[data-resource-categories]");
      if (!container) return;
      container.innerHTML = "";
      if (!displayProfiles || displayProfiles.length === 0) {
        container.hidden = true;
        return;
      }
      container.hidden = false;
      const allButton = document.createElement("button");
      allButton.type = "button";
      allButton.className = `resource-category-link${activeCategory === null ? " is-active" : ""}`;
      allButton.dataset.category = "";
      allButton.setAttribute("role", "tab");
      allButton.setAttribute("aria-expanded", String(activeCategory === null));
      allButton.textContent = t("all");
      allButton.title = t("showAllCategories");
      allButton.addEventListener("click", () => {
        activeCategory = null;
        renderCategoryTiles();
        renderResources();
      });
      container.appendChild(allButton);
      for (const profile of displayProfiles) {
        const tile = document.createElement("button");
        tile.type = "button";
        tile.className = `resource-category-link${activeCategory === profile.id ? " is-active" : ""}`;
        tile.dataset.category = profile.id;
        tile.setAttribute("role", "tab");
        tile.setAttribute("aria-expanded", String(activeCategory === profile.id));
        tile.textContent = profile.label;
        tile.title = `Show only ${profile.label} resources`;
        tile.addEventListener("click", () => {
          activeCategory = profile.id;
          renderCategoryTiles();
          renderResources();
        });
        container.appendChild(tile);
      }
    }

    function updateAnnotateButton(node) {
      // annotationRuntime is only ever created lazily, the first time
      // annotateResource()/openWorkspace()/openSubjectPage() actually
      // runs (see ensureAnnotationRuntime()) — nothing creates it
      // proactively on folder-connect. Gating readiness on it being
      // already non-null left these buttons permanently disabled
      // (silently swallowing clicks, no error) until the user happened
      // to open Workspace or Subjects first. Folder connection is the
      // only real precondition; the runtime creates itself on demand.
      const ready = Boolean(runtime && runtime.state && runtime.state.datasetHandle && typeof OntoBDCAnnotations !== "undefined");
      const viewButton = card.querySelector(".view-annotations-btn");
      const createButton = card.querySelector(".create-annotation-btn");
      const kind = node ? resourceMimeKind(node) : null;

      viewButton.hidden = !node;
      if (!node) {
        viewButton.disabled = true;
        createButton.hidden = true;
        createButton.disabled = true;
        updateWorkspaceButtonState();
        return;
      }

      const annotatable = kind === "image" || kind === "pdf";
      createButton.hidden = !annotatable;
      createButton.disabled = !ready;

      loadResourceAnnotations(node)
        .then((annotations) => {
          viewButton.disabled = !ready || annotations.length === 0;
        })
        .catch((error) => {
          console.error(error);
          viewButton.disabled = !ready;
        })
        .finally(() => {
          updateWorkspaceButtonState();
        });

      // Only tear down the annotations panel when it would now be listing
      // a different resource than the newly active tab — opening/switching
      // to a tab that's already the one the panel is showing (e.g. closing
      // an unrelated background tab) must not spuriously close it.
      if (annotationsTile && annotationsTileResourceId !== (node ? node["@id"] : null)) {
        closeAnnotationsTile();
      }
    }

    function updateWorkspaceButtonState() {
      const headerWorkspaceBtn = document.querySelector(".workspace-btn");
      if (!headerWorkspaceBtn) return;
      if (!runtime || !runtime.state || !runtime.state.datasetHandle) {
        headerWorkspaceBtn.disabled = true;
        return;
      }
      const runtimeInstance = ensureAnnotationRuntime();
      if (!runtimeInstance) {
        headerWorkspaceBtn.disabled = true;
        return;
      }
      try {
        const scratch = document.createElement("div");
        runtimeInstance.openWorkspace(scratch, { containerHandle: runtime.state.datasetHandle })
          .then(() => {
            try {
              const all = (typeof runtime.listAnnotations === "function" ? runtime.listAnnotations() : []) || [];
              headerWorkspaceBtn.disabled = all.length === 0;
            } catch {
              headerWorkspaceBtn.disabled = !runtime.state?.datasetHandle;
            }
          })
          .catch((error) => {
            console.error(error);
            headerWorkspaceBtn.disabled = !runtime.state?.datasetHandle;
          });
      } catch (error) {
        console.error(error);
        headerWorkspaceBtn.disabled = !runtime.state?.datasetHandle;
      }
    }

    // Builds one tab's mounted preview content. Called once per resource,
    // the first time its tab opens — the result is cached in `previewPanes`
    // so switching tabs is a pure visibility toggle (updatePreviewVisibility),
    // never a re-fetch/re-render.
    function mountPreviewPane(node) {
      const pane = document.createElement("div");
      pane.className = "preview-pane";
      pane.dataset.resourceId = node["@id"];

      const filePath = literal(node, `${OBDC_NS}filePath`);
      const href = filePath ? `../../../${filePath}` : null;
      const kind = resourceMimeKind(node);

      if (href && kind === "image") {
        const img = document.createElement("img");
        img.src = href;
        img.alt = resourceLabel(node);
        pane.appendChild(img);
        return pane;
      }
      if (href && kind === "pdf") {
        const frame = document.createElement("iframe");
        frame.src = href;
        frame.title = resourceLabel(node);
        pane.appendChild(frame);
        return pane;
      }
      if (href && kind === "csv") {
        pane.classList.add("is-csv");
        const loading = document.createElement("div");
        loading.className = "resource-preview-empty";
        loading.textContent = t("loadingCsv");
        pane.appendChild(loading);
        fetch(href)
          .then((response) => {
            if (!response.ok) throw new Error(`HTTP ${response.status}`);
            return response.text();
          })
          .then((text) => {
            // The tab may have been closed (and this pane discarded) while
            // the fetch was in flight — only the still-current pane for
            // this resource id may be mutated.
            if (previewPanes.get(node["@id"]) !== pane) return;
            const rows = parseCsv(text);
            pane.innerHTML = "";
            pane.appendChild(rows.length ? renderCsvTable(rows) : renderCsvFallback(href));
          })
          .catch((error) => {
            console.error(error);
            if (previewPanes.get(node["@id"]) !== pane) return;
            pane.innerHTML = "";
            pane.appendChild(renderCsvFallback(href));
          });
        return pane;
      }

      const link = document.createElement("a");
      link.className = "resource-open-link";
      link.href = href || "#";
      link.target = "_blank";
      link.rel = "noopener";
      link.textContent = href ? t("openResource", { name: resourceLabel(node) }) : t("noPreviewAvailable");
      pane.appendChild(link);
      return pane;
    }

    function renderPreviewTabStrip() {
      tabStripEl.innerHTML = "";
      for (const resourceId of openTabs) {
        const node = runtime.state.graphNodes.find((item) => item["@id"] === resourceId);
        if (!node) continue;
        const tab = document.createElement("div");
        tab.className = `preview-tab${resourceId === activeTabId ? " is-active" : ""}`;
        tab.setAttribute("role", "tab");
        tab.setAttribute("aria-selected", String(resourceId === activeTabId));

        const label = document.createElement("span");
        label.className = "preview-tab-label";
        label.textContent = resourceLabel(node);
        label.title = resourceLabel(node);

        const closeBtn = document.createElement("span");
        closeBtn.className = "preview-tab-close";
        closeBtn.textContent = "×";
        closeBtn.setAttribute("role", "button");
        closeBtn.setAttribute("aria-label", `Close ${resourceLabel(node)}`);
        closeBtn.addEventListener("click", (event) => {
          event.stopPropagation();
          closeTab(resourceId);
        });

        tab.append(label, closeBtn);
        tab.addEventListener("click", () => activateTab(resourceId));
        tabStripEl.appendChild(tab);
      }
    }

    function updatePreviewVisibility() {
      emptyPane.classList.toggle("is-active", activeTabId === null);
      for (const [resourceId, pane] of previewPanes) {
        pane.classList.toggle("is-active", resourceId === activeTabId);
      }
    }

    function activateTab(resourceId) {
      activeTabId = resourceId;
      renderPreviewTabStrip();
      updatePreviewVisibility();
      renderList();
      const node = runtime.state.graphNodes.find((item) => item["@id"] === resourceId) || null;
      updateAnnotateButton(node);
    }

    // Single click both opens (if new) and switches to (always) a resource's
    // tab — there is no double-click distinction here, unlike the main
    // Surface's file tree (single-select / double-activate), since this
    // page has no separate "reveal a pre-existing Tile" step to gate.
    function openResourceTab(node) {
      const resourceId = node["@id"];
      if (!openTabs.includes(resourceId)) {
        openTabs.push(resourceId);
        const pane = mountPreviewPane(node);
        previewPanes.set(resourceId, pane);
        previewEl.appendChild(pane);
      }
      activateTab(resourceId);
    }

    function closeTab(resourceId) {
      const index = openTabs.indexOf(resourceId);
      if (index === -1) return;
      openTabs.splice(index, 1);
      const pane = previewPanes.get(resourceId);
      if (pane) pane.remove();
      previewPanes.delete(resourceId);

      if (activeTabId === resourceId) {
        // Activate the tab that slides into the closed one's slot — the
        // neighbor now at the same index, or the new last tab if the
        // closed tab was rightmost, matching common browser tab-close
        // behavior. `null` when no tabs remain.
        activeTabId = openTabs[Math.min(index, openTabs.length - 1)] ?? null;
      }
      renderPreviewTabStrip();
      updatePreviewVisibility();
      renderList();
      const node = activeTabId ? runtime.state.graphNodes.find((item) => item["@id"] === activeTabId) : null;
      updateAnnotateButton(node || null);
    }

    async function toggleRelation(resourceId, button) {
      if (!runtime.state.datasetHandle) return;
      const isRelated = relatedResourceIds.has(resourceId);
      const action = isRelated ? "unrelate" : "relate";
      const originalLabel = button.textContent;
      button.disabled = true;
      button.textContent = isRelated ? t("unrelating") : t("relating");
      try {
        const result = await runLinksetOperation("related", action, dimensionUri, resourceId);
        relatedResourceIds = new Set(result.entries[dimensionUri] || []);
        _dedupSuggestedVsRelated();
        renderResources();
      } catch (error) {
        console.error(error);
        button.textContent = t("error");
        setTimeout(() => { button.textContent = originalLabel; }, 2500);
      } finally {
        button.disabled = !runtime.state.datasetHandle;
      }
    }

    async function toggleSuggestion(resourceId, button) {
      if (!runtime.state.datasetHandle) return;
      const isSuggested = suggestedResourceIds.has(resourceId);
      const action = isSuggested ? "unsuggest" : "suggest";
      const originalLabel = button.textContent;
      button.disabled = true;
      button.textContent = isSuggested ? t("unsuggesting") : t("suggesting");
      try {
        const result = await runLinksetOperation("suggested", action, dimensionUri, resourceId);
        suggestedResourceIds = new Set(result.entries[dimensionUri] || []);
        _dedupSuggestedVsRelated();
        renderResources();
      } catch (error) {
        console.error(error);
        button.textContent = t("error");
        setTimeout(() => { button.textContent = originalLabel; }, 2500);
      } finally {
        button.disabled = !runtime.state.datasetHandle;
      }
    }

    // Same shape/traversal as onto-file-tree-tile.js's #buildTree/#renderNode
    // (the main Surface's file tree) — folders group by filePath segments,
    // sorted directories-first, so a resource's location reads the same way
    // here as it does there. Leaf nodes carry the actual JSON-LD entity.
    function buildResourceTree(nodes) {
      const root = { children: new Map(), isFile: false, path: "" };
      for (const resourceNode of nodes) {
        const filePath = literal(resourceNode, `${OBDC_NS}filePath`);
        if (!filePath) continue;
        const segments = filePath.split("/").filter(Boolean);
        let current = root;
        let builtPath = "";
        segments.forEach((segment, index) => {
          builtPath = builtPath ? `${builtPath}/${segment}` : segment;
          if (!current.children.has(segment)) {
            current.children.set(segment, { children: new Map(), isFile: true, path: builtPath, node: null });
          }
          current = current.children.get(segment);
          const isLeaf = index === segments.length - 1;
          current.isFile = isLeaf;
          if (isLeaf) current.node = resourceNode;
        });
      }
      return root;
    }

    function renderTreeNode(treeNode) {
      const list = document.createElement("ul");
      list.className = "resource-tree-list";
      const entries = [...treeNode.children.entries()].sort((a, b) => {
        if (a[1].isFile !== b[1].isFile) return a[1].isFile ? 1 : -1;
        return a[0].localeCompare(b[0]);
      });

      for (const [name, child] of entries) {
        const item = document.createElement("li");
        const row = document.createElement("div");
        const expanded = expandedPaths.has(child.path);
        row.className = `resource-node ${child.isFile ? "file" : "dir"}`;
        if (child.isFile && child.node && child.node["@id"] === activeTabId) {
          row.classList.add("is-selected");
        }

        const icon = document.createElement("span");
        icon.className = "resource-node-icon";
        icon.setAttribute("aria-hidden", "true");
        icon.innerHTML = child.isFile
          ? RESOURCE_NODE_ICONS.file
          : expanded
            ? RESOURCE_NODE_ICONS.folderOpen
            : RESOURCE_NODE_ICONS.folder;
        const nameEl = document.createElement("span");
        nameEl.className = "resource-node-name";
        nameEl.textContent = name;
        nameEl.title = name;
        row.append(icon, nameEl);

        if (child.isFile) {
          const nodeId = child.node["@id"];

          if (!relatedResourceIds.has(nodeId)) {
            const suggestBtn = document.createElement("button");
            suggestBtn.type = "button";
            suggestBtn.className = "resource-suggest-btn";
            suggestBtn.textContent = suggestedResourceIds.has(nodeId) ? t("unsuggest") : t("suggest");
            suggestBtn.disabled = !runtime.state.datasetHandle;
            suggestBtn.title = runtime.state.datasetHandle ? "" : t("connectFolderFirst");
            suggestBtn.addEventListener("click", (event) => {
              event.stopPropagation();
              toggleSuggestion(nodeId, suggestBtn);
            });
            row.appendChild(suggestBtn);
          }

          const relateBtn = document.createElement("button");
          relateBtn.type = "button";
          relateBtn.className = "resource-relate-btn";
          relateBtn.textContent = relatedResourceIds.has(nodeId) ? t("unrelate") : t("relate");
          relateBtn.disabled = !runtime.state.datasetHandle;
          relateBtn.title = runtime.state.datasetHandle ? "" : t("connectFolderFirst");
          relateBtn.addEventListener("click", (event) => {
            event.stopPropagation();
            toggleRelation(nodeId, relateBtn);
          });
          row.appendChild(relateBtn);

          row.addEventListener("click", () => {
            openResourceTab(child.node);
          });
        } else {
          row.addEventListener("click", () => {
            if (expandedPaths.has(child.path)) expandedPaths.delete(child.path);
            else expandedPaths.add(child.path);
            renderList();
          });
        }

        item.appendChild(row);
        if (!child.isFile && child.children.size && expanded) {
          item.appendChild(renderTreeNode(child));
        }
        list.appendChild(item);
      }

      return list;
    }

    function renderList() {
      const treeEl = card.querySelector(".resource-tree-container");
      if (!treeEl) return;
      treeEl.innerHTML = "";
      const nodes = currentTabNodes();
      if (!nodes.length) return;
      treeEl.appendChild(renderTreeNode(buildResourceTree(nodes)));
    }

    function renderResources() {
      card.querySelectorAll(".resource-tab").forEach((tab) => {
        const isActive = tab.dataset.tab === activeTab;
        tab.classList.toggle("is-active", isActive);
        tab.setAttribute("aria-selected", String(isActive));
      });
      renderCategoryTiles();
      // Neither the Related/Suggested/Found filter nor the category
      // tiles may touch which preview tabs are currently open.
      renderList();
    }

    function closeAnnotationsTile() {
      if (!annotationsTile) return;
      annotationsTile.remove();
      annotationsTile = null;
      annotationsTileResourceId = null;
    }

    function annotationCategoryLabel(annotation) {
      return OntoBDCAnnotationModel.localName(annotation.type).replace("Annotation", "") || t("note");
    }

    function formatTimestamp(iso) {
      if (!iso) return "";
      try {
        return new Date(iso).toLocaleString();
      } catch {
        return iso;
      }
    }

    // The runtime never exposes its internal store.load() directly — only
    // methods that also build UI around it (open/openWorkspace/...). Reuses
    // openWorkspace into a throwaway, never-attached element purely for its
    // store.load() side effect, then reads the same shared store back out
    // via listAnnotations() for this tile's own rendering.
    async function loadResourceAnnotations(node) {
      const instance = ensureAnnotationRuntime();
      if (!instance || !runtime || !runtime.state || !runtime.state.datasetHandle || !node) return [];
      const scratch = document.createElement("div");
      await instance.openWorkspace(scratch, { containerHandle: runtime.state.datasetHandle });
      const context = { logicalSource: node["@id"], representationSource: node["@id"] };
      return (typeof runtime.listAnnotations === "function" ? runtime.listAnnotations() : [])
        .filter((annotation) => OntoBDCAnnotationModel.matchesContext(annotation, context));
    }

    function buildAnnotationsTile(node, annotations) {
      const tile = document.createElement("section");
      tile.className = "tile annotations-tile";
      tile.innerHTML = `
        <div class="annotations-tile-head">
          <span class="dimension-label">${t("annotations")} · ${resourceLabel(node)}</span>
          <button type="button" class="icon-btn close-annotations-btn" title="${t("closeAnnotations")}" aria-label="${t("closeAnnotations")}">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>
          </button>
        </div>
        <div class="annotations-tile-list"></div>
      `;
      tile.querySelector(".close-annotations-btn").addEventListener("click", closeAnnotationsTile);

      const listEl = tile.querySelector(".annotations-tile-list");
      if (!annotations.length) {
        const empty = document.createElement("div");
        empty.className = "annotations-tile-empty";
        empty.textContent = t("noAnnotationsYet");
        listEl.appendChild(empty);
      }
      for (const annotation of annotations) {
        const row = document.createElement("div");
        row.className = "annotations-tile-row";

        const summary = document.createElement("div");
        summary.className = "annotations-tile-summary";
        const badge = document.createElement("span");
        badge.className = "annotations-tile-badge";
        badge.textContent = annotationCategoryLabel(annotation);
        const body = document.createElement("span");
        body.className = "annotations-tile-body-preview";
        body.textContent = annotation.body || "(no text)";
        summary.append(badge, body);

        const detail = document.createElement("div");
        detail.className = "annotations-tile-detail";
        detail.hidden = true;
        const bodyFull = document.createElement("div");
        bodyFull.className = "field-value";
        bodyFull.textContent = annotation.body || "(no text)";
        const meta = document.createElement("div");
        meta.className = "annotations-tile-meta";
        meta.textContent = [
          formatTimestamp(annotation.modified || annotation.created),
          annotation.properties && annotation.properties.markerColor ? annotation.properties.markerColor : null,
        ].filter(Boolean).join(" · ");
        detail.append(bodyFull, meta);

        summary.addEventListener("click", () => { detail.hidden = !detail.hidden; });
        row.append(summary, detail);
        listEl.appendChild(row);
      }
      return tile;
    }

    async function toggleAnnotationsTile(node, button) {
      if (annotationsTile) {
        closeAnnotationsTile();
        return;
      }
      const originalLabel = button.innerHTML;
      button.disabled = true;
      try {
        const annotations = await loadResourceAnnotations(node);
        annotationsTile = buildAnnotationsTile(node, annotations);
        annotationsTileResourceId = node["@id"];
        card.insertAdjacentElement("afterend", annotationsTile);
        annotationsTile.scrollIntoView({ behavior: "smooth", block: "nearest" });
      } catch (error) {
        console.error(error);
      } finally {
        button.disabled = !(ensureAnnotationRuntime() && runtime && runtime.state && runtime.state.datasetHandle);
        button.innerHTML = originalLabel;
      }
    }

    card.querySelectorAll(".resource-tab").forEach((tab) => {
      tab.addEventListener("click", () => {
        activeTab = tab.dataset.tab;
        renderResources();
      });
    });
    card.querySelector(".create-annotation-btn").addEventListener("click", (event) => {
      const node = runtime.state.graphNodes.find((item) => item["@id"] === activeTabId);
      if (node) annotateResource(node, dimensionUri, event.currentTarget);
    });
    card.querySelector(".view-annotations-btn").addEventListener("click", (event) => {
      const node = runtime.state.graphNodes.find((item) => item["@id"] === activeTabId);
      if (node) toggleAnnotationsTile(node, event.currentTarget);
    });

    renderResources();
    updateAnnotateButton(null);

    return {
      element: card,
      async onConnected() {
        connected = true;
        renderResources();
        const [relatedResult, suggestedResult, displayResult] = await Promise.all([
          loadAllLinks("related"),
          loadAllLinks("suggested"),
          loadFileDisplayProfiles(),
        ]);
        relatedResourceIds = new Set(relatedResult.entries[dimensionUri] || []);
        suggestedResourceIds = new Set(suggestedResult.entries[dimensionUri] || []);
        _dedupSuggestedVsRelated();
        displayProfiles = displayResult.profiles || [];
        if (activeCategory && displayProfiles.every((p) => p.id !== activeCategory)) {
          activeCategory = null;
        }
        renderResources();
      },
    };
  }


  function renderDimensionCards() {
    const container = document.querySelector(".dimension-cards");
    container.innerHTML = "";
    dimensionCards.length = 0;
    for (const dimension of DIMENSIONS) {
      const value = literal(runtime.state.selfNode, dimension.property);
      const card = createDimensionCard(dimension, value);
      dimensionCards.push(card);
      container.appendChild(card.element);
    }
  }

  function notifyDimensionCardsConnected() {
    for (const card of dimensionCards) {
      if (card.onConnected) card.onConnected();
    }
  }

  // Each setup step below is independent: a resource tree failing to build
  // must not stop Relate/Suggest from wiring up, and vice versa. Before this
  // guard, render() was one unbroken call chain — any single step throwing
  // (sync or via an unguarded promise elsewhere reacting to this render)
  // silently skipped every step after it, so a single unrelated failure
  // could look like "nothing on the page works" instead of what it was.
  //
  // `step` may be async (e.g. tryReconnectSilently touches IndexedDB and
  // the File System Access API). A plain try/catch here only guards the
  // *synchronous* prefix up to `step`'s first `await` -- anything that
  // rejects after that runs outside this try block entirely and becomes an
  // unhandled promise rejection instead of a caught, labeled error. Await
  // the result and catch on it too so async steps get the same isolation
  // sync ones do.
  function runRenderStep(label, step) {
    try {
      const result = step();
      if (result && typeof result.catch === "function") {
        result.catch((error) => {
          console.error(`[work_stream_view] "${label}" step failed:`, error);
        });
      }
    } catch (error) {
      console.error(`[work_stream_view] "${label}" step failed:`, error);
    }
  }

  function render() {
    runRenderStep("applyStaticI18n", applyStaticI18n);

    runtime.state.graphNodes = loadGraph();
    const resourceId = document.querySelector(".onto-page")?.getAttribute("data-ontobdc-resource") || "";
    runtime.state.selfNode = runtime.state.graphNodes.find((node) => node && node["@id"] === resourceId) || null;

    let identifier = "";
    runRenderStep("renderHeader", () => { identifier = renderHeader(); });

    runRenderStep("workStreamContext", () => {
      if (typeof OntoBDCWorkStreamContext !== "undefined" && runtime.state.selfNode) {
        workStreamContext = OntoBDCWorkStreamContext.create({
          workStreamId: identifier || runtime.state.selfNode["@id"],
          workStreamUri: runtime.state.selfNode["@id"],
        });
      }
    });

    runRenderStep("renderDimensionCards", renderDimensionCards);
    runRenderStep("wireAnnotationControls", wireAnnotationControls);

    if (typeof OntoBDCAnnotations !== "undefined") {
      runRenderStep("ensureAnnotationRuntime", ensureAnnotationRuntime);
      runRenderStep("tryReconnectSilently", tryReconnectSilently);
    } else {
      const connectBtn = document.querySelector(".connect-btn");
      if (connectBtn) connectBtn.hidden = true;
    }
  }

  if (document.readyState === "loading") {
    document.addEventListener("DOMContentLoaded", render);
  } else {
    render();
  }

  // No language tile lives on this standalone Page today, but re-applying
  // static chrome + dimension labels on the event keeps this Page in sync
  // for free if one is ever added, or if `document.documentElement.lang`
  // is changed by anything else.
  document.addEventListener("language-changed", () => {
    applyStaticI18n();
    renderDimensionCards();
  });


  Object.assign(runtime, {
    createDimensionCard: createDimensionCard,
    renderDimensionCards: renderDimensionCards,
    notifyDimensionCardsConnected: notifyDimensionCardsConnected,
    runRenderStep: runRenderStep,
    render: render
  });
  console.log("[work-stream-view] state end: dimension_card_script_generated");
}(window));
"""

    _BUILDERS: Dict[str, Callable[["WorkStreamScriptAdapter"], str]] = {
        # Same vendored SheetJS build the Gantt Page loads, resolved the
        # same way: this adapter's own name list already includes it.
        VENDOR_SHEET_JS_NAME: lambda self: vendor_asset_source(
            VENDOR_SHEET_JS_NAME
        ),
        "i18n_apply": _i18n_apply_source,
        "graph_reader": _graph_reader_source,
        "csv_preview": _csv_preview_source,
        "container_connection": _container_connection_source,
        "connection_state": _connection_state_source,
        "chrome_controls": lambda self: chrome_controls_source(WORK_STREAM_RUNTIME),
        "annotation_bridge": _annotation_bridge_source,
        "pyodide_runtime": _pyodide_runtime_source,
        "linkset_operations": _linkset_operations_source,
        "file_category": _file_category_source,
        "dimension_card": _dimension_card_source,
    }
